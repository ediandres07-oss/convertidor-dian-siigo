"""
Liquidador de Formulario 210 - Declaración de Renta
Mapea datos de Exógena y calcula liquidaciones tributarias colombianas
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
from pathlib import Path

class Formulario210Liquidador:
    """Liquidador completo del Formulario 210"""

    # UVT 2025 (Unidad de Valor Tributario)
    UVT_2025 = 49799
    UVT_2024 = 44654
    UVT_2022 = 38004  # Para año gravable 2022

    # Rangos IRPF 2024
    RANGOS_IRPF = [
        (0, 66950000, 0),
        (66950000, 134900000, 0.05),
        (134900000, 404700000, 0.12),
        (404700000, 673500000, 0.25),
        (673500000, 1347000000, 0.32),
        (1347000000, float('inf'), 0.37),
    ]

    # Mapping de campos Exógena a Formulario 210
    MAPEO_EXOGENA = {
        'Honorarios y Comisiones': 200,
        'Rendimientos Financieros': 300,
        'Ingresos por Arrendamiento': 400,
        'Ganancia en Venta de Bienes': 500,
        'Otros Ingresos': 600,
        'Retención en la Fuente': 110,  # Descuento
    }

    def __init__(self, ano_gravable=2024):
        self.ano_gravable = ano_gravable
        if ano_gravable == 2022:
            self.uvt = self.UVT_2022
        elif ano_gravable >= 2024:
            self.uvt = self.UVT_2025
        else:
            self.uvt = self.UVT_2024
        self.datos = {
            'contribuyente': {},
            'ingresos': {},
            'retenciones': {},
            'calculos': {}
        }

    def leer_exogena(self, ruta_exogena):
        """Lee el archivo de Exógena reportada por terceros"""
        try:
            df = pd.read_excel(ruta_exogena, sheet_name='Reporte', header=None)

            # Extraer datos del titular (filas 5-8)
            datos_person = {}
            for i in range(len(df)):
                row = df.iloc[i]
                if pd.notna(row.iloc[0]):
                    contenido = str(row.iloc[0]).lower()
                    if 'identificación:' in contenido and pd.notna(row.iloc[2]):
                        datos_person['cedula'] = str(row.iloc[2])
                    elif 'nombres' in contenido and pd.notna(row.iloc[2]):
                        datos_person['nombre'] = str(row.iloc[2])
                    elif 'año' in contenido and 'refiere' in contenido and pd.notna(row.iloc[2]):
                        datos_person['ano'] = int(row.iloc[2])

            # Tabla de ingresos comienza en fila 15 (índice 14)
            # Columnas: [Nit reporta, Nombre reporta, Nit contrib, Nombre contrib, Detalle, Valor, Uso, Info adicional]
            self.datos['ingresos'] = {}
            self.datos['retenciones'] = {}

            for i in range(14, len(df)):
                row = df.iloc[i]
                if pd.notna(row.iloc[4]) and pd.notna(row.iloc[5]):
                    concepto = str(row.iloc[4]).strip()
                    try:
                        valor = float(row.iloc[5])
                        if valor > 0:
                            # Clasificar como ingreso o retención
                            uso = str(row.iloc[6]).lower() if pd.notna(row.iloc[6]) else ""
                            if 'retención' in uso or 'r132' in uso or 'r28' in uso:
                                self._agregar_a_dict(self.datos['retenciones'], concepto, valor)
                            else:
                                self._agregar_a_dict(self.datos['ingresos'], concepto, valor)
                    except (ValueError, TypeError):
                        pass

            self.datos['contribuyente'] = datos_person
            return True

        except Exception as e:
            print(f"Error leyendo Exógena: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _agregar_a_dict(self, diccionario, clave, valor):
        """Suma valores si la clave ya existe"""
        if clave in diccionario:
            diccionario[clave] += valor
        else:
            diccionario[clave] = valor

    def leer_formulario_210(self, ruta_210):
        """Lee el Formulario 210 base"""
        try:
            # Intenta con openpyxl
            wb = load_workbook(ruta_210)
            self.wb_original = wb
            return True
        except:
            # Si falla, intentará recrear desde cero
            self.wb_original = None
            return False

    def calcular_irpf(self, base_liquida):
        """Calcula el IRPF según rangos colombianos"""
        impuesto = 0
        for inicio, fin, tasa in self.RANGOS_IRPF:
            if base_liquida <= inicio:
                break
            renta_tramo = min(base_liquida, fin) - inicio
            if renta_tramo > 0:
                impuesto += renta_tramo * tasa
        return max(0, impuesto)

    def calcular_aporte_solidario(self, base_liquida):
        """Aporte solidario para rentas altas"""
        if base_liquida > 16000 * self.uvt:  # 16.000 UVT
            return (base_liquida - 16000 * self.uvt) * 0.01
        return 0

    def liquidar(self):
        """Realiza la liquidación completa"""

        # 1. Sumar ingresos por tipo
        total_ingresos = sum(self.datos['ingresos'].values())

        # 2. Calcular deducciones (10% de ingresos laborales o gastos comprobados)
        deducciones = total_ingresos * 0.10  # Aproximado, se ajusta con datos reales

        # 3. Base renta trabajo
        base_trabajo = max(0, total_ingresos - deducciones)

        # 4. IRPF
        irpf = self.calcular_irpf(base_trabajo)

        # 5. Aporte solidario
        aporte_solidario = self.calcular_aporte_solidario(base_trabajo)

        # 6. Total impuesto
        total_impuesto = irpf + aporte_solidario

        # 7. Deducción de retenciones
        retenciones_totales = self.datos.get('retenciones', {})
        retenciones_valor = sum(retenciones_totales.values()) if isinstance(retenciones_totales, dict) else retenciones_totales

        # 8. Saldo a pagar o acreencia
        saldo = max(0, total_impuesto - retenciones_valor)
        acreencia = max(0, retenciones_valor - total_impuesto)

        self.datos['calculos'] = {
            'total_ingresos': total_ingresos,
            'deducciones': deducciones,
            'base_liquida': base_trabajo,
            'irpf': irpf,
            'aporte_solidario': aporte_solidario,
            'total_impuesto': total_impuesto,
            'retenciones': retenciones_valor,
            'saldo_pagar': saldo,
            'acreencia': acreencia,
        }

        return self.datos['calculos']

    def generar_excel(self, ruta_salida):
        """Genera Excel con el Formulario 210 completo"""

        if not self.datos['calculos']:
            self.liquidar()

        wb = Workbook()
        ws = wb.active
        ws.title = "Formulario 210"

        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        titulo_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        titulo_font = Font(bold=True, size=11)
        currency_format = '#,##0;-#,##0'

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Encabezado
        ws.merge_cells('A1:D1')
        ws['A1'] = "FORMULARIO 210 - DECLARACIÓN DE RENTA"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:D2')
        ws['A2'] = f"Año Gravable: {self.ano_gravable}"
        ws['A2'].alignment = Alignment(horizontal='center')

        # Datos del contribuyente
        row = 4
        ws[f'A{row}'] = "IDENTIFICACIÓN DEL CONTRIBUYENTE"
        ws[f'A{row}'].font = titulo_font
        ws[f'A{row}'].fill = titulo_fill

        row += 1
        ws[f'A{row}'] = "Cédula:"
        ws[f'B{row}'] = self.datos['contribuyente'].get('cedula', '')

        row += 1
        ws[f'A{row}'] = "Nombre:"
        ws[f'B{row}'] = self.datos['contribuyente'].get('nombre', '')

        # Ingresos
        row += 2
        ws[f'A{row}'] = "INGRESOS BRUTOS"
        ws[f'A{row}'].font = titulo_font
        ws[f'A{row}'].fill = titulo_fill

        row += 1
        ws[f'A{row}'] = "Concepto"
        ws[f'B{row}'] = "Valor"
        for cell in [ws[f'A{row}'], ws[f'B{row}']]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.font = header_font

        for concepto, valor in self.datos['ingresos'].items():
            row += 1
            ws[f'A{row}'] = concepto
            ws[f'B{row}'] = valor
            ws[f'B{row}'].number_format = currency_format

        row += 1
        ws[f'A{row}'] = "TOTAL INGRESOS"
        ws[f'B{row}'] = f"={chr(66)}{row-len(self.datos['ingresos'])}:{chr(66)}{row-1}"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = currency_format

        # Liquidación
        row += 2
        ws[f'A{row}'] = "LIQUIDACIÓN DEL IMPUESTO"
        ws[f'A{row}'].font = titulo_font
        ws[f'A{row}'].fill = titulo_fill

        calculos = self.datos['calculos']

        liquidacion_items = [
            ("Total Ingresos", calculos['total_ingresos']),
            ("Deducciones", calculos['deducciones']),
            ("Base Líquida", calculos['base_liquida']),
            ("IRPF Calculado", calculos['irpf']),
            ("Aporte Solidario", calculos['aporte_solidario']),
            ("Total Impuesto", calculos['total_impuesto']),
            ("Retenciones", calculos['retenciones']),
            ("Saldo a Pagar", calculos['saldo_pagar']),
            ("Acreencia", calculos['acreencia']),
        ]

        for concepto, valor in liquidacion_items:
            row += 1
            ws[f'A{row}'] = concepto
            ws[f'B{row}'] = valor
            ws[f'B{row}'].number_format = currency_format
            if "Total" in concepto or "Saldo" in concepto or "Acreencia" in concepto:
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18

        wb.save(ruta_salida)
        return ruta_salida

def procesar_liquidacion(ruta_exogena, ruta_formulario_base, ano_gravable=2024):
    """Función principal para procesar la liquidación completa"""

    liquidador = Formulario210Liquidador(ano_gravable)

    # Leer datos
    if not liquidador.leer_exogena(ruta_exogena):
        return {'error': 'No se pudo leer archivo de Exógena'}

    liquidador.leer_formulario_210(ruta_formulario_base)

    # Calcular
    resultado = liquidador.liquidar()

    # Generar Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta_salida = f"formulario_210_liquidado_{timestamp}.xlsx"
    liquidador.generar_excel(ruta_salida)

    return {
        'exito': True,
        'contribuyente': liquidador.datos['contribuyente'],
        'calculos': resultado,
        'archivo': ruta_salida
    }

if __name__ == "__main__":
    # Prueba
    resultado = procesar_liquidacion(
        "/Users/edison/Downloads/reporteExogena2024Elizabeth.xlsx",
        "/Users/edison/Downloads/VA23-Formulario-210-AG2022-PN-SIMONVELASQUEZ.xlsx",
        2024
    )
    print(json.dumps(resultado, indent=2, default=str))
