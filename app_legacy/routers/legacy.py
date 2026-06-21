"""
Router Legacy - Endpoints para app antigua
Compatibilidad con frontend/app.py
"""

from fastapi import APIRouter, File, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel
from typing import List, Optional
import pandas as pd
from app.services.liquidacion import CalculadoraLiquidacion
import os
import io
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class EmpleadoLiquidacion(BaseModel):
    documento: str
    nombre: str
    salario_mensual: float
    auxilio_transporte: float
    dias_laborados: int
    cesantias: float
    intereses: float
    prima: float
    vacaciones: float
    total_devengos: float
    total_deducciones: float
    neto_pagar: float


class RespuestaProceso(BaseModel):
    total_empleados: int
    empleados: List[EmpleadoLiquidacion]
    total_neto: float


class RespuestaPrima(BaseModel):
    total_empleados: int
    empleados: List[dict]
    total_prima: float


class RespuestaVacaciones(BaseModel):
    total_empleados: int
    empleados: List[dict]
    total_vacaciones: float


router = APIRouter(prefix="/api", tags=["legacy"])


@router.post("/procesar-completo", response_model=RespuestaProceso)
async def procesar_completo(file: UploadFile = File(...)):
    """Procesa archivo completo (compatibilidad app antigua)"""
    contents = await file.read()
    df = pd.read_excel(contents, sheet_name='Empleados')

    empleados = []
    total_neto = 0

    for idx, row in df.iterrows():
        doc = str(row.get('documento', ''))
        nombre = str(row.get('nombre', ''))
        salario = float(row.get('salario_mensual', 0) or 0)
        dias = float(row.get('dias_laborados', 30) or 30)
        auxilio = float(row.get('auxilio_transporte', 0) or 0)

        base = salario + auxilio

        cesantias = (base * dias) / 360
        intereses = (cesantias * 0.12 * dias) / 360
        prima = (base * dias) / 360
        vacaciones = (salario * dias) / 720

        total_dev = cesantias + intereses + prima + vacaciones

        empleados.append({
            'documento': doc,
            'nombre': nombre,
            'salario_mensual': salario,
            'auxilio_transporte': auxilio,
            'dias_laborados': int(dias),
            'cesantias': round(cesantias, 2),
            'intereses': round(intereses, 2),
            'prima': round(prima, 2),
            'vacaciones': round(vacaciones, 2),
            'total_devengos': round(total_dev, 2),
            'total_deducciones': 0,
            'neto_pagar': round(total_dev, 2),
        })

        total_neto += total_dev

    return {
        "total_empleados": len(empleados),
        "empleados": empleados,
        "total_neto": round(total_neto, 2)
    }


@router.post("/calcular-prima", response_model=RespuestaPrima)
async def calcular_prima(file: UploadFile = File(...)):
    """Calcula prima de servicios"""
    contents = await file.read()
    df = pd.read_excel(contents, sheet_name='Empleados')

    empleados = []
    total_prima = 0

    for idx, row in df.iterrows():
        doc = str(row.get('documento', ''))
        nombre = str(row.get('nombre', ''))
        salario = float(row.get('salario_mensual', 0) or 0)
        dias = float(row.get('dias_laborados', 30) or 30)
        auxilio = float(row.get('auxilio_transporte', 0) or 0)

        base = salario + auxilio
        valor_prima = (base * dias) / 360

        empleados.append({
            'documento': doc,
            'nombre': nombre,
            'salario_mensual': salario,
            'auxilio_transporte': auxilio,
            'dias_laborados': int(dias),
            'valor_prima': round(valor_prima, 2),
        })

        total_prima += valor_prima

    return {
        "total_empleados": len(empleados),
        "empleados": empleados,
        "total_prima": round(total_prima, 2)
    }


@router.post("/calcular-vacaciones", response_model=RespuestaVacaciones)
async def calcular_vacaciones(file: UploadFile = File(...)):
    """Calcula vacaciones proporcionales"""
    contents = await file.read()
    df = pd.read_excel(contents, sheet_name='Empleados')

    empleados = []
    total_vacaciones = 0

    for idx, row in df.iterrows():
        doc = str(row.get('documento', ''))
        nombre = str(row.get('nombre', ''))
        salario = float(row.get('salario_mensual', 0) or 0)
        dias = float(row.get('dias_laborados', 30) or 30)

        valor_vacaciones = (salario * dias) / 720

        empleados.append({
            'documento': doc,
            'nombre': nombre,
            'salario_mensual': salario,
            'dias_laborados': int(dias),
            'valor_vacaciones': round(valor_vacaciones, 2),
        })

        total_vacaciones += valor_vacaciones

    return {
        "total_empleados": len(empleados),
        "empleados": empleados,
        "total_vacaciones": round(total_vacaciones, 2)
    }


@router.post("/exportar-pdf-prima-individual")
async def exportar_pdf_prima_individual(file: UploadFile = File(...), documento: str = ""):
    """Exporta PDF individual de prima - Formato media carta profesional"""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from datetime import datetime

    contents = await file.read()
    df = pd.read_excel(contents, sheet_name='Empleados')

    # Buscar empleado
    empleado = None
    for idx, row in df.iterrows():
        if str(row.get('documento', '')) == documento:
            empleado = row
            break

    if empleado is None:
        return {"error": "Empleado no encontrado"}

    doc_num = str(empleado.get('documento', ''))
    nombre = str(empleado.get('nombre', ''))
    empresa = str(empleado.get('empresa', 'EMPRESA S.A.S'))
    salario = float(empleado.get('salario_mensual', 0) or 0)
    dias = float(empleado.get('dias_laborados', 30) or 30)
    auxilio = float(empleado.get('auxilio_transporte', 0) or 0)

    base = salario + auxilio
    valor_prima = (base * dias) / 360

    # Crear PDF - MEDIA CARTA
    pdf_buffer = io.BytesIO()
    doc_pdf = SimpleDocTemplate(pdf_buffer, pagesize=letter,
                               topMargin=0.4*inch, bottomMargin=0.4*inch,
                               rightMargin=0.4*inch, leftMargin=0.4*inch)
    styles = getSampleStyleSheet()
    elements = []

    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f4788'),
        alignment=TA_CENTER,
        spaceAfter=12
    )
    elements.append(Paragraph("LIQUIDACIÓN PRIMA DE SERVICIOS", title_style))

    # Empresa
    empresa_style = ParagraphStyle(
        'EmpresaStyle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f4788'),
        alignment=TA_CENTER,
        spaceAfter=12
    )
    elements.append(Paragraph(empresa, empresa_style))

    # SECCIÓN 1: DATOS DEL EMPLEADO
    empleado_header = [['DATOS DEL EMPLEADO']]
    empleado_header_table = Table(empleado_header, colWidths=[6*inch])
    empleado_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(empleado_header_table)

    # Datos empleado
    datos_emp = [
        ['Nombre:', nombre],
        ['Documento:', doc_num],
        ['Salario Mensual:', f"${salario:,.2f}"],
        ['Auxilio Transporte:', f"${auxilio:,.2f}"],
        ['Días Laborados:', f"{int(dias)}"],
    ]

    datos_emp_table = Table(datos_emp, colWidths=[2*inch, 4*inch])
    datos_emp_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))
    elements.append(datos_emp_table)
    elements.append(Spacer(1, 0.2 * inch))

    # SECCIÓN 2: CONCEPTO Y VALOR
    concepto_header = [['CONCEPTO', 'VALOR']]
    concepto_header_table = Table(concepto_header, colWidths=[4*inch, 2*inch])
    concepto_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(concepto_header_table)

    # Conceptos
    conceptos = [
        ['Salario + Auxilio:', f"${base:,.2f}"],
        ['Días Laborados:', f"{int(dias)}"],
        ['PRIMA DE SERVICIOS', f"${valor_prima:,.2f}"],
    ]

    conceptos_table = Table(conceptos, colWidths=[4*inch, 2*inch])
    conceptos_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#B8D4E8')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
    ]))
    elements.append(conceptos_table)
    elements.append(Spacer(1, 0.15 * inch))

    # Fórmula
    elements.append(Paragraph(
        f"<font size=8>Prima = (Salario + Auxilio) × Días ÷ 360</font>",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.08 * inch))

    # Fecha de generación
    elements.append(Paragraph(
        f"<font size=8>Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</font>",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.3 * inch))

    # SECCIÓN 3: FIRMAS
    firmas_header = [['FIRMAS']]
    firmas_header_table = Table(firmas_header, colWidths=[6*inch])
    firmas_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(firmas_header_table)

    # Firmas
    firmas = [
        ['_' * 35, '', '_' * 35],
        ['Firma Empleado', '', 'Firma Empleador'],
        ['Nombre: _________________', '', 'Nombre: _________________'],
        ['Cédula: _________________', '', 'Cédula: _________________'],
        ['Fecha: _________________', '', 'Fecha: _________________'],
    ]

    firmas_table = Table(firmas, colWidths=[2*inch, 0.3*inch, 2*inch])
    firmas_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(firmas_table)

    doc_pdf.build(elements)

    pdf_buffer.seek(0)
    return StreamingResponse(
        iter([pdf_buffer.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=prima_{nombre.replace(' ', '_')}.pdf"}
    )


@router.post("/exportar-excel-prima")
async def exportar_excel_prima(file: UploadFile = File(...)):
    """Exporta prima a Excel"""
    contents = await file.read()
    df = pd.read_excel(contents, sheet_name='Empleados')

    empleados = []
    for idx, row in df.iterrows():
        doc = str(row.get('documento', ''))
        nombre = str(row.get('nombre', ''))
        salario = float(row.get('salario_mensual', 0) or 0)
        dias = float(row.get('dias_laborados', 30) or 30)
        auxilio = float(row.get('auxilio_transporte', 0) or 0)

        base = salario + auxilio
        valor_prima = (base * dias) / 360

        empleados.append({
            'Documento': doc,
            'Nombre': nombre,
            'Salario': salario,
            'Auxilio': auxilio,
            'Días': int(dias),
            'Prima': round(valor_prima, 2),
        })

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Prima"

    # Headers
    headers = ['Documento', 'Nombre', 'Salario', 'Auxilio', 'Días', 'Prima']
    ws.append(headers)

    # Formato headers
    header_fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Datos
    for emp in empleados:
        ws.append([emp['Documento'], emp['Nombre'], emp['Salario'], emp['Auxilio'], emp['Días'], emp['Prima']])

    # Total
    total_prima = sum(e['Prima'] for e in empleados)
    ws.append(['', '', '', '', 'TOTAL', total_prima])

    # Formato total
    last_row = ws.max_row
    for col in range(1, 7):
        cell = ws.cell(last_row, col)
        cell.font = Font(bold=True)
        if col == 6:
            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

    # Ancho columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15

    # Guardar en memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=prima_servicios.xlsx"}
    )


@router.post("/exportar-pdf-prima-zip")
async def exportar_pdf_prima_zip(file: UploadFile = File(...)):
    """Exporta PDFs de prima en ZIP - Formato media carta profesional"""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from datetime import datetime

    contents = await file.read()
    df = pd.read_excel(contents, sheet_name='Empleados')

    # Crear ZIP en memoria
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        for idx, row in df.iterrows():
            doc = str(row.get('documento', ''))
            nombre = str(row.get('nombre', ''))
            empresa = str(row.get('empresa', 'EMPRESA S.A.S'))
            salario = float(row.get('salario_mensual', 0) or 0)
            dias = float(row.get('dias_laborados', 30) or 30)
            auxilio = float(row.get('auxilio_transporte', 0) or 0)

            base = salario + auxilio
            valor_prima = (base * dias) / 360

            # Crear PDF en memoria - MEDIA CARTA
            pdf_buffer = io.BytesIO()
            doc_pdf = SimpleDocTemplate(pdf_buffer, pagesize=letter,
                                       topMargin=0.4*inch, bottomMargin=0.4*inch,
                                       rightMargin=0.4*inch, leftMargin=0.4*inch)
            styles = getSampleStyleSheet()
            elements = []

            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1f4788'),
                alignment=TA_CENTER,
                spaceAfter=12
            )
            elements.append(Paragraph("LIQUIDACIÓN PRIMA DE SERVICIOS", title_style))

            # Empresa
            empresa_style = ParagraphStyle(
                'EmpresaStyle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=TA_CENTER,
                spaceAfter=8
            )
            elements.append(Paragraph(f"<b>{empresa}</b>", empresa_style))

            # SECCIÓN 1: DATOS DEL EMPLEADO
            empleado_header = [['DATOS DEL EMPLEADO']]
            empleado_header_table = Table(empleado_header, colWidths=[6*inch])
            empleado_header_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(empleado_header_table)

            # Datos empleado
            datos_emp = [
                ['Nombre:', nombre],
                ['Documento:', doc],
                ['Salario Mensual:', f"${salario:,.2f}"],
                ['Auxilio Transporte:', f"${auxilio:,.2f}"],
                ['Días Laborados:', f"{int(dias)}"],
            ]

            datos_emp_table = Table(datos_emp, colWidths=[2*inch, 4*inch])
            datos_emp_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ]))
            elements.append(datos_emp_table)
            elements.append(Spacer(1, 0.2 * inch))

            # SECCIÓN 2: CONCEPTO Y VALOR
            concepto_header = [['CONCEPTO', 'VALOR']]
            concepto_header_table = Table(concepto_header, colWidths=[4*inch, 2*inch])
            concepto_header_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(concepto_header_table)

            # Conceptos
            conceptos = [
                ['Salario + Auxilio:', f"${base:,.2f}"],
                ['Días Laborados:', f"{int(dias)}"],
                ['PRIMA DE SERVICIOS', f"${valor_prima:,.2f}"],
            ]

            conceptos_table = Table(conceptos, colWidths=[4*inch, 2*inch])
            conceptos_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#B8D4E8')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 11),
            ]))
            elements.append(conceptos_table)
            elements.append(Spacer(1, 0.15 * inch))

            # Fórmula
            elements.append(Paragraph(
                f"<font size=8>Prima = (Salario + Auxilio) × Días ÷ 360</font>",
                styles['Normal']
            ))
            elements.append(Spacer(1, 0.08 * inch))

            # Fecha de generación
            elements.append(Paragraph(
                f"<font size=8>Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</font>",
                styles['Normal']
            ))
            elements.append(Spacer(1, 0.3 * inch))

            # SECCIÓN 3: FIRMAS
            firmas_header = [['FIRMAS']]
            firmas_header_table = Table(firmas_header, colWidths=[6*inch])
            firmas_header_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(firmas_header_table)

            # Firmas
            firmas = [
                ['_' * 35, '', '_' * 35],
                ['Firma Empleado', '', 'Firma Empleador'],
                ['Nombre: _________________', '', 'Nombre: _________________'],
                ['Cédula: _________________', '', 'Cédula: _________________'],
                ['Fecha: _________________', '', 'Fecha: _________________'],
            ]

            firmas_table = Table(firmas, colWidths=[2*inch, 0.3*inch, 2*inch])
            firmas_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(firmas_table)

            doc_pdf.build(elements)

            # Agregar PDF al ZIP
            pdf_buffer.seek(0)
            filename = f"Prima_{nombre.replace(' ', '_')}.pdf"
            zip_file.writestr(filename, pdf_buffer.getvalue())

    zip_buffer.seek(0)
    return StreamingResponse(
        iter([zip_buffer.getvalue()]),
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=prima_servicios.zip"}
    )


@router.post("/exportar-excel-vacaciones")
async def exportar_excel_vacaciones(file: UploadFile = File(...)):
    """Exporta vacaciones a Excel"""
    contents = await file.read()
    df = pd.read_excel(contents, sheet_name='Empleados')

    empleados = []
    for idx, row in df.iterrows():
        doc = str(row.get('documento', ''))
        nombre = str(row.get('nombre', ''))
        salario = float(row.get('salario_mensual', 0) or 0)
        dias = float(row.get('dias_laborados', 30) or 30)

        valor_vacaciones = (salario * dias) / 720

        empleados.append({
            'Documento': doc,
            'Nombre': nombre,
            'Salario': salario,
            'Días': int(dias),
            'Vacaciones': round(valor_vacaciones, 2),
        })

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Vacaciones"

    # Headers
    headers = ['Documento', 'Nombre', 'Salario', 'Días', 'Vacaciones']
    ws.append(headers)

    # Formato headers
    header_fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Datos
    for emp in empleados:
        ws.append([emp['Documento'], emp['Nombre'], emp['Salario'], emp['Días'], emp['Vacaciones']])

    # Total
    total_vacaciones = sum(e['Vacaciones'] for e in empleados)
    ws.append(['', '', '', 'TOTAL', total_vacaciones])

    # Formato total
    last_row = ws.max_row
    for col in range(1, 6):
        cell = ws.cell(last_row, col)
        cell.font = Font(bold=True)
        if col == 5:
            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

    # Ancho columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15

    # Guardar en memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=vacaciones_proporcionales.xlsx"}
    )


@router.post("/exportar-pdf-vacaciones-zip")
async def exportar_pdf_vacaciones_zip(file: UploadFile = File(...)):
    """Exporta PDFs de vacaciones en ZIP - Formato media carta profesional"""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from datetime import datetime

    contents = await file.read()
    df = pd.read_excel(contents, sheet_name='Empleados')

    # Crear ZIP en memoria
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        for idx, row in df.iterrows():
            doc = str(row.get('documento', ''))
            nombre = str(row.get('nombre', ''))
            empresa = str(row.get('empresa', 'EMPRESA S.A.S'))
            salario = float(row.get('salario_mensual', 0) or 0)
            dias = float(row.get('dias_laborados', 30) or 30)

            valor_vacaciones = (salario * dias) / 720

            # Crear PDF en memoria - MEDIA CARTA
            pdf_buffer = io.BytesIO()
            doc_pdf = SimpleDocTemplate(pdf_buffer, pagesize=letter,
                                       topMargin=0.4*inch, bottomMargin=0.4*inch,
                                       rightMargin=0.4*inch, leftMargin=0.4*inch)
            styles = getSampleStyleSheet()
            elements = []

            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1f4788'),
                alignment=TA_CENTER,
                spaceAfter=12
            )
            elements.append(Paragraph("LIQUIDACIÓN VACACIONES", title_style))

            # Empresa
            empresa_style = ParagraphStyle(
                'EmpresaStyle',
                parent=styles['Normal'],
                fontSize=10,
                alignment=TA_CENTER,
                spaceAfter=8
            )
            elements.append(Paragraph(f"<b>{empresa}</b>", empresa_style))

            # SECCIÓN 1: DATOS DEL EMPLEADO
            empleado_header = [['DATOS DEL EMPLEADO']]
            empleado_header_table = Table(empleado_header, colWidths=[6*inch])
            empleado_header_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(empleado_header_table)

            # Datos empleado
            datos_emp = [
                ['Nombre:', nombre],
                ['Documento:', doc],
                ['Salario Mensual:', f"${salario:,.2f}"],
                ['Días Laborados:', f"{int(dias)}"],
            ]

            datos_emp_table = Table(datos_emp, colWidths=[2*inch, 4*inch])
            datos_emp_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ]))
            elements.append(datos_emp_table)
            elements.append(Spacer(1, 0.2 * inch))

            # SECCIÓN 2: CONCEPTO Y VALOR
            concepto_header = [['CONCEPTO', 'VALOR']]
            concepto_header_table = Table(concepto_header, colWidths=[4*inch, 2*inch])
            concepto_header_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(concepto_header_table)

            # Conceptos
            conceptos = [
                ['Salario Mensual:', f"${salario:,.2f}"],
                ['Días Laborados:', f"{int(dias)}"],
                ['VACACIONES PROPORCIONALES', f"${valor_vacaciones:,.2f}"],
            ]

            conceptos_table = Table(conceptos, colWidths=[4*inch, 2*inch])
            conceptos_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#B8D4E8')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 11),
            ]))
            elements.append(conceptos_table)
            elements.append(Spacer(1, 0.15 * inch))

            # Fórmula
            elements.append(Paragraph(
                f"<font size=8>Vacaciones = Salario × Días ÷ 720</font>",
                styles['Normal']
            ))
            elements.append(Spacer(1, 0.08 * inch))

            # Fecha de generación
            elements.append(Paragraph(
                f"<font size=8>Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</font>",
                styles['Normal']
            ))
            elements.append(Spacer(1, 0.3 * inch))

            # SECCIÓN 3: FIRMAS
            firmas_header = [['FIRMAS']]
            firmas_header_table = Table(firmas_header, colWidths=[6*inch])
            firmas_header_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(firmas_header_table)

            # Firmas
            firmas = [
                ['_' * 35, '', '_' * 35],
                ['Firma Empleado', '', 'Firma Empleador'],
                ['Nombre: _________________', '', 'Nombre: _________________'],
                ['Cédula: _________________', '', 'Cédula: _________________'],
                ['Fecha: _________________', '', 'Fecha: _________________'],
            ]

            firmas_table = Table(firmas, colWidths=[2*inch, 0.3*inch, 2*inch])
            firmas_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(firmas_table)

            doc_pdf.build(elements)

            # Agregar PDF al ZIP
            pdf_buffer.seek(0)
            filename = f"Vacaciones_{nombre.replace(' ', '_')}.pdf"
            zip_file.writestr(filename, pdf_buffer.getvalue())

    zip_buffer.seek(0)
    return StreamingResponse(
        iter([zip_buffer.getvalue()]),
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=vacaciones_proporcionales.zip"}
    )


@router.post("/exportar-pdf-individual-desde-excel")
async def exportar_pdf_individual(file: UploadFile = File(...), documento: str = ""):
    """Exporta PDF con formato profesional media carta de nómina"""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from datetime import datetime

    contents = await file.read()
    df = pd.read_excel(contents, sheet_name='Empleados')

    # Buscar empleado
    empleado = None
    for idx, row in df.iterrows():
        if str(row.get('documento', '')) == documento:
            empleado = row
            break

    if empleado is None:
        return {"error": "Empleado no encontrado"}

    doc_num = str(empleado.get('documento', ''))
    nombre = str(empleado.get('nombre', ''))
    cargo = str(empleado.get('cargo', 'N/A'))
    empresa = str(empleado.get('empresa', 'EMPRESA S.A.S'))
    salario = float(empleado.get('salario_mensual', 0) or 0)
    dias = float(empleado.get('dias_laborados', 30) or 30)
    auxilio = float(empleado.get('auxilio_transporte', 0) or 0)

    # Cálculos con días variables (prorrateado)
    salario_devengado = (salario / 30) * dias
    auxilio_devengado = (auxilio / 30) * dias
    base = salario_devengado + auxilio_devengado
    cesantias = (base * dias) / 360
    intereses = (cesantias * 0.12 * dias) / 360
    prima = (base * dias) / 360
    vacaciones = (salario_devengado * dias) / 720
    total_devengado = salario_devengado + auxilio_devengado + cesantias + intereses + prima + vacaciones
    total_deducciones = 0
    neto = total_devengado - total_deducciones

    # Crear PDF - MEDIA CARTA
    pdf_buffer = io.BytesIO()
    doc_pdf = SimpleDocTemplate(pdf_buffer, pagesize=letter,
                               topMargin=0.4*inch, bottomMargin=0.4*inch,
                               rightMargin=0.4*inch, leftMargin=0.4*inch)
    styles = getSampleStyleSheet()
    elements = []

    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f4788'),
        alignment=TA_CENTER,
        spaceAfter=12
    )
    elements.append(Paragraph("LIQUIDACIÓN DE NÓMINA", title_style))

    # Empresa
    empresa_style = ParagraphStyle(
        'EmpresaStyle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f4788'),
        alignment=TA_CENTER,
        spaceAfter=12
    )
    elements.append(Paragraph(empresa, empresa_style))

    # SECCIÓN 1: DATOS DEL EMPLEADO
    empleado_header = [['DATOS DEL EMPLEADO']]
    empleado_header_table = Table(empleado_header, colWidths=[6*inch])
    empleado_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(empleado_header_table)

    # Datos empleado
    datos_emp = [
        ['Nombre:', nombre],
        ['Documento:', doc_num],
        ['Cargo:', cargo],
        ['Salario Mensual:', f"${salario:,.2f}"],
        ['Auxilio Transporte:', f"${auxilio:,.2f}"],
        ['Días Laborados:', f"{int(dias)}"],
    ]

    datos_emp_table = Table(datos_emp, colWidths=[2*inch, 4*inch])
    datos_emp_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))
    elements.append(datos_emp_table)
    elements.append(Spacer(1, 0.2 * inch))

    # SECCIÓN 2: CONCEPTO Y VALOR
    concepto_header = [['CONCEPTO', 'VALOR']]
    concepto_header_table = Table(concepto_header, colWidths=[4*inch, 2*inch])
    concepto_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(concepto_header_table)

    # Conceptos
    conceptos = [
        ['Salario Devengado:', f"${salario_devengado:,.2f}"],
        ['Auxilio Devengado:', f"${auxilio_devengado:,.2f}"],
        ['Cesantías:', f"${cesantias:,.2f}"],
        ['Intereses Cesantías:', f"${intereses:,.2f}"],
        ['Prima de Servicios:', f"${prima:,.2f}"],
        ['Vacaciones:', f"${vacaciones:,.2f}"],
        ['TOTAL DEVENGADO', f"${total_devengado:,.2f}"],
    ]

    conceptos_table = Table(conceptos, colWidths=[4*inch, 2*inch])
    conceptos_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#B8D4E8')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
    ]))
    elements.append(conceptos_table)
    elements.append(Spacer(1, 0.15 * inch))

    # Deducciones
    elements.append(Paragraph(
        f"<font size=9><b>Deducciones:</b> ${total_deducciones:,.2f}</font>",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.08 * inch))

    # Neto
    elements.append(Paragraph(
        f"<font size=10><b>NETO A PAGAR: ${neto:,.2f}</b></font>",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.15 * inch))

    # Fecha de generación
    elements.append(Paragraph(
        f"<font size=8>Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</font>",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.3 * inch))

    # SECCIÓN 3: FIRMAS
    firmas_header = [['FIRMAS']]
    firmas_header_table = Table(firmas_header, colWidths=[6*inch])
    firmas_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(firmas_header_table)

    # Firmas
    firmas = [
        ['_' * 35, '', '_' * 35],
        ['Firma Empleado', '', 'Firma Empleador'],
        ['Nombre: _________________', '', 'Nombre: _________________'],
        ['Cédula: _________________', '', 'Cédula: _________________'],
        ['Fecha: _________________', '', 'Fecha: _________________'],
    ]

    firmas_table = Table(firmas, colWidths=[2*inch, 0.3*inch, 2*inch])
    firmas_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(firmas_table)

    doc_pdf.build(elements)

    pdf_buffer.seek(0)
    return StreamingResponse(
        iter([pdf_buffer.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=nomina_{nombre.replace(' ', '_')}.pdf"}
    )


@router.post("/exportar-pdf-vacaciones-individual")
async def exportar_pdf_vacaciones_individual(file: UploadFile = File(...), documento: str = ""):
    """Exporta PDF individual de vacaciones - Formato media carta profesional"""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from datetime import datetime

    contents = await file.read()
    df = pd.read_excel(contents, sheet_name='Empleados')

    # Buscar empleado
    empleado = None
    for idx, row in df.iterrows():
        if str(row.get('documento', '')) == documento:
            empleado = row
            break

    if empleado is None:
        return {"error": "Empleado no encontrado"}

    doc_num = str(empleado.get('documento', ''))
    nombre = str(empleado.get('nombre', ''))
    empresa = str(empleado.get('empresa', 'EMPRESA S.A.S'))
    salario = float(empleado.get('salario_mensual', 0) or 0)
    dias = float(empleado.get('dias_laborados', 30) or 30)

    valor_vacaciones = (salario * dias) / 720

    # Crear PDF - MEDIA CARTA
    pdf_buffer = io.BytesIO()
    doc_pdf = SimpleDocTemplate(pdf_buffer, pagesize=letter,
                               topMargin=0.4*inch, bottomMargin=0.4*inch,
                               rightMargin=0.4*inch, leftMargin=0.4*inch)
    styles = getSampleStyleSheet()
    elements = []

    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f4788'),
        alignment=TA_CENTER,
        spaceAfter=12
    )
    elements.append(Paragraph("LIQUIDACIÓN VACACIONES", title_style))

    # Empresa
    empresa_style = ParagraphStyle(
        'EmpresaStyle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1f4788'),
        alignment=TA_CENTER,
        spaceAfter=12
    )
    elements.append(Paragraph(empresa, empresa_style))

    # SECCIÓN 1: DATOS DEL EMPLEADO
    empleado_header = [['DATOS DEL EMPLEADO']]
    empleado_header_table = Table(empleado_header, colWidths=[6*inch])
    empleado_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(empleado_header_table)

    # Datos empleado
    datos_emp = [
        ['Nombre:', nombre],
        ['Documento:', doc_num],
        ['Salario Mensual:', f"${salario:,.2f}"],
        ['Días Laborados:', f"{int(dias)}"],
    ]

    datos_emp_table = Table(datos_emp, colWidths=[2*inch, 4*inch])
    datos_emp_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))
    elements.append(datos_emp_table)
    elements.append(Spacer(1, 0.2 * inch))

    # SECCIÓN 2: CONCEPTO Y VALOR
    concepto_header = [['CONCEPTO', 'VALOR']]
    concepto_header_table = Table(concepto_header, colWidths=[4*inch, 2*inch])
    concepto_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(concepto_header_table)

    # Conceptos
    conceptos = [
        ['Salario Mensual:', f"${salario:,.2f}"],
        ['Días Laborados:', f"{int(dias)}"],
        ['VACACIONES PROPORCIONALES', f"${valor_vacaciones:,.2f}"],
    ]

    conceptos_table = Table(conceptos, colWidths=[4*inch, 2*inch])
    conceptos_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        # Fila total (última)
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#B8D4E8')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
    ]))
    elements.append(conceptos_table)
    elements.append(Spacer(1, 0.15 * inch))

    # Fórmula
    elements.append(Paragraph(
        f"<font size=8>Vacaciones = Salario × Días ÷ 720</font>",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.08 * inch))

    # Fecha de generación
    elements.append(Paragraph(
        f"<font size=8>Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</font>",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.3 * inch))

    # SECCIÓN 3: FIRMAS
    firmas_header = [['FIRMAS']]
    firmas_header_table = Table(firmas_header, colWidths=[6*inch])
    firmas_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(firmas_header_table)

    # Firmas
    firmas = [
        ['_' * 35, '', '_' * 35],
        ['Firma Empleado', '', 'Firma Empleador'],
        ['Nombre: _________________', '', 'Nombre: _________________'],
        ['Cédula: _________________', '', 'Cédula: _________________'],
        ['Fecha: _________________', '', 'Fecha: _________________'],
    ]

    firmas_table = Table(firmas, colWidths=[2*inch, 0.3*inch, 2*inch])
    firmas_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(firmas_table)

    doc_pdf.build(elements)

    pdf_buffer.seek(0)
    return StreamingResponse(
        iter([pdf_buffer.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=vacaciones_{nombre.replace(' ', '_')}.pdf"}
    )


@router.post("/generar-plano-siigo")
async def generar_plano_siigo(file: UploadFile = File(...)):
    """Genera plano SIIGO en TXT"""
    contents = await file.read()
    df = pd.read_excel(contents, sheet_name='Empleados')

    lineas = []
    for idx, row in df.iterrows():
        doc = str(row.get('documento', ''))
        nombre = str(row.get('nombre', ''))
        salario = float(row.get('salario_mensual', 0) or 0)
        dias = float(row.get('dias_laborados', 30) or 30)
        auxilio = float(row.get('auxilio_transporte', 0) or 0)

        base = salario + auxilio
        cesantias = (base * dias) / 360
        intereses = (cesantias * 0.12 * dias) / 360
        prima = (base * dias) / 360
        vacaciones = (salario * dias) / 720

        # Formato: TIPO;CUENTA;DOCUMENTO;NOMBRE;VALOR;TIPO_MOVIMIENTO;LOTE
        # Salarios
        lineas.append(f"NOMINA;5105;{doc};{nombre};{salario:,.2f};D;01")
        # Cesantías
        lineas.append(f"NOMINA;510530;{doc};{nombre};{cesantias:,.2f};D;01")
        # Intereses
        lineas.append(f"NOMINA;510533;{doc};{nombre};{intereses:,.2f};D;01")
        # Prima
        lineas.append(f"NOMINA;510536;{doc};{nombre};{prima:,.2f};D;01")
        # Vacaciones
        lineas.append(f"NOMINA;510539;{doc};{nombre};{vacaciones:,.2f};D;01")

    contenido_txt = "\n".join(lineas)

    return StreamingResponse(
        iter([contenido_txt.encode('utf-8')]),
        media_type="text/plain",
        headers={"Content-Disposition": "attachment; filename=plano_siigo.txt"}
    )


@router.post("/exportar-nomina-individual")
async def exportar_nomina_individual(file: UploadFile = File(...), documento: str = ""):
    """Exporta PDF con nómina completa formato colombiano"""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from datetime import datetime

    contents = await file.read()
    df = pd.read_excel(contents, sheet_name='Empleados')

    # Buscar empleado
    empleado = None
    for idx, row in df.iterrows():
        if str(row.get('documento', '')) == documento:
            empleado = row
            break

    if empleado is None:
        return {"error": "Empleado no encontrado"}

    doc_num = str(empleado.get('documento', ''))
    nombre = str(empleado.get('nombre', ''))
    cargo = str(empleado.get('cargo', 'N/A'))
    empresa = str(empleado.get('empresa', 'EMPRESA S.A.S'))
    salario = float(empleado.get('salario_mensual', 0) or 0)
    dias = float(empleado.get('dias_laborados', 30) or 30)
    auxilio = float(empleado.get('auxilio_transporte', 0) or 0)

    # Cálculos
    salario_devengado = (salario / 30) * dias
    auxilio_devengado = (auxilio / 30) * dias if dias < 30 else auxilio
    total_devengos = salario_devengado + auxilio_devengado

    # Deducciones aproximadas
    salud = total_devengos * 0.04
    pension = total_devengos * 0.04
    total_deducciones = salud + pension

    neto_pagar = total_devengos - total_deducciones

    # Crear PDF
    pdf_buffer = io.BytesIO()
    doc_pdf = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.4*inch, bottomMargin=0.4*inch,
                                rightMargin=0.5*inch, leftMargin=0.5*inch)
    styles = getSampleStyleSheet()
    elements = []

    # Encabezado
    encabezado_data = [
        ['', 'LIQUIDACIÓN DE NÓMINA', f'Nombre: {empresa}'],
        ['', 'Período de pago: Enero 2026 - Enero 2026', f'NIT: N/A'],
        ['', f'Fargo de expedición: {datetime.now().strftime("%d/%m/%Y")}', f'Dirección: N/A'],
        ['', '', f'Teléfono: N/A'],
    ]

    encabezado_table = Table(encabezado_data, colWidths=[1*inch, 2.5*inch, 2.5*inch])
    encabezado_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (1, 0), (1, 0), 14),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(encabezado_table)
    elements.append(Spacer(1, 0.15 * inch))
    elements.append(Paragraph('<hr/>', styles['Normal']))
    elements.append(Spacer(1, 0.1 * inch))

    # Sección del empleado y detalles
    empleado_data = [
        ['SECCIÓN DEL EMPLEADO', '', 'DETALLES DE DEVENGOS', 'DETALLES DE DEDUCCIONES'],
        [f'Nombre completo: {nombre}', '', f'Salario base: ${salario_devengado:,.2f}', f'Salud (4%): ${salud:,.2f}'],
        [f'Cédula: {doc_num}', '', f'Auxilio transporte: ${auxilio_devengado:,.2f}', f'Pensión (4%): ${pension:,.2f}'],
        [f'Cargo: {cargo}', '', '', 'Otros descuentos: $0.00'],
        [f'Salario base mensual: ${salario:,.0f}', '', '', ''],
        [f'Días trabajados: {int(dias)}', '', '', ''],
    ]

    empleado_table = Table(empleado_data, colWidths=[2*inch, 0.5*inch, 1.75*inch, 1.75*inch])
    empleado_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(empleado_table)
    elements.append(Spacer(1, 0.15 * inch))
    elements.append(Paragraph('<hr/>', styles['Normal']))
    elements.append(Spacer(1, 0.1 * inch))

    # Resumen final
    resumen_data = [
        ['Total Devengado: ${:,.2f}'.format(total_devengos),
         'Total Deducciones: ${:,.2f}'.format(total_deducciones),
         'NETO A PAGAR: ${:,.2f}'.format(neto_pagar)],
    ]

    resumen_table = Table(resumen_data, colWidths=[1.8*inch, 1.8*inch, 2.4*inch])
    resumen_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (1, 0), 'Helvetica'),
        ('FONTNAME', (2, 0), (2, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (2, 0), (2, 0), colors.white),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(resumen_table)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph('<hr/>', styles['Normal']))
    elements.append(Spacer(1, 0.1 * inch))

    # SECCIÓN: FIRMAS
    elements.append(Spacer(1, 0.1 * inch))
    firmas_header = [['FIRMAS']]
    firmas_header_table = Table(firmas_header, colWidths=[6*inch])
    firmas_header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(firmas_header_table)

    # Firmas
    pie_data = [
        ['_' * 35, '', '_' * 35],
        ['Firma Empleado', '', 'Firma Empleador'],
        ['Nombre: _________________', '', 'Nombre: _________________'],
        ['Cédula: _________________', '', 'Cédula: _________________'],
        ['Fecha: _________________', '', 'Fecha: _________________'],
    ]

    pie_table = Table(pie_data, colWidths=[1.8*inch, 0.3*inch, 1.8*inch])
    pie_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(pie_table)
    elements.append(Spacer(1, 0.1 * inch))
    elements.append(Paragraph('<font size=7><i>Nota legal: Los aportes parafiscales en obligatorios concerne a la normación colombina vigente.</i></font>', styles['Normal']))

    doc_pdf.build(elements)

    pdf_buffer.seek(0)
    return StreamingResponse(
        iter([pdf_buffer.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=nomina_{nombre.replace(' ', '_')}.pdf"}
    )


@router.post("/exportar-plano-siigo-excel")
async def exportar_plano_siigo_excel(file: UploadFile = File(...)):
    """Exporta plano SIIGO a Excel"""
    contents = await file.read()
    df = pd.read_excel(contents, sheet_name='Empleados')

    plano_data = []
    for idx, row in df.iterrows():
        doc = str(row.get('documento', ''))
        nombre = str(row.get('nombre', ''))
        salario = float(row.get('salario_mensual', 0) or 0)
        dias = float(row.get('dias_laborados', 30) or 30)
        auxilio = float(row.get('auxilio_transporte', 0) or 0)

        base = salario + auxilio
        cesantias = (base * dias) / 360
        intereses = (cesantias * 0.12 * dias) / 360
        prima = (base * dias) / 360
        vacaciones = (salario * dias) / 720

        # Agregar líneas para cada movimiento
        plano_data.append({
            'Tipo': 'NOMINA',
            'Cuenta': '5105',
            'Documento': doc,
            'Nombre': nombre,
            'Valor': round(salario, 2),
            'Movimiento': 'D',
            'Lote': '01',
            'Concepto': 'Salarios'
        })
        plano_data.append({
            'Tipo': 'NOMINA',
            'Cuenta': '510530',
            'Documento': doc,
            'Nombre': nombre,
            'Valor': round(cesantias, 2),
            'Movimiento': 'D',
            'Lote': '01',
            'Concepto': 'Cesantías'
        })
        plano_data.append({
            'Tipo': 'NOMINA',
            'Cuenta': '510533',
            'Documento': doc,
            'Nombre': nombre,
            'Valor': round(intereses, 2),
            'Movimiento': 'D',
            'Lote': '01',
            'Concepto': 'Intereses'
        })
        plano_data.append({
            'Tipo': 'NOMINA',
            'Cuenta': '510536',
            'Documento': doc,
            'Nombre': nombre,
            'Valor': round(prima, 2),
            'Movimiento': 'D',
            'Lote': '01',
            'Concepto': 'Prima'
        })
        plano_data.append({
            'Tipo': 'NOMINA',
            'Cuenta': '510539',
            'Documento': doc,
            'Nombre': nombre,
            'Valor': round(vacaciones, 2),
            'Movimiento': 'D',
            'Lote': '01',
            'Concepto': 'Vacaciones'
        })

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Plano SIIGO"

    # Headers
    headers = ['Tipo', 'Cuenta', 'Documento', 'Nombre', 'Valor', 'Movimiento', 'Lote', 'Concepto']
    ws.append(headers)

    # Formato headers
    header_fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Datos
    for row_data in plano_data:
        ws.append([row_data[h] for h in headers])

    # Total
    total_valor = sum(r['Valor'] for r in plano_data)
    ws.append(['', '', '', '', total_valor, '', '', 'TOTAL'])

    # Formato total
    last_row = ws.max_row
    for col in range(1, 9):
        cell = ws.cell(last_row, col)
        cell.font = Font(bold=True)
        if col == 5:
            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

    # Ancho columnas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 15

    # Guardar en memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plano_siigo.xlsx"}
    )
