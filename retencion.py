"""
Módulo para cálculo de retenciones colombianas según tabla oficial DIAN 2026
Tabla de tarifas de retención en la fuente - UVT 2026: $52.374 COP
Con validación de bases mínimas por concepto
"""

# UVT 2026 en Pesos Colombianos
UVT_2026 = 52374  # COP

# Tabla oficial de retenciones 2026 por concepto
# Formato: 'concepto': {'tasa': %, 'base_uvt': UVT, 'base_minima_cop': $}
RETENCIONES_2026 = {
    # COMPRAS
    'compras_generales_declarantes': {'tasa': 2.5, 'base_uvt': 27, 'base_minima_cop': 1414098},
    'compras_generales_no_declarantes': {'tasa': 3.5, 'base_uvt': 27, 'base_minima_cop': 1414098},
    'compras_tarjeta_debito_credito': {'tasa': 1.5, 'base_uvt': 0, 'base_minima_cop': 0},
    'compras_agricolas_sin_procesamiento': {'tasa': 1.5, 'base_uvt': 92, 'base_minima_cop': 4818408},
    'compras_agricolas_con_procesamiento_declarantes': {'tasa': 2.5, 'base_uvt': 27, 'base_minima_cop': 1414098},
    'compras_agricolas_con_procesamiento_no_declarantes': {'tasa': 3.5, 'base_uvt': 27, 'base_minima_cop': 1414098},
    'compras_cafe_pergamino': {'tasa': 0.5, 'base_uvt': 160, 'base_minima_cop': 8379840},
    'compras_combustibles': {'tasa': 0.1, 'base_uvt': 0, 'base_minima_cop': 0},
    'compras_vehiculos': {'tasa': 1.0, 'base_uvt': 0, 'base_minima_cop': 0},
    'compras_oro': {'tasa': 2.5, 'base_uvt': 0, 'base_minima_cop': 0},
    'compras_bienes_raices_vivienda_hasta_10000uvt': {'tasa': 1.0, 'base_uvt': 0, 'base_minima_cop': 0},
    'compras_bienes_raices_vivienda_exceso': {'tasa': 2.5, 'base_uvt': 20000, 'base_minima_cop': 1047480000},
    'compras_bienes_raices_otro_uso': {'tasa': 2.5, 'base_uvt': 0, 'base_minima_cop': 0},

    # SERVICIOS
    'servicios_generales_declarantes': {'tasa': 4.0, 'base_uvt': 4, 'base_minima_cop': 209496},
    'servicios_generales_no_declarantes': {'tasa': 6.0, 'base_uvt': 4, 'base_minima_cop': 209496},
    'servicios_transporte_carga': {'tasa': 1.0, 'base_uvt': 4, 'base_minima_cop': 209496},
    'servicios_transporte_pasajeros_terrestre': {'tasa': 3.5, 'base_uvt': 27, 'base_minima_cop': 1414098},
    'servicios_transporte_pasajeros_aereo_maritimo': {'tasa': 1.0, 'base_uvt': 4, 'base_minima_cop': 209496},
    'servicios_temporales_aiu': {'tasa': 1.0, 'base_uvt': 4, 'base_minima_cop': 209496},
    'servicios_vigilancia_aseo_aiu': {'tasa': 2.0, 'base_uvt': 4, 'base_minima_cop': 209496},
    'servicios_salud_ips': {'tasa': 2.0, 'base_uvt': 4, 'base_minima_cop': 209496},
    'servicios_hoteles_restaurantes': {'tasa': 3.5, 'base_uvt': 4, 'base_minima_cop': 209496},
    'servicios_software_licenciamiento': {'tasa': 3.5, 'base_uvt': 0, 'base_minima_cop': 0},

    # ARRENDAMIENTO
    'arrendamiento_muebles': {'tasa': 4.0, 'base_uvt': 0, 'base_minima_cop': 0},
    'arrendamiento_inmuebles_declarantes': {'tasa': 3.5, 'base_uvt': 27, 'base_minima_cop': 1414098},
    'arrendamiento_inmuebles_no_declarantes': {'tasa': 3.5, 'base_uvt': 27, 'base_minima_cop': 1414098},

    # HONORARIOS Y COMISIONES
    'honorarios_comisiones_juridicas': {'tasa': 11.0, 'base_uvt': 0, 'base_minima_cop': 0},
    'honorarios_comisiones_naturales': {'tasa': 11.0, 'base_uvt': 0, 'base_minima_cop': 0},
    'honorarios_comisiones_no_declarantes': {'tasa': 10.0, 'base_uvt': 0, 'base_minima_cop': 0},

    # INGRESOS TRIBUTARIOS
    'otros_ingresos_declarantes': {'tasa': 2.5, 'base_uvt': 27, 'base_minima_cop': 1414098},
    'otros_ingresos_no_declarantes': {'tasa': 3.5, 'base_uvt': 27, 'base_minima_cop': 1414098},

    # RENDIMIENTOS FINANCIEROS
    'intereses_financieros': {'tasa': 7.0, 'base_uvt': 0, 'base_minima_cop': 0},
    'rendimientos_renta_fija': {'tasa': 4.0, 'base_uvt': 0, 'base_minima_cop': 0},

    # RETENCIÓN IVA
    'retencion_iva_servicios': {'tasa': 15.0, 'base_uvt': 4, 'base_minima_cop': 209496},
    'retencion_iva_compras': {'tasa': 15.0, 'base_uvt': 27, 'base_minima_cop': 1414098},

    # CONSTRUCCIÓN
    'construccion_urbanizacion': {'tasa': 2.0, 'base_uvt': 27, 'base_minima_cop': 1414098},

    # VENTAS (autoretención)
    'ventas_autorretencion': {'tasa': 1.2, 'base_uvt': 0, 'base_minima_cop': 0},
}

# Mapeo de actividades económicas (CIIU)
ACTIVIDADES_ECONOMICAS = {
    '4631': 'Comercio al por mayor',
    '4632': 'Comercio al por menor',
    '7820': 'Servicios de TI',
    '6820': 'Arrendamiento de bienes',
}

# Mapeo de NITs especiales con retenciones específicas
NITS_ESPECIALES = {
    # 'NIT': {'concepto': 'servicios_software_licenciamiento', 'tasa': 3.5}

# Mapeo de tipos de operación a conceptos DIAN 2026
TIPOS_OPERACION_DIAN = {
    # COMPRAS
    'Compras Generales (Declarantes)': 'compras_generales_declarantes',
    'Compras Generales (No Declarantes)': 'compras_generales_no_declarantes',
    'Compras con Tarjeta Débito/Crédito': 'compras_tarjeta_debito_credito',
    'Compras Agrícolas sin Procesamiento': 'compras_agricolas_sin_procesamiento',
    'Compras Café Pergamino': 'compras_cafe_pergamino',
    'Compras Combustibles': 'compras_combustibles',
    'Compras Vehículos': 'compras_vehiculos',
    'Compras Oro': 'compras_oro',
    'Compras Bienes Raíces Vivienda': 'compras_bienes_raices_vivienda_hasta_10000uvt',

    # SERVICIOS
    'Servicios Generales (Declarantes)': 'servicios_generales_declarantes',
    'Servicios Generales (No Declarantes)': 'servicios_generales_no_declarantes',
    'Servicios Transporte Carga': 'servicios_transporte_carga',
    'Servicios Transporte Pasajeros': 'servicios_transporte_pasajeros_terrestre',
    'Servicios Temporales': 'servicios_temporales_aiu',
    'Servicios Vigilancia y Aseo': 'servicios_vigilancia_aseo_aiu',
    'Servicios Salud': 'servicios_salud_ips',
    'Servicios Hoteles y Restaurantes': 'servicios_hoteles_restaurantes',
    'Servicios Software/Licenciamiento': 'servicios_software_licenciamiento',

    # ARRENDAMIENTO
    'Arrendamiento Muebles': 'arrendamiento_muebles',
    'Arrendamiento Inmuebles (Declarantes)': 'arrendamiento_inmuebles_declarantes',
    'Arrendamiento Inmuebles (No Declarantes)': 'arrendamiento_inmuebles_no_declarantes',

    # HONORARIOS
    'Honorarios y Comisiones': 'honorarios_comisiones_naturales',

    # OTROS
    'Intereses Financieros': 'intereses_financieros',
    'Construcción y Urbanización': 'construccion_urbanizacion',
}
}

def obtener_concepto_por_tipo(tipo_documento, es_declarante=True):
    """
    Obtiene el concepto de retención según el tipo de documento.

    Args:
        tipo_documento: Tipo de documento ('compras', 'servicios', 'honorarios', etc.)
        es_declarante: Si es declarante de renta (afecta tasa)

    Returns:
        Concepto de retención (clave en RETENCIONES_2026)
    """

    tipo_lower = tipo_documento.lower().replace(' ', '_')

    # Mapeo de tipos generales a conceptos específicos
    tipo_map = {
        'compras': 'compras_generales_declarantes' if es_declarante else 'compras_generales_no_declarantes',
        'servicios': 'servicios_generales_declarantes' if es_declarante else 'servicios_generales_no_declarantes',
        'software': 'servicios_software_licenciamiento',
        'arrendamiento': 'arrendamiento_inmuebles_declarantes' if es_declarante else 'arrendamiento_inmuebles_no_declarantes',
        'honorarios': 'honorarios_comisiones_no_declarantes' if not es_declarante else 'honorarios_comisiones_naturales',
        'ventas': 'ventas_autorretencion',
    }

    # Intentar encontrar mapeo exacto
    if tipo_lower in RETENCIONES_2026:
        return tipo_lower

    # Intentar mapeo por palabra clave
    for clave, concepto in tipo_map.items():
        if clave in tipo_lower:
            return concepto

    # Por defecto según tipo
    if 'compra' in tipo_lower:
        return 'compras_generales_declarantes'
    elif 'venta' in tipo_lower:
        return 'ventas_autorretencion'
    elif 'servicio' in tipo_lower or 'software' in tipo_lower:
        return 'servicios_software_licenciamiento'

    # Default: compras genéricas declarantes
    return 'compras_generales_declarantes'

def calcular_retencion(base_gravable, tipo_operacion='compras', nit_proveedor='', actividad='4631', es_declarante=True):
    """
    Calcula la retención aplicable según la operación.
    Valida que la base sea superior a la base mínima del concepto en COP.

    Args:
        base_gravable: Base sobre la cual se calcula (en COP)
        tipo_operacion: 'ventas', 'compras', 'servicios', 'software', etc.
        nit_proveedor: NIT del proveedor (para buscar en tabla especial)
        actividad: Actividad económica (ej: '4631')
        es_declarante: Si es declarante de renta

    Returns:
        dict con: base, tasa, valor_retencion, aplica (bool), concepto, base_minima_cop
    """

    # Obtener concepto
    if nit_proveedor in NITS_ESPECIALES:
        concepto = NITS_ESPECIALES[nit_proveedor].get('concepto')
    else:
        concepto = obtener_concepto_por_tipo(tipo_operacion, es_declarante)

    # Obtener datos del concepto
    concepto_data = RETENCIONES_2026.get(concepto, RETENCIONES_2026['compras_generales_declarantes'])
    tasa = concepto_data['tasa']
    base_minima_cop = concepto_data['base_minima_cop']

    # Validar si aplica retención según base mínima
    aplica = base_gravable >= base_minima_cop if base_minima_cop > 0 else True

    if aplica:
        valor_retencion = round(base_gravable * tasa / 100, 2)
    else:
        valor_retencion = 0

    return {
        'base': round(base_gravable, 2),
        'tasa': tasa,
        'valor': valor_retencion,
        'tipo': tipo_operacion,
        'actividad': actividad,
        'es_declarante': es_declarante,
        'concepto': concepto,
        'base_minima_cop': base_minima_cop,
        'aplica': aplica
    }

def generar_reporte_retenciones_excel(planos_data, actividad='4631', es_declarante=True):
    """
    Genera reporte de retenciones a partir de datos de planos.
    Utiliza tabla oficial 2026 DIAN con validación de bases mínimas en COP.
    Crea dos hojas: Detalle y Consolidadas

    Args:
        planos_data: Dict con datos de planos {'compras': [...], 'ventas': [...]}
        actividad: Actividad económica
        es_declarante: Si es declarante de renta

    Returns:
        Excel workbook con retenciones calculadas en pesos colombianos
    """

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
    except ImportError:
        return None

    wb = Workbook()
    
    # ===== HOJA 1: DETALLE =====
    ws_detalle = wb.active
    ws_detalle.title = 'Detalle'

    # Encabezados
    headers = ['Tipo', 'Documento', 'Fecha', 'NIT', 'Tercero', 'Base Gravable (COP)', 'Base Mín. (COP)', 'Tasa %', 'Aplica', 'Valor Retención (COP)']

    # Estilo encabezado
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for col, header in enumerate(headers, 1):
        cell = ws_detalle.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Ancho de columnas
    ws_detalle.column_dimensions['A'].width = 10
    ws_detalle.column_dimensions['B'].width = 12
    ws_detalle.column_dimensions['C'].width = 12
    ws_detalle.column_dimensions['D'].width = 12
    ws_detalle.column_dimensions['E'].width = 20
    ws_detalle.column_dimensions['F'].width = 16
    ws_detalle.column_dimensions['G'].width = 16
    ws_detalle.column_dimensions['H'].width = 10
    ws_detalle.column_dimensions['I'].width = 8
    ws_detalle.column_dimensions['J'].width = 16

    row = 2
    totales = {'base': 0, 'retencion': 0, 'excluidas': 0, 'compras_base': 0, 'compras_ret': 0, 'ventas_base': 0, 'ventas_ret': 0}

    # Formato de moneda COP
    currency_format = '#,##0'

    # Procesar compras
    for doc in planos_data.get('compras', []):
        base = doc.get('total', 0) - doc.get('iva', 0)
        ret = calcular_retencion(base, 'compras', doc.get('nit', ''), actividad, es_declarante)

        ws_detalle.cell(row=row, column=1, value='Compra')
        ws_detalle.cell(row=row, column=2, value=doc.get('folio', ''))
        ws_detalle.cell(row=row, column=3, value=doc.get('fecha', ''))
        ws_detalle.cell(row=row, column=4, value=doc.get('nit', ''))
        ws_detalle.cell(row=row, column=5, value=doc.get('nombre', ''))
        
        # Base Gravable
        cell_base = ws_detalle.cell(row=row, column=6, value=ret['base'])
        cell_base.number_format = currency_format
        
        # Base Mínima
        cell_min = ws_detalle.cell(row=row, column=7, value=ret['base_minima_cop'] if ret['base_minima_cop'] > 0 else None)
        cell_min.number_format = currency_format
        
        # Tasa
        ws_detalle.cell(row=row, column=8, value=ret['tasa'])
        
        # Aplica
        ws_detalle.cell(row=row, column=9, value='✓' if ret['aplica'] else '✗')
        
        # Valor Retención
        cell_ret = ws_detalle.cell(row=row, column=10, value=ret['valor'] if ret['aplica'] else 0)
        cell_ret.number_format = currency_format

        totales['base'] += ret['base']
        totales['compras_base'] += ret['base']
        if ret['aplica']:
            totales['retencion'] += ret['valor']
            totales['compras_ret'] += ret['valor']
        else:
            totales['excluidas'] += ret['base']

        row += 1

    # Procesar ventas
    for doc in planos_data.get('ventas', []):
        base = doc.get('total', 0) - doc.get('iva', 0)
        ret = calcular_retencion(base, 'ventas', doc.get('nit', ''), actividad, es_declarante)

        ws_detalle.cell(row=row, column=1, value='Venta')
        ws_detalle.cell(row=row, column=2, value=doc.get('folio', ''))
        ws_detalle.cell(row=row, column=3, value=doc.get('fecha', ''))
        ws_detalle.cell(row=row, column=4, value=doc.get('nit', ''))
        ws_detalle.cell(row=row, column=5, value=doc.get('nombre', ''))
        
        # Base Gravable
        cell_base = ws_detalle.cell(row=row, column=6, value=ret['base'])
        cell_base.number_format = currency_format
        
        # Base Mínima
        cell_min = ws_detalle.cell(row=row, column=7, value=ret['base_minima_cop'] if ret['base_minima_cop'] > 0 else None)
        cell_min.number_format = currency_format
        
        # Tasa
        ws_detalle.cell(row=row, column=8, value=ret['tasa'])
        
        # Aplica
        ws_detalle.cell(row=row, column=9, value='✓' if ret['aplica'] else '✗')
        
        # Valor Retención
        cell_ret = ws_detalle.cell(row=row, column=10, value=ret['valor'] if ret['aplica'] else 0)
        cell_ret.number_format = currency_format

        totales['base'] += ret['base']
        totales['ventas_base'] += ret['base']
        if ret['aplica']:
            totales['retencion'] += ret['valor']
            totales['ventas_ret'] += ret['valor']
        else:
            totales['excluidas'] += ret['base']

        row += 1

    # Fila de totales
    total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    total_font = Font(bold=True, size=11)

    row += 1
    ws_detalle.cell(row=row, column=5, value='TOTAL').font = total_font
    
    cell_total_base = ws_detalle.cell(row=row, column=6, value=totales['base'])
    cell_total_base.font = total_font
    cell_total_base.fill = total_fill
    cell_total_base.number_format = currency_format
    
    cell_total_ret = ws_detalle.cell(row=row, column=10, value=totales['retencion'])
    cell_total_ret.font = total_font
    cell_total_ret.fill = total_fill
    cell_total_ret.number_format = currency_format

    # Nota de excluidas
    row += 2
    ws_detalle.cell(row=row, column=5, value=f"Base excluida por mínimo: $ {totales['excluidas']:,.2f} COP").font = Font(italic=True, size=9)

    # ===== HOJA 2: CONSOLIDADAS =====
    ws_consolidado = wb.create_sheet('Consolidadas')
    
    # Encabezados consolidado
    headers_cons = ['Tipo Operación', 'Base Gravable (COP)', 'Valor Retención (COP)']
    
    cons_header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    cons_header_font = Font(bold=True, color='FFFFFF', size=11)
    
    for col, header in enumerate(headers_cons, 1):
        cell = ws_consolidado.cell(row=1, column=col, value=header)
        cell.fill = cons_header_fill
        cell.font = cons_header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws_consolidado.column_dimensions['A'].width = 25
    ws_consolidado.column_dimensions['B'].width = 20
    ws_consolidado.column_dimensions['C'].width = 20

    # Datos consolidados
    cons_row = 2
    
    # Compras
    ws_consolidado.cell(row=cons_row, column=1, value='COMPRAS')
    cell_comp_base = ws_consolidado.cell(row=cons_row, column=2, value=totales['compras_base'])
    cell_comp_base.number_format = currency_format
    cell_comp_ret = ws_consolidado.cell(row=cons_row, column=3, value=totales['compras_ret'])
    cell_comp_ret.number_format = currency_format
    cons_row += 1
    
    # Ventas
    ws_consolidado.cell(row=cons_row, column=1, value='VENTAS')
    cell_vent_base = ws_consolidado.cell(row=cons_row, column=2, value=totales['ventas_base'])
    cell_vent_base.number_format = currency_format
    cell_vent_ret = ws_consolidado.cell(row=cons_row, column=3, value=totales['ventas_ret'])
    cell_vent_ret.number_format = currency_format
    cons_row += 1
    
    # Total
    cons_row += 1
    total_cons_fill = PatternFill(start_color='C5E1A5', end_color='C5E1A5', fill_type='solid')
    total_cons_font = Font(bold=True, size=11)
    
    cell_total_label = ws_consolidado.cell(row=cons_row, column=1, value='TOTAL')
    cell_total_label.font = total_cons_font
    cell_total_label.fill = total_cons_fill
    
    cell_total_cons_base = ws_consolidado.cell(row=cons_row, column=2, value=totales['base'])
    cell_total_cons_base.font = total_cons_font
    cell_total_cons_base.fill = total_cons_fill
    cell_total_cons_base.number_format = currency_format
    
    cell_total_cons_ret = ws_consolidado.cell(row=cons_row, column=3, value=totales['retencion'])
    cell_total_cons_ret.font = total_cons_font
    cell_total_cons_ret.fill = total_cons_fill
    cell_total_cons_ret.number_format = currency_format

    return wb
