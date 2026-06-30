#!/usr/bin/env python3
"""
Mapeador Formulario 210 DIAN - Lector y Editor de Casillas
Mapea todas las celdas del formulario según normativa DIAN
"""

from openpyxl import load_workbook
from dataclasses import dataclass
from typing import Dict, Optional

@dataclass
class Casilla:
    """Representa una casilla del formulario"""
    numero: int
    descripcion: str
    fila: int
    columna: str
    es_calculada: bool = False
    valor: Optional[float] = None

class Formulario210Mapper:
    """Mapea y lee/escribe el Formulario 210"""

    def __init__(self, ruta_archivo: str):
        self.ruta_archivo = ruta_archivo
        self.wb = None
        self.ws = None
        self.casillas: Dict[int, Casilla] = {}
        self.uvt_2024 = 47065

    def abrir(self):
        """Abre el archivo Excel"""
        try:
            self.wb = load_workbook(self.ruta_archivo, data_only=False)
            self.ws = self.wb['210']
            self.definir_casillas()  # Definir casillas al abrir
            print("✓ Archivo abierto correctamente")
            return True
        except Exception as e:
            print(f"✗ Error abriendo archivo: {e}")
            return False

    def obtener_valor(self, fila: int, columna: str) -> Optional[float]:
        """Obtiene valor de una celda"""
        try:
            celda = self.ws[f'{columna}{fila}']
            if celda.value is None:
                return 0
            if isinstance(celda.value, (int, float)):
                return float(celda.value)
            try:
                return float(celda.value)
            except:
                return 0
        except:
            return 0

    def definir_casillas(self):
        """Define todas las casillas del formulario"""

        # PATRIMONIO
        self.casillas[29] = Casilla(29, "Total patrimonio bruto", 9, "H")
        self.casillas[30] = Casilla(30, "Deudas", 9, "O")
        self.casillas[31] = Casilla(31, "Total patrimonio líquido", 9, "X")

        # CÉDULA GENERAL - RENTAS DE TRABAJO
        self.casillas[32] = Casilla(32, "Ingresos brutos - Rentas de trabajo", 11, "H")
        self.casillas[33] = Casilla(33, "Ingresos no constitutivos - Rentas de trabajo", 13, "H")
        self.casillas[34] = Casilla(34, "Renta líquida - Rentas de trabajo", 15, "H", es_calculada=True)
        self.casillas[35] = Casilla(35, "Rentas exentas AFC/AVC - Rentas de trabajo", 17, "H")
        self.casillas[36] = Casilla(36, "Otras rentas exentas - Rentas de trabajo", 18, "H")
        self.casillas[37] = Casilla(37, "Total rentas exentas - Rentas de trabajo", 19, "H", es_calculada=True)
        self.casillas[38] = Casilla(38, "Intereses de vivienda - Rentas de trabajo", 20, "H")
        self.casillas[39] = Casilla(39, "Otras deducciones - Rentas de trabajo", 21, "H")
        self.casillas[40] = Casilla(40, "Total deducciones - Rentas de trabajo", 22, "H", es_calculada=True)
        self.casillas[41] = Casilla(41, "Rtas exentas/ded. limit. - Rentas de trabajo", 23, "H", es_calculada=True)
        self.casillas[42] = Casilla(42, "Renta líq. ord. ejercicio - Rentas de trabajo", 27, "H", es_calculada=True)

        # CÉDULA GENERAL - RENTAS DE TRABAJO SIN RELACIÓN LABORAL
        self.casillas[43] = Casilla(43, "Ingresos brutos - Trabajo sin relación laboral", 11, "N")
        self.casillas[44] = Casilla(44, "Ingresos no constitutivos - Trabajo sin relación", 13, "N")
        self.casillas[45] = Casilla(45, "Costos y deducciones - Trabajo sin relación", 14, "N")
        self.casillas[46] = Casilla(46, "Renta líquida - Trabajo sin relación", 15, "N", es_calculada=True)
        self.casillas[47] = Casilla(47, "Rentas exentas AFC/AVC - Trabajo sin relación", 17, "N")
        self.casillas[48] = Casilla(48, "Otras rentas exentas - Trabajo sin relación", 18, "N")
        self.casillas[49] = Casilla(49, "Total rentas exentas - Trabajo sin relación", 19, "N", es_calculada=True)
        self.casillas[50] = Casilla(50, "Intereses de vivienda - Trabajo sin relación", 20, "N")
        self.casillas[51] = Casilla(51, "Otras deducciones - Trabajo sin relación", 21, "N")
        self.casillas[52] = Casilla(52, "Total deducciones - Trabajo sin relación", 22, "N", es_calculada=True)
        self.casillas[53] = Casilla(53, "Rtas exentas/ded. limit. - Trabajo sin relación", 23, "N", es_calculada=True)
        self.casillas[54] = Casilla(54, "Renta líq. ord. ejercicio - Trabajo sin relación", 24, "N", es_calculada=True)
        self.casillas[55] = Casilla(55, "Pérdida líquida - Trabajo sin relación", 25, "N", es_calculada=True)
        self.casillas[56] = Casilla(56, "Compensaciones pérdidas - Trabajo sin relación", 26, "N")
        self.casillas[57] = Casilla(57, "Renta líquida ordinaria - Trabajo sin relación", 27, "N", es_calculada=True)

        # CÉDULA GENERAL - RENTAS DE CAPITAL
        self.casillas[58] = Casilla(58, "Ingresos brutos - Rentas de capital", 11, "U")
        self.casillas[59] = Casilla(59, "Ingresos no constitutivos - Rentas de capital", 13, "U")
        self.casillas[60] = Casilla(60, "Costos y deducciones - Rentas de capital", 14, "U")
        self.casillas[61] = Casilla(61, "Renta líquida - Rentas de capital", 15, "U", es_calculada=True)
        self.casillas[62] = Casilla(62, "Rentas líquidas pasivas - Rentas de capital", 16, "U")
        self.casillas[63] = Casilla(63, "Rentas exentas AFC/AVC - Rentas de capital", 17, "U")
        self.casillas[64] = Casilla(64, "Otras rentas exentas - Rentas de capital", 18, "U")
        self.casillas[65] = Casilla(65, "Total rentas exentas - Rentas de capital", 19, "U", es_calculada=True)
        self.casillas[66] = Casilla(66, "Intereses de vivienda - Rentas de capital", 20, "U")
        self.casillas[67] = Casilla(67, "Otras deducciones - Rentas de capital", 21, "U")
        self.casillas[68] = Casilla(68, "Total deducciones - Rentas de capital", 22, "U", es_calculada=True)
        self.casillas[69] = Casilla(69, "Rtas exentas/ded. limit. - Rentas de capital", 23, "U", es_calculada=True)
        self.casillas[70] = Casilla(70, "Renta líq. ord. ejercicio - Rentas de capital", 24, "U", es_calculada=True)
        self.casillas[71] = Casilla(71, "Pérdida líquida - Rentas de capital", 25, "U", es_calculada=True)
        self.casillas[72] = Casilla(72, "Compensaciones pérdidas - Rentas de capital", 26, "U")
        self.casillas[73] = Casilla(73, "Renta líquida ordinaria - Rentas de capital", 27, "U", es_calculada=True)

        # CÉDULA GENERAL - RENTAS NO LABORALES
        self.casillas[74] = Casilla(74, "Ingresos brutos - Rentas no laborales", 11, "Y")
        self.casillas[75] = Casilla(75, "Devol., rebajas, desc. - Rentas no laborales", 12, "Y")
        self.casillas[76] = Casilla(76, "Ingresos no constitutivos - Rentas no laborales", 13, "Y")
        self.casillas[77] = Casilla(77, "Costos y deducciones - Rentas no laborales", 14, "Y")
        self.casillas[78] = Casilla(78, "Renta líquida - Rentas no laborales", 15, "Y", es_calculada=True)
        self.casillas[79] = Casilla(79, "Rentas líquidas pasivas - Rentas no laborales", 16, "Y")
        self.casillas[80] = Casilla(80, "Rentas exentas AFC/AVC - Rentas no laborales", 17, "Y")
        self.casillas[81] = Casilla(81, "Otras rentas exentas - Rentas no laborales", 18, "Y")
        self.casillas[82] = Casilla(82, "Total rentas exentas - Rentas no laborales", 19, "Y", es_calculada=True)
        self.casillas[83] = Casilla(83, "Intereses de vivienda - Rentas no laborales", 20, "Y")
        self.casillas[84] = Casilla(84, "Otras deducciones - Rentas no laborales", 21, "Y")
        self.casillas[85] = Casilla(85, "Total deducciones - Rentas no laborales", 22, "Y", es_calculada=True)
        self.casillas[86] = Casilla(86, "Rtas exentas/ded. limit. - Rentas no laborales", 23, "Y", es_calculada=True)
        self.casillas[87] = Casilla(87, "Renta líq. ord. ejercicio - Rentas no laborales", 24, "Y", es_calculada=True)
        self.casillas[88] = Casilla(88, "Pérdida líquida - Rentas no laborales", 25, "Y", es_calculada=True)
        self.casillas[89] = Casilla(89, "Compensaciones pérdidas - Rentas no laborales", 26, "Y")
        self.casillas[90] = Casilla(90, "Renta líquida ordinaria - Rentas no laborales", 27, "Y", es_calculada=True)

        # CÉDULA PENSIONES
        self.casillas[99] = Casilla(99, "Ingresos brutos pensiones", 30, "K")
        self.casillas[100] = Casilla(100, "Ingresos no constitutivos pensiones", 31, "K")
        self.casillas[101] = Casilla(101, "Renta líquida pensiones", 32, "K", es_calculada=True)
        self.casillas[102] = Casilla(102, "Rentas exentas pensiones", 33, "K")
        self.casillas[103] = Casilla(103, "Renta líq. grav. cédula pensiones", 34, "K", es_calculada=True)

        # CÉDULA DIVIDENDOS
        self.casillas[104] = Casilla(104, "Dividendos 2016 y anteriores", 35, "K")
        self.casillas[105] = Casilla(105, "Ingresos no constitutivos dividendos", 36, "K")
        self.casillas[106] = Casilla(106, "Renta líq. ord. 2016 y anteriores", 37, "K", es_calculada=True)
        self.casillas[107] = Casilla(107, "1a subcédula 2017+", 38, "K")
        self.casillas[108] = Casilla(108, "2a subcédula 2017+", 39, "K")
        self.casillas[109] = Casilla(109, "Dividendos exterior", 40, "K")
        self.casillas[110] = Casilla(110, "Rentas exentas casilla 109", 41, "K")
        self.casillas[111] = Casilla(111, "Renta líq. grav. art. 241", 42, "K", es_calculada=True)

        # CÉDULA GANANCIAS OCASIONALES
        self.casillas[112] = Casilla(112, "Ingresos ganancias ocasionales", 43, "K")
        self.casillas[113] = Casilla(113, "Costos ganancias ocasionales", 44, "K")
        self.casillas[114] = Casilla(114, "Ganancias no gravadas/exentas", 45, "K")
        self.casillas[115] = Casilla(115, "Ganancias ocasionales gravables", 46, "K", es_calculada=True)

        # LIQUIDACIÓN PRIVADA
        self.casillas[116] = Casilla(116, "Imp. cédula general/pensiones/dividendos", 30, "X", es_calculada=True)
        self.casillas[117] = Casilla(117, "Imp. renta presuntiva/pensiones/dividendos", 31, "X", es_calculada=True)
        self.casillas[118] = Casilla(118, "Imp. dividendos 2017+ 2a subcédula", 32, "W")
        self.casillas[119] = Casilla(119, "Imp. dividendos 2016", 33, "W")
        self.casillas[120] = Casilla(120, "Imp. dividendos exterior", 34, "X", es_calculada=True)
        self.casillas[121] = Casilla(121, "Total impuesto rentas líq. grav.", 35, "X", es_calculada=True)
        self.casillas[122] = Casilla(122, "Imp. pagados exterior", 36, "T")
        self.casillas[123] = Casilla(123, "Donaciones", 36, "Y")
        self.casillas[124] = Casilla(124, "Dividendos/particip./otros", 37, "T")
        self.casillas[125] = Casilla(125, "Total descuentos tributarios", 37, "Z", es_calculada=True)
        self.casillas[126] = Casilla(126, "Impuesto neto de renta", 38, "X", es_calculada=True)
        self.casillas[127] = Casilla(127, "Imp. ganancias ocasionales", 39, "W")
        self.casillas[128] = Casilla(128, "Desc. imp. exterior ganancias ocasionales", 40, "W")
        self.casillas[129] = Casilla(129, "Total impuesto a cargo", 41, "X", es_calculada=True)
        self.casillas[130] = Casilla(130, "Anticipo año anterior", 42, "W")
        self.casillas[131] = Casilla(131, "Saldo favor año anterior", 43, "W")
        self.casillas[132] = Casilla(132, "Retenciones año gravable", 44, "W")
        self.casillas[133] = Casilla(133, "Anticipo año siguiente", 45, "W")

        # SALDO / RESULTADO FINAL
        self.casillas[134] = Casilla(134, "Saldo a pagar por impuesto", 47, "F", es_calculada=True)
        self.casillas[135] = Casilla(135, "Sanciones", 47, "M")
        self.casillas[136] = Casilla(136, "Total saldo a pagar", 47, "T", es_calculada=True)
        self.casillas[137] = Casilla(137, "Total saldo a favor", 47, "Z", es_calculada=True)

        # DATOS ADICIONALES
        self.casillas[138] = Casilla(138, "Número de dependientes económicos", 48, "F")
        self.casillas[139] = Casilla(139, "Adición por dependientes a casilla 92", 48, "L")
        self.casillas[140] = Casilla(140, "Superó tope art. 336-1", 48, "S")
        self.casillas[141] = Casilla(141, "Aporte voluntario", 48, "Y")

        # PAGO TOTAL
        self.casillas[980] = Casilla(980, "Pago total", 51, "X", es_calculada=True)

    def leer_valores(self):
        """Lee todos los valores del formulario"""
        if not self.casillas:
            self.definir_casillas()

        for num, casilla in self.casillas.items():
            valor = self.obtener_valor(casilla.fila, casilla.columna)
            casilla.valor = valor

    def obtener_casilla(self, numero: int) -> Optional[Casilla]:
        """Obtiene una casilla, la crea si no existe"""
        if numero not in self.casillas:
            self.definir_casillas()
        return self.casillas.get(numero)

    def guardar_valor(self, numero_casilla: int, valor: float):
        """Guarda un valor en el formulario"""
        if numero_casilla not in self.casillas:
            return False

        casilla = self.casillas[numero_casilla]
        if casilla.es_calculada:
            print(f"⚠️  Casilla {numero_casilla} es calculada, no se puede editar")
            return False

        try:
            self.ws[f'{casilla.columna}{casilla.fila}'] = valor
            casilla.valor = valor
            return True
        except Exception as e:
            print(f"Error guardando casilla {numero_casilla}: {e}")
            return False

    def guardar_archivo(self):
        """Guarda los cambios en el archivo"""
        try:
            self.wb.save(self.ruta_archivo)
            print("✓ Archivo guardado correctamente")
            return True
        except Exception as e:
            print(f"✗ Error guardando archivo: {e}")
            return False

    def cerrar(self):
        """Cierra el archivo"""
        if self.wb:
            self.wb.close()
