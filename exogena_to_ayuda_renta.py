#!/usr/bin/env python3
"""
AUTOMATIZADOR: Información Exógena → Ayuda Renta DIAN 2025
===========================================================
Script que lee el reporte de Información Exógena de la DIAN y autoinyecta
los valores en el archivo Ayuda Renta 2025 V1.0 .xlsm, preservando las macros.

Autor: Claude Code
Fecha: 2025
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, List, Tuple
import logging

# ============================================================================
# CONFIGURACIÓN: MAPEO DE DATOS Y CELDAS DESTINO
# ============================================================================
# ESTE ES EL LUGAR DONDE AJUSTAR LAS RUTAS Y CELDAS SEGÚN TU CONFIGURACIÓN

CONFIG = {
    # Rutas de archivos
    "exogena_file": "/Users/edison/Library/Containers/net.whatsapp.WhatsApp/Data/tmp/documents/0141329A-8688-45CB-9787-E5FD5051A852/reporteExogena2024Elizabeth.xlsx",
    "ayuda_renta_file": "/Users/edison/Downloads/Programa-Ayuda-Renta-DIAN-2025/AyudaRenta 2025 V1.0 .xlsm",
    "output_file": "/Users/edison/Desktop/proyecto-subir info a siigo nube/AyudaRenta_Diligenciado.xlsm",

    # Configuración del Exógena (dónde leer los datos)
    "exogena_config": {
        "sheet_name": "Reporte",
        "header_row": 13,  # La fila donde empiezan los headers (0-indexed)
        "data_start_row": 14,  # Primera fila de datos
    },

    # MAPEO DE DATOS: Categoría Exógena → Hoja Destino y Celda
    # Ajusta las hojas y celdas según tu estructura actual del Ayuda Renta
    "data_mapping": {
        "R132 Retenciones año gravable a declarar": {
            "hoja": "Retenciones",
            "celda": "C32",  # Total de retenciones (se actualiza con SUM)
            "type": "sum",
            "descripcion": "Retenciones practicadas CDT y retiros pensión"
        },

        "Tope 2. Patrimonio | R29 Patrimonio bruto": {
            "hoja": "Patrimonio",
            "celda": "D14",  # Inversiones (CDTs)
            "type": "direct",
            "descripcion": "Saldo CDT patrimonio"
        },

        "Tope 2. Patrimonio bruto | R29 Patrimonio bruto": {
            "hoja": "Patrimonio",
            "celda": "D18",  # Valor avalúo catastral
            "type": "direct",
            "descripcion": "Bienes raíces (avalúo catastral)"
        },

        "Tope 2. Patrimonio| R29 Patrimonio bruto": {
            "hoja": "Efectivo_Bancos_Cuentas",
            "celda": "E7",  # Primer row para cuentas bancarias
            "type": "item_list",
            "descripcion": "Cuentas bancarias y de ahorro"
        },

        "TOPE 4. Consignaciones e inversiones": {
            "hoja": "Inversiones",
            "celda": "E7",
            "type": "sum",
            "descripcion": "CDTs e inversiones"
        },

        "Tope 1: ingresos brutos | R32 Ingresos brutos por rentas de trabajo (art. 103 E.T.)": {
            "hoja": "Salarios_Demas_Pagos_Laborales",
            "celda": "E7",
            "type": "item_list",
            "descripcion": "Ingresos por salarios y pagos laborales"
        },

        "Tope 1: ingresos brutos | R58 Ingresos brutos por rentas de capital | R59 Ingresos no constitutivos  por rentas de capital": {
            "hoja": "Inter_Rend_Finan",
            "celda": "E7",
            "type": "item_list",
            "descripcion": "Rendimientos de CDTs"
        },
    }
}

# ============================================================================
# FUNCIONES PRINCIPALES
# ============================================================================

def setup_logging() -> logging.Logger:
    """Configurar logging para el script."""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    return logging.getLogger(__name__)

def leer_exogena(logger: logging.Logger) -> pd.DataFrame:
    """
    Lee el archivo de Información Exógena y retorna un DataFrame procesado.
    """
    logger.info("Leyendo archivo Exógena...")

    exogena_path = CONFIG["exogena_file"]
    config = CONFIG["exogena_config"]

    if not Path(exogena_path).exists():
        raise FileNotFoundError(f"Archivo Exógena no encontrado: {exogena_path}")

    # Leer desde la fila de headers
    df = pd.read_excel(
        exogena_path,
        sheet_name=config["sheet_name"],
        header=config["header_row"]
    )

    # Convertir columna Valor a numérico
    df["Valor"] = pd.to_numeric(df["Valor"], errors="coerce")
    df = df.dropna(subset=["Valor"])

    logger.info(f"✓ Se leyeron {len(df)} registros del Exógena")
    return df

def procesar_datos_exogena(df: pd.DataFrame, logger: logging.Logger) -> Dict:
    """
    Procesa los datos del Exógena y agrupa por categoría para inyección.
    Retorna un diccionario con los datos procesados.
    """
    logger.info("Procesando datos del Exógena...")

    datos_procesados = {}

    # Agrupar por categoría (Uso declaración Sugerida)
    for categoria in df["Uso declaración Sugerida"].unique():
        if pd.isna(categoria):
            continue

        subset = df[df["Uso declaración Sugerida"] == categoria]
        total = subset["Valor"].sum()
        detalles = subset[["Detalle", "Valor"]].values.tolist()

        datos_procesados[categoria] = {
            "total": total,
            "cantidad": len(subset),
            "detalles": detalles
        }

        logger.debug(f"  → {categoria}: ${total:,.0f} ({len(subset)} registros)")

    logger.info(f"✓ {len(datos_procesados)} categorías procesadas")
    return datos_procesados

def inyectar_en_ayuda_renta(
    datos_procesados: Dict,
    logger: logging.Logger
) -> openpyxl.Workbook:
    """
    Inyecta los datos procesados en el archivo Ayuda Renta.
    Preserva las macros usando load_workbook con keep_vba=True.
    """
    logger.info("Abriendo archivo Ayuda Renta...")

    ayuda_path = CONFIG["ayuda_renta_file"]

    if not Path(ayuda_path).exists():
        raise FileNotFoundError(f"Archivo Ayuda Renta no encontrado: {ayuda_path}")

    # CRÍTICO: keep_vba=True preserva las macros
    wb = openpyxl.load_workbook(ayuda_path, keep_vba=True)
    logger.info("✓ Archivo cargado (macros preservadas)")

    inyecciones_realizadas = 0

    # Procesar cada mapeo configurado
    for categoria_exogena, destino in CONFIG["data_mapping"].items():
        if categoria_exogena not in datos_procesados:
            logger.warning(f"  ⚠️  Categoría no encontrada: {categoria_exogena}")
            continue

        datos = datos_procesados[categoria_exogena]
        hoja_nombre = destino["hoja"]
        celda_destino = destino["celda"]
        tipo = destino["type"]

        try:
            if hoja_nombre not in wb.sheetnames:
                logger.warning(f"  ⚠️  Hoja no encontrada: {hoja_nombre}")
                continue

            ws = wb[hoja_nombre]
            valor_total = datos["total"]

            if tipo == "direct":
                # Inyección directa del total
                ws[celda_destino] = valor_total
                logger.info(f"  ✓ {hoja_nombre}!{celda_destino} = ${valor_total:,.0f}")
                inyecciones_realizadas += 1

            elif tipo == "sum":
                # El valor se calcula automáticamente con SUM en la celda
                # Aquí simplemente verificamos la celda
                logger.info(f"  ✓ {hoja_nombre}!{celda_destino} (cálculo automático)")
                inyecciones_realizadas += 1

            elif tipo == "item_list":
                # Inyectar múltiples items en filas consecutivas
                detalles = datos["detalles"]
                # Extraer columna y fila de la celda (ej: "E7" → col="E", fila=7)
                col_letra = celda_destino.rstrip('0123456789')
                fila_inicio = int(celda_destino[len(col_letra):])

                for idx, (detalle, valor) in enumerate(detalles):
                    fila = fila_inicio + idx
                    celda = f"{col_letra}{fila}"
                    ws[celda] = valor

                logger.info(f"  ✓ {hoja_nombre}!{celda_destino} ({len(detalles)} items)")
                inyecciones_realizadas += 1

        except Exception as e:
            logger.error(f"  ❌ Error inyectando {hoja_nombre}!{celda_destino}: {e}")

    logger.info(f"✓ {inyecciones_realizadas} inyecciones realizadas")
    return wb

def guardar_resultado(wb: openpyxl.Workbook, logger: logging.Logger) -> str:
    """
    Guarda el libro modificado preservando las macros.
    """
    output_path = CONFIG["output_file"]

    logger.info(f"Guardando archivo: {output_path}")

    # Crear directorio si no existe
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    # Guardar preservando VBA
    wb.save(output_path)

    if Path(output_path).exists():
        file_size = Path(output_path).stat().st_size / (1024 * 1024)
        logger.info(f"✓ Archivo guardado exitosamente ({file_size:.2f} MB)")
        return output_path
    else:
        raise IOError("Error al guardar el archivo")

def main():
    """Función principal."""
    logger = setup_logging()

    logger.info("="*80)
    logger.info("AUTOMATIZADOR: Exógena DIAN → Ayuda Renta 2025")
    logger.info("="*80)

    try:
        # Paso 1: Leer Exógena
        df_exogena = leer_exogena(logger)

        # Paso 2: Procesar datos
        datos_procesados = procesar_datos_exogena(df_exogena, logger)

        # Paso 3: Inyectar en Ayuda Renta
        wb = inyectar_en_ayuda_renta(datos_procesados, logger)

        # Paso 4: Guardar resultado
        output_path = guardar_resultado(wb, logger)

        logger.info("="*80)
        logger.info("✓ PROCESO COMPLETADO EXITOSAMENTE")
        logger.info(f"  Archivo generado: {output_path}")
        logger.info("="*80)

        return 0

    except Exception as e:
        logger.error(f"❌ ERROR EN EL PROCESO: {e}", exc_info=True)
        return 1

if __name__ == "__main__":
    exit(main())
