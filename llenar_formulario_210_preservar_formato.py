#!/usr/bin/env python3
"""
Llenador de Formulario 210 - Preserva formato original
Solo llena celdas específicas DIAN
"""

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

class LlenadorFormulario210Formato:
    """Llena formulario 210 preservando formato original"""

    def __init__(self, ruta_formulario, ruta_exogena):
        self.ruta_formulario = ruta_formulario
        self.ruta_exogena = ruta_exogena
        self.wb = None
        self.ws = None
        self.datos = {}

    def extraer_exogena(self):
        """Extrae datos de exógena"""
        df = pd.read_excel(self.ruta_exogena, sheet_name='Reporte', header=None)

        # Datos del contribuyente
        for i in range(len(df)):
            row = df.iloc[i]
            if pd.notna(row.iloc[0]):
                txt = str(row.iloc[0]).lower()
                if 'identificación:' in txt and pd.notna(row.iloc[2]):
                    self.datos['cedula'] = str(row.iloc[2])
                elif 'nombres' in txt and pd.notna(row.iloc[2]):
                    self.datos['nombre'] = str(row.iloc[2])
                elif 'año' in txt and 'refiere' in txt and pd.notna(row.iloc[2]):
                    self.datos['ano'] = int(row.iloc[2])

        # Ingresos y retenciones
        self.datos['ingresos'] = {}
        self.datos['retenciones'] = {}

        for i in range(14, len(df)):
            row = df.iloc[i]
            if pd.notna(row.iloc[4]) and pd.notna(row.iloc[5]):
                concepto = str(row.iloc[4]).strip()
                try:
                    valor = float(row.iloc[5])
                    if valor > 0:
                        uso = str(row.iloc[6]).lower() if pd.notna(row.iloc[6]) else ""

                        # Clasificar
                        if 'retención' in uso or 'r132' in uso or 'cdt retención' in concepto.lower():
                            if concepto not in self.datos['retenciones']:
                                self.datos['retenciones'][concepto] = 0
                            self.datos['retenciones'][concepto] += valor
                        else:
                            if concepto not in self.datos['ingresos']:
                                self.datos['ingresos'][concepto] = 0
                            self.datos['ingresos'][concepto] += valor
                except:
                    pass

        self.datos['total_ingresos'] = sum(self.datos['ingresos'].values())
        self.datos['total_retenciones'] = sum(self.datos['retenciones'].values())

        return self.datos

    def abrir_formulario(self):
        """Abre el formulario sin modificarlo"""
        try:
            self.wb = load_workbook(self.ruta_formulario)
            self.ws = self.wb.active
            return True
        except Exception as e:
            print(f"Error: {e}")
            return False

    def llenar_celdas_especificas(self):
        """Llena SOLO las celdas específicas del formulario 210"""
        print("\n📝 LLENANDO FORMULARIO 210")
        print("="*70)

        # Línea 1: Año (E14)
        self.ws['E14'].value = self.datos.get('ano', 2024)
        print(f"✓ Año (E14): {self.datos.get('ano')}")

        # Línea 5: NIT/Cédula - buscar celda cercana a D20
        # Normalmente después de "Número de Identificación Tributaria (NIT)"
        for row_idx in range(20, 25):
            cell = self.ws.cell(row=row_idx, column=7)  # Columna G
            if cell.value is None:
                cell.value = self.datos.get('cedula', '')
                print(f"✓ Cédula (G{row_idx}): {self.datos.get('cedula')}")
                break

        # Línea 32: Ingresos brutos (N27)
        self.ws['N27'].value = int(self.datos.get('total_ingresos', 0))
        print(f"✓ Ingresos Brutos (N27): ${self.datos.get('total_ingresos', 0):,.0f}")

        # Línea 33: Devoluciones/descuentos (N28) - si es necesario
        # self.ws['N28'].value = 0

        # Línea 34: Renta Líquida (N31)
        deduccion = self.datos.get('total_ingresos', 0) * 0.10
        base_liquida = self.datos.get('total_ingresos', 0) - deduccion
        self.ws['N31'].value = int(base_liquida)
        print(f"✓ Renta Líquida (N31): ${base_liquida:,.0f}")

        # Calcular retenciones total
        total_retenciones = self.datos.get('total_retenciones', 0)

        print(f"\n✓ Total Retenciones: ${total_retenciones:,.0f}")
        print(f"✓ Base Líquida: ${base_liquida:,.0f}")

        return {
            'base_liquida': base_liquida,
            'retenciones': total_retenciones
        }

    def guardar(self, ruta_salida=None):
        """Guarda sin modificar formato"""
        if not ruta_salida:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            cedula = self.datos.get('cedula', 'XXX')
            ruta_salida = f"Formulario_210_Llenado_{cedula}_{timestamp}.xlsx"

        try:
            self.wb.save(ruta_salida)
            return ruta_salida
        except Exception as e:
            print(f"Error guardando: {e}")
            return None

    def procesar(self):
        """Proceso completo"""
        print("\n╔════════════════════════════════════════════════════════════╗")
        print("║  LLENADOR FORMULARIO 210 - PRESERVA FORMATO ORIGINAL      ║")
        print("╚════════════════════════════════════════════════════════════╝")

        # Extraer
        print("\n📂 Extrayendo datos de Exógena...")
        self.extraer_exogena()
        print(f"   ✓ {self.datos.get('nombre')} (CC {self.datos.get('cedula')})")
        print(f"   ✓ Ingresos: ${self.datos.get('total_ingresos', 0):,.0f}")
        print(f"   ✓ Retenciones: ${self.datos.get('total_retenciones', 0):,.0f}")

        # Abrir
        print("\n📋 Abriendo Formulario 210...")
        if not self.abrir_formulario():
            return None

        # Llenar
        calcs = self.llenar_celdas_especificas()

        # Guardar
        print("\n💾 Guardando formulario...")
        ruta = self.guardar()
        print(f"   ✓ {ruta}")

        # Resumen
        print("\n" + "="*70)
        print("📊 RESUMEN")
        print("="*70)
        print(f"Ingresos:         ${self.datos.get('total_ingresos', 0):>17,.0f}")
        print(f"Deducción (10%):  ${self.datos.get('total_ingresos', 0)*0.1:>17,.0f}")
        print(f"Base Líquida:     ${calcs['base_liquida']:>17,.0f}")
        print(f"Retenciones:      ${calcs['retenciones']:>17,.0f}")
        print("="*70)

        print(f"\n✅ Formulario llenado exitosamente!")
        print(f"📄 Archivo: {ruta}")
        print(f"\n⚠️  IMPORTANTE:")
        print(f"   • Valida con tu contador")
        print(f"   • Comprueba retenciones")
        print(f"   • Revisa deducciones")
        print(f"   • Presenta a DIAN antes del 10 de abril")

        return ruta

if __name__ == "__main__":
    llenador = LlenadorFormulario210Formato(
        ruta_formulario="/Users/edison/Downloads/formulario 210.xlsx",
        ruta_exogena="/Users/edison/Downloads/reporteExogena2024Elizabeth.xlsx"
    )

    resultado = llenador.procesar()
