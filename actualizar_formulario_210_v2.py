#!/usr/bin/env python3
"""
Actualiza Formulario 210 con UVT 2025 y valores calculados
"""

import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import json

class ActualizadorFormulario210:
    """Actualiza formulario 210 con liquidación completa"""

    def __init__(self, ruta_formulario, ruta_exogena):
        self.ruta_formulario = ruta_formulario
        self.ruta_exogena = ruta_exogena
        self.wb = None
        self.ws = None
        self.datos = {}

    def extraer_exogena(self):
        """Extrae datos de exógena"""
        df = pd.read_excel(self.ruta_exogena, sheet_name='Reporte', header=None)

        for i in range(len(df)):
            row = df.iloc[i]
            if pd.notna(row.iloc[0]):
                txt = str(row.iloc[0]).lower()
                if 'identificación:' in txt and pd.notna(row.iloc[2]):
                    self.datos['cedula'] = str(row.iloc[2])
                elif 'nombres' in txt and pd.notna(row.iloc[2]):
                    self.datos['nombre'] = str(row.iloc[2])
                elif 'año' in txt and 'refiere' in txt and pd.notna(row.iloc[2]):
                    self.datos['ano'] = int(row.iloc[2])

        self.datos['ingresos'] = {}
        self.datos['retenciones'] = {}

        for i in range(14, len(df)):
            row = df.iloc[i]
            if pd.notna(row.iloc[4]) and pd.notna(row.iloc[5]):
                concepto = str(row.iloc[4]).strip()
                try:
                    valor = float(row.iloc[5])
                    if valor > 0:
                        uso = str(row.iloc[6]).lower() if pd.notna(row.iloc[6]) else ""

                        if 'retención' in uso or 'r132' in uso or 'cdt retención' in concepto.lower():
                            if concepto not in self.datos['retenciones']:
                                self.datos['retenciones'][concepto] = 0
                            self.datos['retenciones'][concepto] += valor
                        else:
                            if concepto not in self.datos['ingresos']:
                                self.datos['ingresos'][concepto] = 0
                            self.datos['ingresos'][concepto] += valor
                except:
                    pass

        self.datos['total_ingresos'] = sum(self.datos['ingresos'].values())
        self.datos['total_retenciones'] = sum(self.datos['retenciones'].values())

    def calcular_liquidacion(self, deduccion_pct=10):
        """Calcula liquidación con UVT 2025"""
        UVT_2025 = 49799

        total_ingresos = self.datos.get('total_ingresos', 0)
        deduccion = total_ingresos * (deduccion_pct / 100)
        base_liquida = total_ingresos - deduccion

        # IRPF progresivo
        rangos = [
            (0, 66950000, 0),
            (66950000, 134900000, 0.05),
            (134900000, 404700000, 0.12),
            (404700000, 673500000, 0.25),
            (673500000, 1347000000, 0.32),
            (1347000000, float('inf'), 0.37),
        ]

        irpf = 0
        for inicio, fin, tasa in rangos:
            if base_liquida > inicio:
                tramo = min(base_liquida, fin) - inicio
                irpf += tramo * tasa

        # Aporte solidario
        limite = 16000 * UVT_2025
        aporte_solidario = 0
        if base_liquida > limite:
            aporte_solidario = (base_liquida - limite) * 0.01

        # Impuesto total
        total_impuesto = irpf + aporte_solidario
        retenciones = self.datos.get('total_retenciones', 0)
        saldo = max(0, total_impuesto - retenciones)

        return {
            'total_ingresos': total_ingresos,
            'deduccion': deduccion,
            'deduccion_pct': deduccion_pct,
            'base_liquida': base_liquida,
            'irpf': irpf,
            'aporte_solidario': aporte_solidario,
            'total_impuesto': total_impuesto,
            'retenciones': retenciones,
            'saldo_pagar': saldo,
            'uvt_2025': UVT_2025,
        }

    def abrir_formulario(self):
        """Abre formulario"""
        try:
            self.wb = load_workbook(self.ruta_formulario)
            self.ws = self.wb.active
            return True
        except Exception as e:
            print(f"Error: {e}")
            return False

    def llenar_formulario(self, calculos):
        """Llena el formulario con los cálculos"""
        # Año
        self.ws['E14'].value = self.datos.get('ano', 2024)

        # Cédula
        for row_idx in range(20, 25):
            cell = self.ws.cell(row=row_idx, column=7)
            if cell.value is None:
                cell.value = self.datos.get('cedula', '')
                break

        # Ingresos brutos (N27)
        self.ws['N27'].value = int(calculos['total_ingresos'])

        # Renta líquida (N31)
        self.ws['N31'].value = int(calculos['base_liquida'])

        return True

    def guardar(self, ruta_salida=None):
        """Guarda formulario actualizado"""
        if not ruta_salida:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            cedula = self.datos.get('cedula', 'XXX')
            ruta_salida = f"Formulario_210_UVT2025_{cedula}_{timestamp}.xlsx"

        try:
            self.wb.save(ruta_salida)
            return ruta_salida
        except Exception as e:
            print(f"Error: {e}")
            return None

    def procesar(self, deduccion_pct=10):
        """Proceso completo"""
        print("\n╔════════════════════════════════════════════════════════════╗")
        print("║    ACTUALIZAR FORMULARIO 210 - UVT 2025 ($49.799)         ║")
        print("╚════════════════════════════════════════════════════════════╝\n")

        # Extraer
        print("📂 Extrayendo datos...")
        self.extraer_exogena()
        print(f"   ✓ {self.datos.get('nombre')}")
        print(f"   ✓ Cédula: {self.datos.get('cedula')}")

        # Calcular
        print("\n🧮 Calculando liquidación...")
        calculos = self.calcular_liquidacion(deduccion_pct)

        # Mostrar resultados
        print("\n📊 RESULTADO DE LIQUIDACIÓN")
        print("─" * 70)
        print(f"Ingresos Brutos:      ${calculos['total_ingresos']:>20,.0f}")
        print(f"Deducción ({deduccion_pct}%):         ${calculos['deduccion']:>20,.0f}")
        print(f"Base Líquida:         ${calculos['base_liquida']:>20,.0f}")
        print()
        print(f"IRPF:                 ${calculos['irpf']:>20,.0f}")
        print(f"Aporte Solidario:     ${calculos['aporte_solidario']:>20,.0f}")
        print(f"─────────────────────────────────────")
        print(f"Total Impuesto:       ${calculos['total_impuesto']:>20,.0f}")
        print(f"Retenciones:          ${calculos['retenciones']:>20,.0f}")
        print(f"═════════════════════════════════════")
        print(f"✅ SALDO A PAGAR:     ${calculos['saldo_pagar']:>20,.0f}")
        print(f"UVT 2025:             $                  {calculos['uvt_2025']:>8,.0f}")

        # Abrir formulario
        print("\n📋 Abriendo formulario...")
        if not self.abrir_formulario():
            return None

        # Llenar
        print("✏️  Llenando formulario...")
        self.llenar_formulario(calculos)

        # Guardar
        print("💾 Guardando...")
        ruta = self.guardar()
        print(f"   ✓ {ruta}")

        # Guardar JSON
        resultado = {
            'contribuyente': {
                'cedula': self.datos.get('cedula'),
                'nombre': self.datos.get('nombre'),
                'ano': self.datos.get('ano'),
            },
            'calculos': calculos,
            'fecha': datetime.now().isoformat(),
        }

        json_file = ruta.replace('.xlsx', '.json')
        with open(json_file, 'w', encoding='utf-8') as f:
            import json
            json.dump(resultado, f, indent=2, ensure_ascii=False, default=str)

        print(f"   ✓ {json_file}")

        print("\n✅ PROCESO COMPLETADO")
        print(f"\n📥 Archivos generados:")
        print(f"   • {ruta}")
        print(f"   • {json_file}")

        return ruta, calculos

if __name__ == "__main__":
    actualizador = ActualizadorFormulario210(
        ruta_formulario="/Users/edison/Downloads/formulario 210.xlsx",
        ruta_exogena="/Users/edison/Downloads/reporteExogena2024Elizabeth.xlsx"
    )

    actualizador.procesar(deduccion_pct=10)
