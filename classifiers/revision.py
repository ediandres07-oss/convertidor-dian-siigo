"""
Nivel 3: revisión humana. Genera un Excel con las clasificaciones pendientes
(confianza baja o sin regla/IA) y reimporta las correcciones como reglas
nuevas para el Nivel 1.
"""

import io
import openpyxl
from openpyxl.styles import Font, PatternFill

from .reglas import agregar_regla_nit

COLUMNAS = ['NIT', 'Tercero', 'Descripción', 'Valor', 'Cuenta sugerida',
            'Confianza', 'Justificación', 'cuenta_aprobada']


def generar_excel_revision(pendientes):
    """
    pendientes: lista de dicts con 'nit', 'nombre', 'detalle', 'total',
    'cuenta', 'confianza', 'justificacion'. Retorna BytesIO del Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Revisión Pendiente'

    header_fill = PatternFill(fill_type='solid', fgColor='dc2626')
    header_font = Font(bold=True, color='FFFFFF')
    for col, nombre in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=1, column=col, value=nombre)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, doc in enumerate(pendientes, 2):
        ws.cell(row=row_idx, column=1, value=doc.get('nit'))
        ws.cell(row=row_idx, column=2, value=doc.get('nombre'))
        ws.cell(row=row_idx, column=3, value=doc.get('detalle'))
        ws.cell(row=row_idx, column=4, value=doc.get('total'))
        ws.cell(row=row_idx, column=5, value=doc.get('cuenta'))
        ws.cell(row=row_idx, column=6, value=doc.get('confianza'))
        ws.cell(row=row_idx, column=7, value=doc.get('justificacion'))
        # cuenta_aprobada queda vacía para que el usuario la llene

    anchos = [14, 30, 40, 14, 14, 10, 40, 16]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def importar_revision(empresa_nit, excel_stream):
    """
    Lee un Excel de revisión corregido (con 'cuenta_aprobada' llenado) y
    agrega una regla NIT->cuenta por cada fila corregida.
    Retorna la cantidad de reglas agregadas.
    """
    wb = openpyxl.load_workbook(excel_stream, data_only=True)
    ws = wb.active

    agregadas = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        nit, _nombre, _desc, _valor, _cuenta_sugerida, _confianza, _just, cuenta_aprobada = row[:8]
        if not nit or not cuenta_aprobada:
            continue
        agregar_regla_nit(empresa_nit, nit, str(cuenta_aprobada).strip(),
                           tipo='gasto', notas='Aprobado en revisión manual')
        agregadas += 1

    return agregadas
