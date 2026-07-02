"""
Generador de Planos para Contai-Ilimitada desde el reporte DIAN.

Toma el mismo archivo DIAN que usa el convertidor Siigo y genera el plano de
ventas, compras y notas crédito de compras en formato Contai-Ilimitada
(contabilidad de partida doble por documento).

Nota: el reporte DIAN solo trae el IVA total y el total por documento, por lo
que todo se contabiliza al 19% (no se separa 19%/5%) y no se incluyen
retenciones en la fuente. Si un documento no trae IVA, se omite la línea de IVA.

Gastos/servicios: cada proveedor usa una cuenta contable distinta (arriendo,
honorarios, servicios públicos, etc.). La clasificación compra-vs-gasto y la
cuenta específica por NIT ya NO están hardcodeadas aquí: se resuelven por
empresa (identificada por su NIT) usando classifiers/ — reglas
(reglas_empresas/<nit_empresa>.json), caché de clasificaciones previas, e IA
como fallback para NITs nuevos. Ver [[project-multi-empresa]] en memoria:
cada empresa tiene su propio archivo de reglas, no se reusan cuentas de otra.
"""

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from converter import _leer_todo, _fecha
from classifiers import clasificar_lote_documentos
from classifiers.reglas import cargar_reglas


# ========================================
# Cuentas contables Contai (INVERPLAS)
# ========================================
CTA_INGRESO_0  = '412051'      # Ingresos no gravados (sin IVA)
CTA_INGRESO_19 = '412054'      # Ingresos gravados 19%
CTA_INGRESO_5  = '412055'      # Ingresos gravados 5%
CTA_IVA_19     = '24080104'    # IVA generado 19%
CTA_IVA_5      = '24080105'    # IVA generado 5%
CTA_CLIENTES   = '130505'      # Clientes (CxC)
CTA_DEV_VENTAS    = '417502'   # Devolución en ventas 19%
CTA_DEV_VENTAS_5  = '417503'   # Devolución en ventas 5%
CTA_IVA_DEV    = '24080207'    # IVA en devoluciones (19% y 5%: no hay cuenta
                               # separada confirmada para 5%, se reutiliza esta)

COMP_VENTA   = 4               # Comprobante de venta
COMP_DEVOL   = 16              # Comprobante de devolución (NC ventas)
COMP_COMPRA  = 1               # Comprobante de compra
COMP_NC_COMPRA = 2             # Comprobante de NC de compra
COMP_GASTO   = 12              # Comprobante de gasto/servicio

TIPO_DEBITO  = 1
TIPO_CREDITO = 2

HEADERS_CONTAI = [
    'Cuenta', 'Comprobante', 'Fecha(mm/dd/yyyy)', 'Comprobante', 'Documento Ref.',
    'Nit', 'Detalle', 'Tipo', 'VALORES', 'BASE', 'Centro de Costo',
    'Trans. Ext', 'Plazo'
]

_HEADER_FILL = PatternFill(fill_type='solid', fgColor='1a3a5c')
_HEADER_FONT = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
_THIN_BORDER = Border(
    left=Side(style='thin', color='BFCCD8'),
    right=Side(style='thin', color='BFCCD8'),
    top=Side(style='thin', color='BFCCD8'),
    bottom=Side(style='thin', color='BFCCD8')
)


def _fmt_fecha(val):
    """Devuelve la fecha en formato dd-mm-yyyy (como en el plano Contai)."""
    f = _fecha(val)
    try:
        return f.strftime('%d-%m-%Y')
    except AttributeError:
        return str(val) if val is not None else ''


def _fila_contai(cuenta, comp, fecha, folio, nit, detalle, tipo, valores, base=None):
    """Construye una fila del plano Contai (13 columnas)."""
    return [
        cuenta, comp, fecha, folio, folio,
        str(nit) if nit else None, detalle, tipo, round(valores, 2),
        round(base, 2) if base is not None else None,
        None, None, None
    ]


def _escribir_headers(ws):
    for col, header in enumerate(HEADERS_CONTAI, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _THIN_BORDER
    ws.row_dimensions[1].height = 20


def _escribir_fila(ws, row_idx, fila):
    for col, val in enumerate(fila, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.border = _THIN_BORDER
        # Formato numérico para VALORES (col 9) y BASE (col 10)
        if col in (9, 10) and isinstance(val, (int, float)):
            cell.number_format = '#,##0.00'


def _clasificar_tarifa_venta(base, iva):
    """
    Determina la cuenta de ingreso/IVA según la tarifa efectiva de CADA
    factura (iva/base), ya que el reporte DIAN sí trae base e IVA por
    documento (aunque no por línea/producto). Documentos que mezclan
    productos a 19% y 5% en una misma factura se aproximan a la tarifa
    dominante, ya que el DIAN no discrimina por producto en un solo total.

    La cuenta de IVA solo es None cuando el IVA es realmente ~0 (venta no
    gravada) — si hay IVA aunque su tarifa efectiva no sea exactamente
    19% o 5%, igual se escribe (a la cuenta más cercana) para no perder
    el valor y descuadrar el plano.

    Retorna (cuenta_ingreso, cuenta_iva_o_None).
    """
    if abs(iva) < 0.01:
        return CTA_INGRESO_0, None
    rate = iva / base if abs(base) > 0.01 else 1.0
    if rate < 0.12:
        return CTA_INGRESO_5, CTA_IVA_5
    return CTA_INGRESO_19, CTA_IVA_19


def _escribir_ventas(ws, row_idx, ventas, nc_ventas):
    """Escribe las líneas de facturas de venta y sus notas crédito."""
    # --- Facturas de venta (Comprobante 4) ---
    for v in ventas:
        if abs(v['total']) < 0.01:
            continue
        fecha  = _fmt_fecha(v['fecha'])
        folio  = v['folio']
        nit    = v['nit']
        base   = v['gravado']
        iva    = v['iva']
        total  = v['total']
        cta_ingreso, cta_iva = _clasificar_tarifa_venta(base, iva)

        # Ingreso (crédito): no gravado, 19% o 5% según la tarifa efectiva
        _escribir_fila(ws, row_idx, _fila_contai(
            cta_ingreso, COMP_VENTA, fecha, folio, nit, 'VENTAS',
            TIPO_CREDITO, base))
        row_idx += 1

        # IVA generado (crédito) — solo si la factura tiene IVA
        if cta_iva and abs(iva) > 0.01:
            _escribir_fila(ws, row_idx, _fila_contai(
                cta_iva, COMP_VENTA, fecha, folio, nit, 'VENTAS',
                TIPO_CREDITO, iva, base=base))
            row_idx += 1

        # Clientes / CxC (débito) por el total
        _escribir_fila(ws, row_idx, _fila_contai(
            CTA_CLIENTES, COMP_VENTA, fecha, folio, nit, 'VENTAS',
            TIPO_DEBITO, total))
        row_idx += 1

    # --- Notas crédito / devoluciones (Comprobante 16) ---
    for nc in nc_ventas:
        if abs(nc['total']) < 0.01:
            continue
        fecha  = _fmt_fecha(nc['fecha'])
        folio  = nc['folio']
        nit    = nc['nit']
        base   = nc['gravado']
        iva    = nc['iva']
        total  = nc['total']
        cta_ingreso, cta_iva = _clasificar_tarifa_venta(base, iva)
        # Cuenta de devolución según tarifa. No hay cuenta confirmada para
        # devolución de ventas no gravadas, se usa la de 19% (mejor
        # aproximación disponible) en ese caso.
        cta_dev = CTA_DEV_VENTAS_5 if cta_ingreso == CTA_INGRESO_5 else CTA_DEV_VENTAS

        # Devolución en ventas (débito)
        _escribir_fila(ws, row_idx, _fila_contai(
            cta_dev, COMP_DEVOL, fecha, folio, nit, 'DEV. VENTAS',
            TIPO_DEBITO, base))
        row_idx += 1

        # IVA en devoluciones (débito) — solo si hay IVA
        if cta_iva and abs(iva) > 0.01:
            _escribir_fila(ws, row_idx, _fila_contai(
                CTA_IVA_DEV, COMP_DEVOL, fecha, folio, nit, 'DEV. VENTAS',
                TIPO_DEBITO, iva, base=base))
            row_idx += 1

        # Clientes / CxC (crédito) por el total
        _escribir_fila(ws, row_idx, _fila_contai(
            CTA_CLIENTES, COMP_DEVOL, fecha, folio, nit, 'DEV. VENTAS',
            TIPO_CREDITO, total))
        row_idx += 1

    return row_idx


def _clasificar_tarifa_compra(base, iva):
    """
    Igual que _clasificar_tarifa_venta pero para compras: el reporte DIAN
    trae base e IVA por documento, así que la tarifa efectiva (iva/base) se
    calcula por factura. Retorna '0', '5' o '19' (sufijo de la clave
    cuenta_compras_<tarifa> / cuenta_dev_compras_<tarifa> en el JSON de
    la empresa).
    """
    if abs(iva) < 0.01:
        return '0'
    rate = iva / base if abs(base) > 0.01 else 1.0
    if rate < 0.12:
        return '5'
    return '19'


def _cuenta_compra(cuentas, prefijo, tarifa):
    """Busca cuentas[f'{prefijo}_{tarifa}']; si la empresa no tiene esa
    cuenta configurada (ej. sin compras no gravadas), usa la de 19% como
    mejor aproximación disponible — nunca deja de escribir el valor."""
    return cuentas.get(f'{prefijo}_{tarifa}') or cuentas.get(f'{prefijo}_19')


def _escribir_compras(ws, row_idx, compras, cuentas):
    """Escribe las líneas de facturas de compra (mercancía/inventario),
    separadas por tarifa efectiva (no gravada/5%/19%) de cada factura."""
    for c in compras:
        if abs(c['total']) < 0.01:
            continue
        fecha  = _fmt_fecha(c['fecha'])
        folio  = c['folio']
        nit    = c['nit']
        base   = c['gravado']
        iva    = c['iva']
        total  = c['total']
        tarifa = _clasificar_tarifa_compra(base, iva)
        cta_compra = _cuenta_compra(cuentas, 'cuenta_compras', tarifa)

        # Compras gravadas o no gravadas (débito), según tarifa efectiva
        _escribir_fila(ws, row_idx, _fila_contai(
            cta_compra, COMP_COMPRA, fecha, folio, nit, 'COMPRAS',
            TIPO_DEBITO, base))
        row_idx += 1

        # IVA descontable (débito) — solo si hay IVA
        if abs(iva) > 0.01:
            _escribir_fila(ws, row_idx, _fila_contai(
                cuentas['cuenta_iva_descontable'], COMP_COMPRA, fecha, folio, nit, 'COMPRAS',
                TIPO_DEBITO, iva, base=base))
            row_idx += 1

        # Proveedores / CxP (crédito) por el total
        _escribir_fila(ws, row_idx, _fila_contai(
            cuentas['cuenta_proveedores'], COMP_COMPRA, fecha, folio, nit, 'COMPRAS',
            TIPO_CREDITO, total))
        row_idx += 1

    return row_idx


def _escribir_nc_compras(ws, row_idx, nc_compras, cuentas):
    """Escribe las líneas de notas crédito de compra (devoluciones a proveedor),
    separadas por tarifa efectiva igual que _escribir_compras."""
    for nc in nc_compras:
        if abs(nc['total']) < 0.01:
            continue
        fecha  = _fmt_fecha(nc['fecha'])
        folio  = nc['folio']
        nit    = nc['nit']
        base   = nc['gravado']
        iva    = nc['iva']
        total  = nc['total']
        tarifa = _clasificar_tarifa_compra(base, iva)
        cta_dev = _cuenta_compra(cuentas, 'cuenta_dev_compras', tarifa)

        # Devolución en compras (crédito)
        _escribir_fila(ws, row_idx, _fila_contai(
            cta_dev, COMP_NC_COMPRA, fecha, folio, nit, 'DEV. COMPRAS',
            TIPO_CREDITO, base))
        row_idx += 1

        # IVA descontable en devolución (crédito) — solo si hay IVA
        if abs(iva) > 0.01:
            _escribir_fila(ws, row_idx, _fila_contai(
                cuentas['cuenta_iva_descontable'], COMP_NC_COMPRA, fecha, folio, nit, 'DEV. COMPRAS',
                TIPO_CREDITO, iva, base=base))
            row_idx += 1

        # Proveedores / CxP (débito) por el total
        _escribir_fila(ws, row_idx, _fila_contai(
            cuentas['cuenta_proveedores'], COMP_NC_COMPRA, fecha, folio, nit, 'DEV. COMPRAS',
            TIPO_DEBITO, total))
        row_idx += 1

    return row_idx


def _clasificar_recibidos(empresa_nit, compras_raw, gastos_raw, notas_credito_raw):
    """
    Reclasifica TODOS los documentos recibidos (facturas + NC) de la empresa
    usando el clasificador híbrido (reglas -> caché -> IA, ver classifiers/).
    La clasificación de converter.py (compras_raw/gastos_raw, basada en
    NIT_PROVEEDOR_INVENTARIO de un solo cliente) se ignora aquí — cada
    empresa tiene su propio archivo de reglas (reglas_empresas/<nit>.json).

    Retorna (compras, gastos, nc_compras, nc_gastos, pendientes, resumen_fuentes).
    Los docs de gastos/nc_gastos traen además 'cuenta_asignada'.
    """
    reglas_empresa = cargar_reglas(empresa_nit)
    tipo_default = reglas_empresa.get('tipo_default_recibidos') or 'gasto'

    facturas = compras_raw + gastos_raw
    resultados_facturas = clasificar_lote_documentos(empresa_nit, facturas, tipo_default=tipo_default)
    resultados_nc = clasificar_lote_documentos(empresa_nit, notas_credito_raw, tipo_default=tipo_default)

    compras, gastos, pendientes = [], [], []
    resumen_fuentes = {'regla': 0, 'cache': 0, 'ia': 0, 'default': 0}

    def _procesar(doc, res, destino_compras, destino_gastos):
        resumen_fuentes[res['fuente']] = resumen_fuentes.get(res['fuente'], 0) + 1
        if res['tipo'] == 'compra':
            destino_compras.append(doc)
        else:
            destino_gastos.append({**doc, 'cuenta_asignada': res['cuenta']})
        if res['pendiente']:
            pendientes.append({**doc, 'cuenta': res['cuenta'],
                                'confianza': res['confianza'],
                                'justificacion': res['justificacion']})

    for doc, res in zip(facturas, resultados_facturas):
        _procesar(doc, res, compras, gastos)

    nc_compras, nc_gastos = [], []
    for doc, res in zip(notas_credito_raw, resultados_nc):
        _procesar(doc, res, nc_compras, nc_gastos)

    return compras, gastos, nc_compras, nc_gastos, pendientes, resumen_fuentes


def _escribir_nc_gastos(ws, row_idx, nc_gastos, cuenta_iva_descontable, cuenta_proveedores):
    """Escribe notas crédito de gastos/servicios (devolución al proveedor)."""
    for nc in nc_gastos:
        if abs(nc['total']) < 0.01:
            continue
        fecha  = _fmt_fecha(nc['fecha'])
        folio  = nc['folio']
        nit    = nc['nit']
        base   = nc['gravado']
        iva    = nc['iva']
        total  = nc['total']
        cuenta_gasto = nc['cuenta_asignada']

        # Reverso de la cuenta de gasto (crédito)
        _escribir_fila(ws, row_idx, _fila_contai(
            cuenta_gasto, COMP_NC_COMPRA, fecha, folio, nit, 'DEV. GASTOS',
            TIPO_CREDITO, base))
        row_idx += 1

        # Reverso IVA descontable (crédito) — solo si hay IVA
        if abs(iva) > 0.01:
            _escribir_fila(ws, row_idx, _fila_contai(
                cuenta_iva_descontable, COMP_NC_COMPRA, fecha, folio, nit, 'DEV. GASTOS',
                TIPO_CREDITO, iva, base=base))
            row_idx += 1

        # Proveedores / CxP (débito) por el total
        _escribir_fila(ws, row_idx, _fila_contai(
            cuenta_proveedores, COMP_NC_COMPRA, fecha, folio, nit, 'DEV. GASTOS',
            TIPO_DEBITO, total))
        row_idx += 1

    return row_idx


def _escribir_gastos(ws, row_idx, gastos, cuenta_iva_descontable, cuenta_proveedores):
    """Escribe las líneas de facturas de gastos/servicios. Cada doc ya trae
    su 'cuenta_asignada' (resuelta por classifiers: regla, caché o IA)."""
    for g in gastos:
        if abs(g['total']) < 0.01:
            continue
        fecha  = _fmt_fecha(g['fecha'])
        folio  = g['folio']
        nit    = g['nit']
        base   = g['gravado']
        iva    = g['iva']
        total  = g['total']
        cuenta_gasto = g['cuenta_asignada']

        # Cuenta de gasto (débito)
        _escribir_fila(ws, row_idx, _fila_contai(
            cuenta_gasto, COMP_GASTO, fecha, folio, nit, 'GASTOS',
            TIPO_DEBITO, base))
        row_idx += 1

        # IVA descontable (débito) — solo si hay IVA
        if abs(iva) > 0.01:
            _escribir_fila(ws, row_idx, _fila_contai(
                cuenta_iva_descontable, COMP_GASTO, fecha, folio, nit, 'GASTOS',
                TIPO_DEBITO, iva, base=base))
            row_idx += 1

        # Proveedores / CxP (crédito) por el total
        _escribir_fila(ws, row_idx, _fila_contai(
            cuenta_proveedores, COMP_GASTO, fecha, folio, nit, 'GASTOS',
            TIPO_CREDITO, total))
        row_idx += 1

    return row_idx


def generar_plano_contai_completo(input_stream, incluir=('ventas', 'compras', 'nc_compras', 'gastos')):
    """
    Lee el reporte DIAN y genera el plano contable completo en formato Contai
    (ventas, compras, notas crédito de compras y gastos, según lo solicitado
    en 'incluir'). La clasificación compra/gasto usa classifiers/ (reglas por
    empresa -> caché -> IA), no cuentas hardcodeadas de un solo cliente.

    Retorna un dict: {'plano': BytesIO, 'empresa_nit': str, 'pendientes': [...],
    'resumen': {'por_regla': n, 'por_cache': n, 'por_ia': n, 'pendientes': n}}
    """
    wb_src = openpyxl.load_workbook(input_stream, data_only=True, read_only=True)
    (mi_nit, facturas_compras_raw, facturas_gastos_raw, notas_credito_raw,
     ventas, nc_ventas, _, _, _, _, _, _) = _leer_todo(wb_src)

    cuentas = cargar_reglas(mi_nit)
    necesita_cuentas = ('compras' in incluir or 'gastos' in incluir or 'nc_compras' in incluir)
    if necesita_cuentas:
        # Mínimo indispensable: cuenta_compras_19 y cuenta_dev_compras_19 son
        # el fallback de cualquier tarifa no configurada explícitamente
        # (ver _cuenta_compra), así que son las únicas realmente obligatorias.
        faltantes = [c for c in ('cuenta_compras_19', 'cuenta_proveedores',
                                  'cuenta_iva_descontable', 'cuenta_dev_compras_19')
                     if not cuentas.get(c)]
        if faltantes:
            raise ValueError(
                f"La empresa con NIT {mi_nit} no tiene configuradas las cuentas "
                f"{faltantes} en reglas_empresas/{mi_nit}.json. Configúralas antes "
                f"de generar compras/gastos para esta empresa (no se asumen las "
                f"cuentas de otra empresa)."
            )

    compras, gastos, nc_compras, nc_gastos, pendientes, resumen_fuentes = _clasificar_recibidos(
        mi_nit, facturas_compras_raw, facturas_gastos_raw, notas_credito_raw)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ARCHIVO PLANO'
    _escribir_headers(ws)

    row_idx = 2
    if 'ventas' in incluir:
        row_idx = _escribir_ventas(ws, row_idx, ventas, nc_ventas)
    if 'compras' in incluir:
        row_idx = _escribir_compras(ws, row_idx, compras, cuentas)
    if 'gastos' in incluir:
        row_idx = _escribir_gastos(ws, row_idx, gastos,
                                    cuentas['cuenta_iva_descontable'], cuentas['cuenta_proveedores'])
    if 'nc_compras' in incluir:
        row_idx = _escribir_nc_compras(ws, row_idx, nc_compras, cuentas)
        row_idx = _escribir_nc_gastos(ws, row_idx, nc_gastos,
                                       cuentas['cuenta_iva_descontable'], cuentas['cuenta_proveedores'])

    # Ajustar ancho de columnas
    anchos = [10, 12, 16, 12, 14, 14, 14, 6, 14, 14, 14, 12, 8]
    for col_idx, ancho in enumerate(anchos, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = ancho

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    resumen = {**resumen_fuentes, 'total_pendientes': len(pendientes)}
    return {'plano': output, 'empresa_nit': mi_nit, 'pendientes': pendientes, 'resumen': resumen}


def generar_plano_contai(input_stream, incluir=('ventas', 'compras', 'nc_compras', 'gastos')):
    """Compatibilidad: igual que generar_plano_contai_completo pero retorna
    solo el BytesIO del plano (sin info de pendientes de revisión)."""
    return generar_plano_contai_completo(input_stream, incluir=incluir)['plano']


def generar_plano_ventas_contai(input_stream):
    """Compatibilidad: genera solo el plano de ventas en formato Contai."""
    return generar_plano_contai(input_stream, incluir=('ventas',))
