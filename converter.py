import io
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# NIT del proveedor de inventario — sus facturas van a COMPRAS
NIT_PROVEEDOR_INVENTARIO = "830040709"

# ========================================
# MAPEO CONCEPTO -> CUENTA DE GASTO
# ========================================
# ========================================
# MAPEO NIT -> (CUENTA DE GASTO, CUENTA IVA, COD_IMPUESTO)
# Generado desde Balance de prueba por tercero 2025.
# cod_imp: 1 = IVA descontable bienes (24081001)
#          None = IVA descontable servicios (24081501)
# ========================================
NIT_DETAIL_MAP = {
    # ── Mapeado desde Balance de prueba por tercero 2025 - Helier (85 NITs) ────
    "1000635885":  ("52050601", "24081001", 1),
    "1001651750":  ("52050601", "24081001", 1),
    "1001660115":  ("52050601", "24081001", 1),
    "1001755653":  ("52050601", "24081001", 1),
    "1007291112":  ("52050601", "24081001", 1),
    "1007362118":  ("52050601", "24081001", 1),
    "1007425303":  ("52050601", "24081001", 1),
    "1007565931":  ("52050601", "24081001", 1),
    "1039784009":  ("52050601", "24081001", 1),
    "1040030119":  ("52050601", "24081001", 1),
    "1040030513":  ("52050601", "24081001", 1),
    "1040031271":  ("52050601", "24081001", 1),
    "1040032457":  ("52050601", "24081001", 1),
    "1040033510":  ("52050601", "24081001", 1),
    "1040034303":  ("52050601", "24081001", 1),
    "1040039802":  ("52050601", "24081001", 1),
    "1040047012":  ("52050601", "24081001", 1),
    "1040047136":  ("52050601", "24081001", 1),
    "1040050117":  ("52050601", "24081001", 1),
    "1040051841":  ("52050601", "24081001", 1),
    "1040871544":  ("52050601", "24081001", 1),
    "1054918024":  ("52050601", "24081001", 1),
    "15378117":    ("52454001", "24081501", None),
    "15387696":    ("52050601", "24081001", 1),
    "15388106":    ("52050601", "24081001", 1),
    "15388152":    ("52050601", "24081001", 1),
    "15441517":    ("52454001", "24081501", None),
    "15442412":    ("52454001", "24081501", None),
    "15445134":    ("52454001", "24081501", None),
    "21481563":    ("52454001", "24081501", None),
    "2995970":     ("52603501", "24080501", None),  # HELIER (propietario)
    "39189647":    ("52050601", "24081001", 1),
    "39455072":    ("52454001", "24081501", None),
    "41658695":    ("52454001", "24081501", None),
    "52537318":    ("52050601", "24081001", 1),
    "71111818":    ("52454001", "24081501", None),
    "71277817":    ("52050601", "24081001", 1),
    "71641441":    ("52454001", "24081501", None),
    "800088702":   ("52056901", "24081001", 1),     # EPS
    "800106884":   ("52454001", "24081501", None),
    "800197268":   ("52454001", "24080301", None),  # DIAN
    "800224808":   ("52057001", "24081001", 1),     # Pensión Porvenir
    "800227940":   ("52057001", "24081001", 1),     # Pensión Colfondos
    "800229739":   ("52057001", "24081001", 1),     # Pensión Protección
    "800242106":   ("52454001", "24081501", None),
    "805018674":   ("52454001", "24081501", None),
    "811009893":   ("52454001", "24081501", None),
    "811019540":   ("52454001", "24081501", None),
    "811036256":   ("52454001", "24081501", None),
    "830040709":   ("52454001", "24081001", 1),     # Proveedor principal
    "830048145":   ("52454001", "24081501", None),  # SIIGO SAS
    "830112317":   ("52454001", "24081501", None),
    "830114921":   ("52454001", "24081501", None),
    "860002184":   ("52454001", "24081501", None),  # Seguros
    "860011153":   ("52056801", "24081001", 1),     # ARP
    "860027404":   ("52454001", "24081501", None),  # Seguros
    "860039988":   ("52454001", "24081501", None),  # Seguros
    "860508462":   ("52454001", "24081501", None),  # Seguros
    "890900841":   ("52057201", "24081001", 1),     # COMFAMA
    "890900943":   ("52454001", "24081501", None),
    "890903407":   ("52454001", "24081501", None),
    "899999034":   ("52057801", "24081001", 1),     # SENA
    "899999239":   ("52057501", "24081001", 1),     # ICBF
    "900079535":   ("52454001", "24081501", None),
    "900092385":   ("52353501", "24081501", None),  # Teléfono
    "900219834":   ("52454001", "24081501", None),
    "900276962":   ("52454001", "24081501", None),
    "900336004":   ("52057001", "24081001", 1),     # COLPENSIONES
    "900338277":   ("52454001", "24081501", None),
    "900424941":   ("52454001", "24081501", None),
    "900569531":   ("52454001", "24081501", None),
    "900599627":   ("52454001", "24081501", None),
    "900720039":   ("52454001", "24081501", None),
    "900806620":   ("52454001", "24081501", None),
    "900933915":   ("52454001", "24081501", None),
    "900972974":   ("52454001", "24081501", None),
    "900996769":   ("52454001", "24081501", None),
    "901123242":   ("52454001", "24081501", None),
    "901151055":   ("52454001", "24081501", None),
    "901151070":   ("52454001", "24081501", None),
    "901168195":   ("52454001", "24081501", None),
    "901376585":   ("52454001", "24081501", None),
    "901406550":   ("52454001", "24081501", None),
    "901714655":   ("52454001", "24081501", None),
    "901799395":   ("52454001", "24081501", None),
    # ── Otros NITs históricos (período anterior) ─────────────────────────────
    "15378281":    ("52953501", "24081001", 1),     # Combustible
    "15378576":    ("52454001", "24081001", 1),
    "39188564":    ("52201001", "24081001", 1),
    "42995046":    ("52103001", "24081001", 1),
    "71113787":    ("52454001", "24081001", 1),
    "71117018":    ("52454001", "24081001", 1),
    "800015551":   ("52401001", "24081001", 1),
    "800180330":   ("52950501", "24081001", 1),
    "800251569":   ("52354001", "24081001", 1),
    "830059699":   ("52352001", "24081501", None),
    "860002400":   ("52304001", "24081501", None),
    "890900081":   ("52454001", "24081001", 1),
    "890903790":   ("52304001", "24081001", 1),
    "890939936":   ("52251001", "24081001", 1),
    "890981207":   ("52150501", "24081001", 1),
    "890982616":   ("52150501", "24081001", 1),
    "900201094":   ("52952001", "24081001", 1),
    "900391666":   ("52251001", "24081001", 1),
    "901095934":   ("52953501", "24081001", 1),
    "901113791":   ("52952001", "24081001", 1),
    "901304273":   ("52956001", "24081001", 1),
    "830122566":   ("52353502", "24081001", 1)
}

# Defaults para proveedores nuevos no mapeados
DEFAULT_CUENTA_GASTO = "52454001"
DEFAULT_IVA_CUENTA   = "24081001"
DEFAULT_COD_IMP      = 1

# ========================================
# ENCABEZADOS DEL PLANO SIIGO
def _n(val):
    """Convierte a float de forma segura; devuelve 0.0 si no es numérico."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _fecha(val):
    """Convierte fechas en texto (DD-MM-YYYY o YYYY-MM-DD) a datetime.
    Si ya es datetime o no se puede parsear, lo devuelve tal cual."""
    if val is None or not isinstance(val, str):
        return val
    from datetime import datetime as _dt
    s = val.strip().split(' ')[0]   # quitar hora si viene
    for fmt_str in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return _dt.strptime(s, fmt_str)
        except ValueError:
            continue
    return val


# ========================================
HEADERS = [
    'Tipo de comprobante', 'Consecutivo comprobante', 'Fecha de elaboración ',
    'Sigla moneda', 'Tasa de cambio', 'Código cuenta contable',
    'Identificación tercero', 'Sucursal', 'Código producto',
    'Código de bodega', 'Acción', 'Cantidad producto',
    'Prefijo', 'Consecutivo', 'No. cuota',
    'Fecha vencimiento', 'Código impuesto', 'Código grupo activo fijo',
    'Código activo fijo', 'Descripción', 'Código centro/subcentro de costos',
    'Débito', 'Crédito', 'Observaciones',
    'Base gravable libro compras/ventas  ', 'Base exenta libro compras/ventas',
    'Mes de cierre'
]


def make_row(tipo_comp, consec, fecha, cuenta, nit, cod_imp, descripcion, debito, credito):
    return [
        tipo_comp, consec, fecha, None, None, cuenta,
        str(nit) if nit else None, None, None, None, None, None,
        None, None, None, None, cod_imp, None, None, descripcion, None,
        debito, credito, None, None, None, None
    ]


def _write_headers(ws):
    """Escribe encabezados con estilos: fondo azul claro, texto oscuro, negrita."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    header_fill = PatternFill(fill_type='solid', fgColor='D6EAF8')   # Azul claro
    header_font = Font(bold=True, color='1a3a5c', name='Calibri', size=10)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='BFCCD8'),
        right=Side(style='thin', color='BFCCD8'),
        top=Side(style='thin', color='BFCCD8'),
        bottom=Side(style='thin', color='BFCCD8')
    )

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.row_dimensions[1].height = 20


def _apply_table_colors(ws, start_row=2, end_row=None):
    """Aplica colores alternados a las filas de datos."""
    from openpyxl.styles import PatternFill, Border, Side

    if end_row is None:
        end_row = ws.max_row

    light_gray_fill = PatternFill(fill_type='solid', fgColor='F8F9FA')   # Gris muy claro
    white_fill = PatternFill(fill_type='solid', fgColor='FFFFFF')
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    for row in range(start_row, end_row + 1):
        fill = light_gray_fill if (row - start_row) % 2 == 0 else white_fill
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border


def _insert_table(ws, sheet_name, start_row=1, end_row=None):
    """Inserta una tabla formateada en la hoja con filtros y estilos."""
    from openpyxl.worksheet.table import Table, TableStyleInfo

    if end_row is None:
        end_row = ws.max_row

    if end_row < start_row:
        return  # Sin datos para crear tabla

    # Rango de la tabla: A1:Z{end_row} (ajusta el rango según las columnas usadas)
    max_col = ws.max_column
    end_col = get_column_letter(max_col)
    ref = f"A{start_row}:{end_col}{end_row}"

    # Crear tabla con nombre único
    tab_name = f"Tabla{sheet_name.replace(' ', '')}"
    tab = Table(displayName=tab_name, ref=ref)

    # Estilo claro con líneas y filtros
    style = TableStyleInfo(
        name="TableStyleLight9",  # Estilo claro predefinido de Excel
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


# ========================================
# LECTURA DE LA HOJA REPORTE
# ========================================

def _detect_format_name(wb_src):
    """Detecta el nombre de la hoja y el formato ('reporte' o 'nuevo')."""
    if 'REPORTE' in wb_src.sheetnames:
        return 'reporte', 'REPORTE'
    for name in wb_src.sheetnames:
        if name.startswith('Rp_'):
            return 'nuevo', name
    # Fallback: primera hoja
    name = wb_src.sheetnames[0]
    return 'nuevo', name


def _parse_rows(wb_src):
    """
    Lee todas las filas de la hoja DIAN en un solo paso con iter_rows.
    Retorna (fmt, mi_nit, filas) donde filas es lista de dicts normalizados.
    """
    from collections import Counter
    fmt, sheet_name = _detect_format_name(wb_src)
    ws = wb_src[sheet_name]

    EXCLUIDOS = {'222222222222', '0', '', 'None', 'none'}

    def _es_nit_valido(v):
        """Un NIT válido es dígitos (con guiones opcionales), entre 6 y 15 caracteres."""
        if not v:
            return False
        # Remover guiones y espacios
        v_clean = str(v).replace('-', '').replace(' ', '').strip()
        return v_clean.isdigit() and 6 <= len(v_clean) <= 15
    col_receptor = 8 if fmt == 'reporte' else 12

    conteo_nit = Counter()
    raw_rows   = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if fmt == 'reporte':
            if len(row) < 13:
                continue
            tipo         = row[0]
            folio        = row[2]
            fecha        = row[4]
            nit_emisor   = str(row[5] or '').strip()
            nit_receptor = str(row[7] or '').strip()
            no_gravado   = _n(row[9])
            gravado      = _n(row[10])
            iva          = _n(row[11])
            total        = _n(row[12])
        else:
            if len(row) < 30:
                continue
            tipo         = row[0]
            folio        = row[2]
            fecha        = row[7]
            nit_emisor   = str(row[9]  or '').strip()
            nit_receptor = str(row[11] or '').strip()
            iva          = _n(row[13])
            total        = _n(row[29])
            gravado      = 0.0
            no_gravado   = 0.0

        # Contar emisor Y receptor para detectar mi_nit correctamente
        # (archivos con muchas ventas tienen más emisiones que recepciones)
        for nit_check in (nit_emisor, nit_receptor):
            if nit_check not in EXCLUIDOS and _es_nit_valido(nit_check):
                conteo_nit[nit_check] += 1

        raw_rows.append((tipo, folio, fecha, nit_emisor, nit_receptor,
                         no_gravado, gravado, iva, total))

    if not conteo_nit:
        raise ValueError("No se encontró ningún NIT receptor en el archivo.")

    mi_nit = conteo_nit.most_common(1)[0][0]

    # Segunda pasada sobre los datos ya en memoria (lista Python, muy rápida)
    filas = []
    for (tipo, folio, fecha, nit_emisor, nit_receptor,
         no_gravado, gravado, iva, total) in raw_rows:

        if nit_receptor != mi_nit:
            continue
        if folio is None:
            continue

        if fmt == 'nuevo':
            # Calcular base a partir de total e iva
            iva_r      = round(iva, 2)
            total_r    = round(total, 2)
            gravado    = round(total_r - iva_r, 2)
            no_gravado = 0.0

        filas.append({
            'tipo': str(tipo or '').lower(),
            'folio': folio, 'fecha': fecha,
            'nit_emisor': nit_emisor,
            'no_gravado': no_gravado, 'gravado': gravado,
            'iva': iva, 'total': total,
        })

    return fmt, mi_nit, filas


def read_reporte(wb_src):
    """Lee el reporte DIAN y clasifica en facturas_compras, facturas_gastos, notas_credito."""
    fmt, mi_nit, filas = _parse_rows(wb_src)

    facturas_compras = []
    facturas_gastos  = []
    notas_credito    = []

    for f in filas:
        tipo_str = f['tipo']
        doc = {
            'folio': f['folio'], 'fecha': f['fecha'], 'nit': f['nit_emisor'],
            'no_gravado': f['no_gravado'], 'gravado': f['gravado'],
            'iva': f['iva'], 'total': f['total'],
        }
        if 'nota de cr' in tipo_str:
            notas_credito.append(doc)
        elif f['nit_emisor'] == NIT_PROVEEDOR_INVENTARIO:
            facturas_compras.append(doc)
        else:
            facturas_gastos.append(doc)

    return facturas_compras, facturas_gastos, notas_credito


def _detect_format(wb_src):
    """Compatibilidad: retorna (fmt, ws) para funciones que aún lo usan."""
    fmt, sheet_name = _detect_format_name(wb_src)
    return fmt, wb_src[sheet_name]


def _detect_mi_nit(ws, fmt):
    """Compatibilidad: detecta NIT desde una hoja ya abierta."""
    from collections import Counter
    col_receptor = 7 if fmt == 'reporte' else 11  # índice 0-based en iter_rows
    EXCLUIDOS = {'222222222222', '0', '', 'None'}
    conteo = Counter()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) > col_receptor:
            nit = str(row[col_receptor] or '').strip()
            if nit not in EXCLUIDOS and nit.isdigit() and 6 <= len(nit) <= 15:
                conteo[nit] += 1
    if not conteo:
        raise ValueError("No se encontró ningún NIT receptor en el archivo.")
    return conteo.most_common(1)[0][0]


# ========================================
# GENERACIÓN DE PLANOS
# ========================================

def convert_compras(facturas, wb_dst, consec_inicio):
    ws_dst = wb_dst.create_sheet('COMPRAS')
    _write_headers(ws_dst)

    row_out = 2

    # Pre-calcular para garantizar balance: deb_mercancia + deb_iva = cred_cxp
    fact_calc = []
    for f in facturas:
        cred_cxp     = round(f['total'], 2)
        deb_iva      = round(f['iva'], 2)
        deb_mercancia = round(cred_cxp - deb_iva, 2)   # total - iva = base exacta
        fact_calc.append({**f, 'deb_mercancia': deb_mercancia,
                          'deb_iva': deb_iva, 'cred_cxp': cred_cxp})

    # Bloque 1: mercancía gravada (débito)
    consec = consec_inicio
    for f in fact_calc:
        row = make_row(9, consec, f['fecha'], '14350101', f['nit'], None,
                       f"FC: {f['folio']}", f['deb_mercancia'], 0)
        for col, val in enumerate(row, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1

    # Bloque 2: IVA descontable (débito)
    consec = consec_inicio
    for f in fact_calc:
        row = make_row(9, consec, f['fecha'], '24081001', f['nit'], 1,
                       f"FC: {f['folio']}", f['deb_iva'], 0)
        for col, val in enumerate(row, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1

    # Bloque 3: CxP proveedor (crédito = suma exacta de débitos)
    consec = consec_inicio
    for f in fact_calc:
        row = make_row(9, consec, f['fecha'], '22050501', f['nit'], None,
                       f"FC: {f['folio']}", 0, f['cred_cxp'])
        for col, val in enumerate(row, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1

    _apply_table_colors(ws_dst, start_row=2, end_row=ws_dst.max_row)
    _insert_table(ws_dst, "COMPRAS")
    return consec_inicio + len(facturas)


def convert_nc_compras(notas, wb_dst, consec_inicio):
    if not notas:
        return consec_inicio

    ws_dst = wb_dst.create_sheet('NC COMPRAS')
    _write_headers(ws_dst)

    row_out = 2

    # Pre-calcular para garantizar balance: deb_22050501 = cred_14350101 + cred_24081002
    notas_calc = []
    for n in notas:
        deb_cxp      = round(n['total'], 2)
        cred_iva     = round(n['iva'], 2)
        cred_mercancia = round(deb_cxp - cred_iva, 2)   # total - iva = base exacta
        notas_calc.append({**n, 'deb_cxp': deb_cxp,
                           'cred_mercancia': cred_mercancia, 'cred_iva': cred_iva})

    # Bloque 1: mercancía (crédito)
    consec = consec_inicio
    for n in notas_calc:
        row = make_row(10, consec, n['fecha'], '14350101', n['nit'], None,
                       f"NC: {n['folio']}", 0, n['cred_mercancia'])
        for col, val in enumerate(row, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1

    # Bloque 2: IVA descontable NC (crédito)
    consec = consec_inicio
    for n in notas_calc:
        row = make_row(10, consec, n['fecha'], '24081002', n['nit'], 1,
                       f"NC: {n['folio']}", 0, n['cred_iva'])
        for col, val in enumerate(row, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1

    # Bloque 3: CxP proveedor (débito = suma exacta de créditos)
    consec = consec_inicio
    for n in notas_calc:
        row = make_row(10, consec, n['fecha'], '22050501', n['nit'], None,
                       f"NC: {n['folio']}", n['deb_cxp'], 0)
        for col, val in enumerate(row, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1

    _apply_table_colors(ws_dst, start_row=2, end_row=ws_dst.max_row)
    _insert_table(ws_dst, "NC COMPRAS")
    return consec_inicio + len(notas)


def convert_gastos(gastos, wb_dst, consec_inicio):
    ws_dst = wb_dst.create_sheet('GASTOS')
    _write_headers(ws_dst)

    row_out = 2

    # Pre-calcular débitos redondeados para garantizar balance exacto
    gastos_calc = []
    for g in gastos:
        nit_str = str(g['nit'])
        cuenta, cuenta_iva, cod_imp = NIT_DETAIL_MAP.get(
            nit_str, (DEFAULT_CUENTA_GASTO, DEFAULT_IVA_CUENTA, DEFAULT_COD_IMP))
        deb_gasto = round(g['gravado'] + g['no_gravado'], 2)
        deb_iva   = round(g['iva'], 2)
        cred_cxp  = round(deb_gasto + deb_iva, 2)   # crédito = suma exacta de débitos
        gastos_calc.append({**g, 'cuenta': cuenta, 'cuenta_iva': cuenta_iva,
                            'cod_imp': cod_imp, 'deb_gasto': deb_gasto,
                            'deb_iva': deb_iva, 'cred_cxp': cred_cxp})

    # Bloque 1: cuenta de gasto (débito)
    consec = consec_inicio
    for g in gastos_calc:
        row = make_row(12, consec, g['fecha'], g['cuenta'], g['nit'], None,
                       f"G: {g['folio']}", g['deb_gasto'], 0)
        for col, val in enumerate(row, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1

    # Bloque 2: IVA descontable (débito)
    consec = consec_inicio
    for g in gastos_calc:
        row = make_row(12, consec, g['fecha'], g['cuenta_iva'], g['nit'], g['cod_imp'],
                       f"G: {g['folio']}", g['deb_iva'], 0)
        for col, val in enumerate(row, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1

    # Bloque 3: CxP proveedor (crédito = suma exacta de débitos)
    consec = consec_inicio
    for g in gastos_calc:
        row = make_row(12, consec, g['fecha'], '23359501', g['nit'], None,
                       f"G: {g['folio']}", 0, g['cred_cxp'])
        for col, val in enumerate(row, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1

    _apply_table_colors(ws_dst, start_row=2, end_row=ws_dst.max_row)
    _insert_table(ws_dst, "GASTOS")
    return consec_inicio + len(gastos)


# ========================================
# LECTURA DE VENTAS
# ========================================

def read_ventas(wb_src):
    """
    Lee ventas y NC ventas emitidas por el negocio en un solo paso.
    Retorna (ventas, nc_ventas).
    """
    from collections import Counter
    fmt, sheet_name = _detect_format_name(wb_src)
    ws = wb_src[sheet_name]

    EXCLUIDOS = {'222222222222', '0', '', 'None'}
    col_receptor = 7 if fmt == 'reporte' else 11  # 0-based

    # Primer paso: detectar mi_nit y recolectar filas crudas
    # Contar emisor + receptor para evitar que NITs genéricos ganen
    conteo = Counter()
    raw = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if fmt == 'reporte':
            if len(row) < 13: continue
            nit_emisor   = str(row[5] or '').strip()
            nit_receptor = str(row[7] or '').strip()
        else:
            if len(row) < 30: continue
            nit_emisor   = str(row[9]  or '').strip()
            nit_receptor = str(row[11] or '').strip()
        for nit_check in (nit_emisor, nit_receptor):
            if nit_check not in EXCLUIDOS and _es_nit_valido(nit_check):
                conteo[nit_check] += 1
        raw.append(row)

    if not conteo:
        raise ValueError("No se encontró ningún NIT en el archivo.")
    mi_nit = conteo.most_common(1)[0][0]

    ventas    = []
    nc_ventas = []

    for row in raw:
        if fmt == 'reporte':
            tipo         = str(row[0] or '').lower()
            folio        = row[2]
            fecha        = row[4]
            nit_emisor   = str(row[5] or '').strip()
            nit_cliente  = str(row[7] or '').strip()
            nombre       = ''
            iva          = _n(row[11])
            total        = _n(row[12])
        else:
            tipo         = str(row[0] or '').lower()
            folio        = row[2]
            fecha        = row[7]
            nit_emisor   = str(row[9]  or '').strip()
            nit_cliente  = str(row[11] or '').strip()
            nombre       = str(row[12] or '').strip()
            iva          = _n(row[13])
            total        = _n(row[29])

        if nit_emisor != mi_nit or folio is None:
            continue

        iva_r   = round(iva, 2)
        total_r = round(total, 2)
        gravado = round(total_r - iva_r, 2)

        doc = {'folio': folio, 'fecha': fecha, 'nit': nit_cliente,
               'nombre': nombre, 'gravado': gravado, 'iva': iva_r, 'total': total_r}

        if 'nota de cr' in tipo:
            nc_ventas.append(doc)
        elif 'factura' in tipo:
            ventas.append(doc)

    return ventas, nc_ventas


def convert_nc_ventas(nc_ventas, wb_dst, consec_inicio):
    """
    Genera la hoja NC VENTAS con el plano Siigo consolidado por NIT cliente.
    Entradas inversas a VENTAS:
      Débito  41009501 (Ingresos)
      Débito  24080501 (IVA generado)
      Crédito 13050501 (CxC clientes)
    """
    from collections import defaultdict

    ws_dst = wb_dst.create_sheet('NC VENTAS')
    _write_headers(ws_dst)
    row_out = 2

    consolidado = defaultdict(lambda: {'iva': 0.0, 'total': 0.0, 'fecha': None})
    for n in nc_ventas:
        k = str(n['nit'])
        consolidado[k]['iva']   += n['iva']
        consolidado[k]['total'] += n['total']
        if consolidado[k]['fecha'] is None:
            consolidado[k]['fecha'] = n['fecha']

    clientes = [(nit, d) for nit, d in consolidado.items() if abs(d['total']) > 0.01]

    # Pre-calcular para balance exacto: deb_ingreso + deb_iva = cred_cxc
    calc = []
    for nit, d in clientes:
        deb_iva     = round(d['iva'], 2)
        cred_cxc    = round(d['total'], 2)
        deb_ingreso = round(cred_cxc - deb_iva, 2)
        calc.append((nit, d, deb_ingreso, deb_iva, cred_cxc))

    consec = consec_inicio

    # Bloque 1: Ingresos (débito — reversa de venta)
    for nit, d, deb_ingreso, _, _ in calc:
        row = make_row(2, consec, d['fecha'], '41009501', nit, None,
                       'NC VENTA CONSOLIDADA', deb_ingreso, 0)
        for col, val in enumerate(row, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1

    # Bloque 2: IVA generado (débito — reversa)
    consec = consec_inicio
    for nit, d, _, deb_iva, _ in calc:
        row = make_row(2, consec, d['fecha'], '24080501', nit, 1,
                       'NC VENTA CONSOLIDADA', deb_iva, 0)
        for col, val in enumerate(row, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1

    # Bloque 3: CxC clientes (crédito = suma exacta de débitos)
    consec = consec_inicio
    for nit, d, _, _, cred_cxc in calc:
        row = make_row(2, consec, d['fecha'], '13050501', nit, None,
                       'NC VENTA CONSOLIDADA', 0, cred_cxc)
        for col, val in enumerate(row, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1

    _apply_table_colors(ws_dst, start_row=2, end_row=ws_dst.max_row)
    _insert_table(ws_dst, "VENTAS")
    return consec_inicio + len(calc)


def convert_ventas(ventas, nc_ventas, wb_dst, consec_inicio):
    """
    Genera la hoja VENTAS con el plano Siigo consolidado por NIT cliente.
    Solo facturas de venta (las NC van en hoja NC VENTAS separada).
    Cada NIT cliente → una fila por bloque contable (CxC, Ingresos, IVA).
    """
    from collections import defaultdict

    ws_dst = wb_dst.create_sheet('VENTAS')
    _write_headers(ws_dst)
    row_out = 2

    # Consolidar solo facturas de venta por NIT cliente (sin restar NC)
    consolidado = defaultdict(lambda: {'iva': 0.0, 'total': 0.0, 'fecha': None})
    for v in ventas:
        k = str(v['nit'])
        consolidado[k]['iva']   += v['iva']
        consolidado[k]['total'] += v['total']
        if consolidado[k]['fecha'] is None:
            consolidado[k]['fecha'] = v['fecha']

    # Filtrar NITs con total > 0
    clientes = [(nit, d) for nit, d in consolidado.items() if abs(d['total']) > 0.01]

    # Pre-calcular para balance exacto: deb_cxc = cred_ingreso + cred_iva
    clientes_calc = []
    for nit, d in clientes:
        cred_iva     = round(d['iva'], 2)
        cred_ingreso = round(d['total'], 2) - cred_iva
        clientes_calc.append((nit, d, cred_ingreso, cred_iva))

    consec = consec_inicio

    # Bloque 1: CxC clientes (débito)
    for nit, d, _, _ in clientes_calc:
        row = make_row(1, consec, d['fecha'], '13050501', nit, None,
                       'VENTA CONSOLIDADA', round(d['total'], 2), 0)
        for col, val in enumerate(row, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1

    # Bloque 2: Ingresos por ventas (crédito = total - iva)
    consec = consec_inicio
    for nit, d, cred_ingreso, _ in clientes_calc:
        row = make_row(1, consec, d['fecha'], '41009501', nit, None,
                       'VENTA CONSOLIDADA', 0, cred_ingreso)
        for col, val in enumerate(row, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1

    # Bloque 3: IVA generado (crédito)
    consec = consec_inicio
    for nit, d, _, cred_iva in clientes_calc:
        row = make_row(1, consec, d['fecha'], '24080501', nit, 1,
                       'VENTA CONSOLIDADA', 0, cred_iva)
        for col, val in enumerate(row, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1

    _apply_table_colors(ws_dst, start_row=2, end_row=ws_dst.max_row)
    _insert_table(ws_dst, "NC VENTAS")
    return consec_inicio + len(clientes)


# ========================================
# PUNTO DE ENTRADA PRINCIPAL — PLANOS SIIGO
# ========================================

def _leer_todo(wb_src):
    """
    Lee el archivo DIAN en UN SOLO PASO y devuelve todas las listas necesarias.
    Evita múltiples iteraciones sobre el mismo archivo grande.
    """
    from collections import Counter
    fmt, sheet_name = _detect_format_name(wb_src)
    ws = wb_src[sheet_name]

    EXCLUIDOS = {'222222222222', '0', '', 'None', 'none'}

    def _es_nit_valido(v):
        """Un NIT válido es dígitos (con guiones opcionales), entre 6 y 15 caracteres."""
        if not v:
            return False
        # Remover guiones y espacios
        v_clean = str(v).replace('-', '').replace(' ', '').strip()
        return v_clean.isdigit() and 6 <= len(v_clean) <= 15

    # Primera pasada: detectar mi_nit y acumular filas crudas en memoria
    # Se cuentan EMISOR + RECEPTOR para que archivos con muchas ventas
    # no confundan un NIT genérico (ej: 111111) con el NIT de la empresa.
    conteo = Counter()
    raw = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        nits_encontrados = []

        if fmt == 'reporte':
            if len(row) < 13:
                continue
            nit_emisor   = str(row[5] or '').strip()
            nit_receptor = str(row[7] or '').strip()
            nits_encontrados = [nit_emisor, nit_receptor]
        else:
            # Formato nuevo: buscar NITs en columnas esperadas primero
            if len(row) > 9:
                nits_encontrados.append(str(row[9] or '').strip())   # col J
            if len(row) > 11:
                nits_encontrados.append(str(row[11] or '').strip())  # col L
            # Si no hay NITs válidos, buscar en todo el row (flexible)
            if not any(_es_nit_valido(n) for n in nits_encontrados if n):
                nits_encontrados = [str(v or '').strip() for v in row if v]

        for nit_check in nits_encontrados:
            if nit_check not in EXCLUIDOS and _es_nit_valido(nit_check):
                conteo[nit_check] += 1
        raw.append(row)

    if not conteo:
        # Dar más contexto sobre por qué falló
        fmt_detected, sheet_detected = _detect_format_name(wb_src)
        raise ValueError(
            f"No se encontró ningún NIT válido en el archivo.\n"
            f"Detectado formato: {fmt_detected}, Hoja: {sheet_detected}\n"
            f"Asegúrate de que sea un reporte DIAN válido con NITs correctamente formateados."
        )
    mi_nit = conteo.most_common(1)[0][0]

    # Segunda pasada sobre lista Python (muy rápida, ya en memoria)
    facturas_compras = []; facturas_gastos = []; notas_credito = []
    ventas = []; nc_ventas = []
    base_ventas = 0.0; iva_ventas = 0.0
    base_nc_ventas = 0.0; iva_nc_ventas = 0.0
    fechas_venta = []

    for row in raw:
        if not row or len(row) < 2:
            continue

        if fmt == 'reporte':
            if len(row) < 13:
                continue
            tipo         = str(row[0] or '').lower()
            folio        = row[2]
            fecha        = _fecha(row[4])
            nit_emisor   = str(row[5] or '').strip()
            nit_receptor = str(row[7] or '').strip()
            no_gravado   = _n(row[9])
            gravado_raw  = _n(row[10])
            iva          = _n(row[11])
            total        = _n(row[12])
            nit_cliente  = nit_receptor
            nombre       = ''
            grupo        = str(row[14] or '').strip() if len(row) > 14 else ''
        else:
            if len(row) < 14:
                continue
            tipo         = str(row[0] or '').lower()
            folio        = row[2] if len(row) > 2 else None
            fecha        = _fecha(row[7]) if len(row) > 7 else None
            nit_emisor   = str(row[9]  or '').strip() if len(row) > 9 else ''
            nit_receptor = str(row[11] or '').strip() if len(row) > 11 else ''
            nombre       = str(row[12] or '').strip() if len(row) > 12 else ''
            iva          = _n(row[13]) if len(row) > 13 else 0.0
            total        = _n(row[29]) if len(row) > 29 else 0.0
            no_gravado   = 0.0
            gravado_raw  = 0.0
            nit_cliente  = nit_receptor
            grupo        = str(row[31] or '').strip() if len(row) > 31 else ''

        iva_r   = round(iva,   2)
        total_r = round(total, 2)
        base_r  = round(total_r - iva_r, 2)   # base exacta para balancear

        es_nc      = 'nota de cr' in tipo
        es_factura = 'factura'    in tipo

        # Clasificación Emitido/Recibido: la columna "Grupo" de la DIAN es
        # la fuente oficial. Si no existe, se infiere por NIT. Esto evita
        # que autofacturas (emisor == receptor == mi_nit) se dupliquen
        # como gasto Y venta a la vez.
        if grupo in ('Emitido', 'Recibido'):
            es_recibido = (grupo == 'Recibido')
            es_emitido  = (grupo == 'Emitido')
        else:
            es_recibido = (nit_receptor == mi_nit and nit_emisor != mi_nit)
            es_emitido  = (nit_emisor == mi_nit)

        # ── Compras / gastos (recibidos por mi_nit) ──
        if es_recibido and nit_receptor == mi_nit and folio is not None:
            # Usar base_r (total-iva) siempre para garantizar balance exacto.
            # gravado_raw puede venir de fórmulas Excel que leen como 0.
            grav = base_r; nograv = 0.0

            doc = {'folio': folio, 'fecha': fecha, 'nit': nit_emisor,
                   'no_gravado': nograv, 'gravado': grav, 'iva': iva_r, 'total': total_r}

            if es_nc:
                notas_credito.append(doc)
            elif nit_emisor == NIT_PROVEEDOR_INVENTARIO:
                facturas_compras.append(doc)
            else:
                facturas_gastos.append(doc)

        # ── Ventas (emitidas por mi_nit) ──
        if es_emitido and nit_emisor == mi_nit and folio is not None:
            vdoc = {'folio': folio, 'fecha': fecha, 'nit': nit_cliente,
                    'nombre': nombre, 'gravado': base_r, 'iva': iva_r, 'total': total_r}
            if es_nc:
                nc_ventas.append(vdoc)
                base_nc_ventas += base_r; iva_nc_ventas += iva_r
                if fecha: fechas_venta.append(fecha)
            elif es_factura:
                ventas.append(vdoc)
                base_ventas += base_r; iva_ventas += iva_r
                if fecha: fechas_venta.append(fecha)

    fecha_ini = min(fechas_venta) if fechas_venta else None
    fecha_fin = max(fechas_venta) if fechas_venta else None

    return (mi_nit,
            facturas_compras, facturas_gastos, notas_credito,
            ventas, nc_ventas,
            base_ventas, iva_ventas, base_nc_ventas, iva_nc_ventas,
            fecha_ini, fecha_fin)


# ========================================
# NÓMINA — PLANO SIIGO
# ========================================
# Estructura del plano de nómina Siigo:
#  Tipo comprobante: 13 (Nómina)
#  Por cada empleado, un consecutivo con estas entradas:
#    DÉBITOS  (gastos de nómina)          CRÉDITOS (contrapartes)
#    52050601 Sueldos                      25xxxx  Prestaciones / CxP empleado
#    52051801 Comisiones                   23700501 CxP EPS (aporte empresa)
#    52052701 Auxilio transporte           23700601 CxP ARL
#    52053001 Cesantías                    23701001 CxP Caja/SENA/ICBF
#    52053301 Int. cesantías               23803001 CxP Pensión (aporte empresa)
#    52053601 Prima
#    52053901 Vacaciones
#    52056801 ARL
#    52056901 EPS
#    52057001 Pensión
#    52057201 Caja compensación
#    52057501 ICBF
#    52057801 SENA

# Mapa empleado NIT → cuentas desde balance
NOMINA_EMPLEADOS_BALANCE = None  # se carga desde el archivo de balance

def _cargar_nomina_balance(balance_path):
    """Lee el balance y extrae cuentas de nómina por empleado."""
    import openpyxl as _opx
    from collections import defaultdict as _dd
    wb = _opx.load_workbook(balance_path, data_only=True, read_only=True)
    ws = wb['Sheet1']
    CTAS_NOM = {
        '52050601','52051801','52052701','52053001','52053301',
        '52053601','52053901','52056801','52056901','52057001',
        '52057201','52057501','52057801',
    }
    emp = _dd(lambda: {'nombre': '', 'cuentas': {}})
    for row in ws.iter_rows(min_row=9, values_only=True):
        if row[0] != 'Auxiliar' or not row[4]: continue
        cuenta = str(row[2] or '').strip()
        nit    = str(row[4] or '').strip()
        nombre = str(row[6] or '').strip()
        deb    = float(row[8] or 0)
        if cuenta in CTAS_NOM and deb > 0:
            emp[nit]['nombre'] = nombre
            emp[nit]['cuentas'][cuenta] = deb
    return dict(emp)


def convert_nomina(empleados_data, wb_dst, consec_inicio, fecha_nomina=None):
    """
    Genera la hoja NOMINA con el plano Siigo.
    empleados_data: dict {nit: {nombre, cuentas: {cuenta: monto}}}
    Un consecutivo por empleado. Débitos gastos, crédito 25101001 (CxP empleado).
    Para prestaciones sociales (cesantías, prima, vacaciones) → cuenta 25101xxx.
    Para aportes parafiscales (EPS/ARL/Pensión/Caja) → cuentas 237xx / 238xx.
    """
    import datetime

    ws_dst = wb_dst.create_sheet('NOMINA')
    _write_headers(ws_dst)
    row_out = 2

    # Cuentas que van a contrapartes especiales (no CxP empleado)
    CONTRAP = {
        '52056901': '23700501',  # EPS → CxP EPS
        '52056801': '23700601',  # ARL → CxP ARL
        '52057201': '23701001',  # Caja compensación → CxP Caja/SENA/ICBF
        '52057501': '23701001',  # ICBF → CxP Caja/SENA/ICBF
        '52057801': '23701001',  # SENA → CxP Caja/SENA/ICBF
        '52057001': '23803001',  # Pensión → CxP Fondos de pensión
        '52053001': '25101001',  # Cesantías → CxP Cesantías
        '52053301': '25101002',  # Int. cesantías → CxP Int. cesantías
        '52053901': '25101003',  # Vacaciones → CxP Vacaciones
        '52053601': '25101004',  # Prima → CxP Prima de servicios
    }
    # Cuentas de salario/transporte/comisiones → CxP empleado directo
    CXP_EMPLEADO = '25050101'

    if fecha_nomina is None:
        fecha_nomina = datetime.date.today()

    consec = consec_inicio

    for nit, emp in sorted(empleados_data.items()):
        cuentas = emp.get('cuentas', {})
        if not cuentas:
            continue

        # Pre-calcular total débito y grupos de crédito
        creditos = {}  # cta_credito → monto
        total_debito = 0.0
        for cta, monto in cuentas.items():
            monto_r = round(monto, 2)
            total_debito += monto_r
            cta_cred = CONTRAP.get(cta, CXP_EMPLEADO)
            creditos[cta_cred] = round(creditos.get(cta_cred, 0.0) + monto_r, 2)

        desc = f"NOM: {emp.get('nombre', nit)[:30]}"

        # Filas de débito (una por cuenta de gasto)
        for cta, monto in sorted(cuentas.items()):
            row = make_row(13, consec, fecha_nomina, cta, nit, None,
                           desc, round(monto, 2), 0)
            for col, val in enumerate(row, 1):
                ws_dst.cell(row=row_out, column=col, value=val)
            row_out += 1

        # Filas de crédito (una por cuenta contraparte)
        for cta_cred, monto_cred in sorted(creditos.items()):
            row = make_row(13, consec, fecha_nomina, cta_cred, nit, None,
                           desc, 0, monto_cred)
            for col, val in enumerate(row, 1):
                ws_dst.cell(row=row_out, column=col, value=val)
            row_out += 1

        consec += 1

    _apply_table_colors(ws_dst, start_row=2, end_row=ws_dst.max_row)
    _insert_table(ws_dst, "NOMINA")
    return consec


def process_file(input_stream, consec_compras=371, consec_nc=271, consec_gastos=798, consec_ventas=1, consec_nc_ventas=1):
    """Lee el archivo DIAN y genera los planos Siigo (compras, NC, gastos y ventas)."""
    wb_src = openpyxl.load_workbook(input_stream, data_only=True, read_only=True)

    (mi_nit, facturas_compras, facturas_gastos, notas_credito,
     ventas, nc_ventas, _, _, _, _, _, _) = _leer_todo(wb_src)

    wb_dst = Workbook()
    if wb_dst.active:
        wb_dst.remove(wb_dst.active)

    convert_compras(facturas_compras, wb_dst, consec_compras)
    convert_nc_compras(notas_credito, wb_dst, consec_nc)
    convert_gastos(facturas_gastos, wb_dst, consec_gastos)
    convert_ventas(ventas, nc_ventas, wb_dst, consec_ventas)
    convert_nc_ventas(nc_ventas, wb_dst, consec_nc_ventas)

    output_stream = io.BytesIO()
    wb_dst.save(output_stream)
    output_stream.seek(0)
    return output_stream


# ========================================
# LIQUIDACIÓN DE IVA — LECTURA VENTAS
# ========================================

def read_ventas_iva(wb_src):
    """
    Lee ventas y NC ventas emitidas por el negocio en un solo paso.
    Retorna (base_ventas, iva_ventas, base_nc_ventas, iva_nc_ventas, fecha_inicio, fecha_fin).
    """
    # Reutiliza _parse_rows que ya hizo el trabajo pesado
    fmt, mi_nit, filas = _parse_rows(wb_src)

    base_ventas    = 0.0;  iva_ventas     = 0.0
    base_nc_ventas = 0.0;  iva_nc_ventas  = 0.0
    fechas = []

    for f in filas:
        # filas contiene documentos recibidos por mi_nit (compras/gastos)
        # Para ventas necesitamos documentos EMITIDOS por mi_nit
        pass

    # Segunda lectura enfocada en emisor = mi_nit
    fmt2, sheet_name = _detect_format_name(wb_src)
    ws2 = wb_src[sheet_name]
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if fmt2 == 'reporte':
            if len(row) < 13: continue
            tipo       = str(row[0] or '').lower()
            nit_emisor = str(row[5] or '').strip()
            fecha      = row[4]
            iva        = _n(row[11])
            total      = _n(row[12])
        else:
            if len(row) < 30: continue
            tipo       = str(row[0] or '').lower()
            nit_emisor = str(row[9] or '').strip()
            fecha      = row[7]
            iva        = _n(row[13])
            total      = _n(row[29])

        if nit_emisor != mi_nit:
            continue
        gravado = round(total - iva, 2)
        if fecha:
            fechas.append(fecha)
        if 'nota de cr' in tipo:
            base_nc_ventas += gravado
            iva_nc_ventas  += iva
        elif 'factura' in tipo:
            base_ventas += gravado
            iva_ventas  += iva

    fecha_inicio = min(fechas) if fechas else None
    fecha_fin    = max(fechas) if fechas else None
    return base_ventas, iva_ventas, base_nc_ventas, iva_nc_ventas, fecha_inicio, fecha_fin


# ========================================
# LIQUIDACIÓN DE IVA — GENERACIÓN ARCHIVO
# ========================================

def _fmt_date(d):
    if d is None:
        return ''
    if hasattr(d, 'strftime'):
        return d.strftime('%d/%m/%Y')
    return str(d)


def _make_border():
    thin = Side(style='thin', color='BFBFBF')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_cell(ws, row, col, value, bold=False, fill_hex=None, number_format=None,
                align='left', font_color='000000', font_size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=font_color, size=font_size)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=False)
    cell.border = _make_border()
    if fill_hex:
        cell.fill = PatternFill(fill_type='solid', fgColor=fill_hex)
    if number_format:
        cell.number_format = number_format
    return cell


def generate_liquidacion_iva_file(facturas_compras, facturas_gastos, notas_credito,
                                  base_ventas, iva_ventas,
                                  base_nc_ventas, iva_nc_ventas,
                                  fecha_inicio, fecha_fin, mi_nit=''):
    """Genera un Excel con una hoja RESUMEN de liquidación de IVA."""

    # ── Cálculos ────────────────────────────────────────────────────────────
    base_compras_inv = sum(f['gravado'] for f in facturas_compras)
    iva_compras_inv  = sum(f['iva']     for f in facturas_compras)

    base_gastos_bienes    = 0.0;  iva_gastos_bienes    = 0.0
    base_gastos_servicios = 0.0;  iva_gastos_servicios = 0.0
    for g in facturas_gastos:
        _, cuenta_iva, _ = NIT_DETAIL_MAP.get(
            str(g['nit']), (DEFAULT_CUENTA_GASTO, DEFAULT_IVA_CUENTA, DEFAULT_COD_IMP))
        if cuenta_iva == '24081501':
            base_gastos_servicios += g['gravado']
            iva_gastos_servicios  += g['iva']
        else:
            base_gastos_bienes += g['gravado']
            iva_gastos_bienes  += g['iva']

    base_nc_compras = sum(n['gravado'] for n in notas_credito)
    iva_nc_compras  = sum(n['iva']     for n in notas_credito)

    base_generado_neto    = base_ventas - base_nc_ventas
    iva_generado_neto     = iva_ventas  - iva_nc_ventas

    base_descontable_neto = (base_compras_inv + base_gastos_bienes
                             + base_gastos_servicios - base_nc_compras)
    iva_descontable_neto  = (iva_compras_inv + iva_gastos_bienes
                             + iva_gastos_servicios - iva_nc_compras)

    saldo = iva_generado_neto - iva_descontable_neto  # > 0: a pagar

    periodo = (f"{_fmt_date(fecha_inicio)} – {_fmt_date(fecha_fin)}"
               if fecha_inicio else "N/D")

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('RESUMEN')

    C_AZUL        = '1E3A5F'
    C_AZUL_MED    = '2E5FA3'
    C_VERDE       = '1A6B3C'
    C_ROJO        = '8B0000'
    C_AZUL_CLARO  = 'DCE6F1'
    C_VERDE_CLARO = 'E2EFDA'
    C_ROJO_CLARO  = 'FCE4D6'
    C_BLANCO      = 'FFFFFF'
    FMT_PESO      = '"$"#,##0.00'

    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22

    # ── Título ──
    ws.merge_cells('A1:D1')
    t = ws['A1']
    t.value = 'LIQUIDACIÓN DE IVA — PERÍODO FISCAL'
    t.font = Font(bold=True, color='FFFFFF', size=14)
    t.fill = PatternFill(fill_type='solid', fgColor=C_AZUL)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:D2')
    p = ws['A2']
    p.value = f'Período: {periodo}   |   NIT: {mi_nit}'
    p.font = Font(italic=True, color='FFFFFF', size=10)
    p.fill = PatternFill(fill_type='solid', fgColor=C_AZUL_MED)
    p.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    # ── Helpers ──
    def section_header(row, title, color, ncols=4):
        ws.merge_cells(f'A{row}:{get_column_letter(ncols)}{row}')
        c = ws[f'A{row}']
        c.value = title
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = PatternFill(fill_type='solid', fgColor=color)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c.border = _make_border()
        ws.row_dimensions[row].height = 22

    def col_headers(row):
        for col, txt in enumerate(['Concepto', 'Cuenta', 'Base Gravable', 'IVA (19%)'], 1):
            _apply_cell(ws, row, col, txt, bold=True, fill_hex='D9D9D9',
                        align='center', font_size=9)
        ws.row_dimensions[row].height = 16

    def data_row(row, concepto, cuenta, base, iva_val, fill=None,
                 color_val=None, bold=False, indent=1):
        label = ('   ' * indent) + concepto
        _apply_cell(ws, row, 1, label, bold=bold, fill_hex=fill, font_size=10)
        _apply_cell(ws, row, 2, cuenta, fill_hex=fill, align='center',
                    font_size=9, font_color='555555')
        _apply_cell(ws, row, 3, base, bold=bold, fill_hex=fill,
                    number_format=FMT_PESO, align='right',
                    font_color=color_val or '000000')
        _apply_cell(ws, row, 4, iva_val, bold=bold, fill_hex=fill,
                    number_format=FMT_PESO, align='right',
                    font_color=color_val or '000000')
        ws.row_dimensions[row].height = 17

    def subtotal_row(row, label, base, iva_val, fill, color_val='000000'):
        ws.merge_cells(f'A{row}:B{row}')
        c = ws[f'A{row}']
        c.value = label
        c.font = Font(bold=True, size=10)
        c.fill = PatternFill(fill_type='solid', fgColor=fill)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border = _make_border()
        ws[f'B{row}'].border = _make_border()
        _apply_cell(ws, row, 3, base, bold=True, fill_hex=fill,
                    number_format=FMT_PESO, align='right', font_color=color_val)
        _apply_cell(ws, row, 4, iva_val, bold=True, fill_hex=fill,
                    number_format=FMT_PESO, align='right', font_color=color_val)
        ws.row_dimensions[row].height = 18

    r = 4

    # ── IVA GENERADO ──
    section_header(r, '▶  IVA GENERADO (Ventas)', C_AZUL_MED); r += 1
    col_headers(r); r += 1
    data_row(r, 'Ventas gravadas', '24080501', base_ventas, iva_ventas,
             fill=C_BLANCO); r += 1
    data_row(r, 'Menos: notas crédito ventas', '24080501',
             -base_nc_ventas, -iva_nc_ventas,
             fill=C_BLANCO,
             color_val='8B0000' if iva_nc_ventas > 0 else None); r += 1
    subtotal_row(r, 'IVA GENERADO NETO',
                 base_generado_neto, iva_generado_neto, C_AZUL_CLARO); r += 1

    r += 1

    # ── IVA DESCONTABLE ──
    section_header(r, '▶  IVA DESCONTABLE (Compras y Gastos)', C_VERDE); r += 1
    col_headers(r); r += 1
    data_row(r, 'Compras inventario — bienes (19%)', '24081001',
             base_compras_inv, iva_compras_inv, fill=C_BLANCO); r += 1
    data_row(r, 'Gastos — bienes (19%)', '24081001',
             base_gastos_bienes, iva_gastos_bienes, fill=C_BLANCO); r += 1
    data_row(r, 'Gastos — servicios (19%)', '24081501',
             base_gastos_servicios, iva_gastos_servicios, fill=C_BLANCO); r += 1
    data_row(r, 'Menos: notas crédito compras', '24081002',
             -base_nc_compras, -iva_nc_compras,
             fill=C_BLANCO,
             color_val='8B0000' if iva_nc_compras > 0 else None); r += 1
    subtotal_row(r, 'IVA DESCONTABLE NETO',
                 base_descontable_neto, iva_descontable_neto, C_VERDE_CLARO); r += 1

    r += 1

    # ── SALDO ──
    es_a_pagar   = saldo >= 0
    saldo_label  = 'IVA A PAGAR' if es_a_pagar else 'IVA A FAVOR (saldo próximo período)'
    saldo_fill   = C_ROJO_CLARO  if es_a_pagar else C_VERDE_CLARO
    saldo_color  = C_ROJO        if es_a_pagar else C_VERDE

    section_header(r, '▶  SALDO IVA', C_ROJO if es_a_pagar else C_VERDE); r += 1

    # Fila de saldo — solo columna IVA, base en blanco
    ws.merge_cells(f'A{r}:B{r}')
    c = ws[f'A{r}']
    c.value = saldo_label
    c.font = Font(bold=True, size=11, color=saldo_color)
    c.fill = PatternFill(fill_type='solid', fgColor=saldo_fill)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = _make_border()
    ws[f'B{r}'].border = _make_border()
    _apply_cell(ws, r, 3, None, fill_hex=saldo_fill)
    _apply_cell(ws, r, 4, abs(saldo), bold=True, fill_hex=saldo_fill,
                number_format=FMT_PESO, align='right', font_color=saldo_color,
                font_size=11)
    ws.row_dimensions[r].height = 22; r += 1

    r += 1
    ws.merge_cells(f'A{r}:D{r}')
    nota = ws[f'A{r}']
    nota.value = ('* Valores calculados desde la hoja REPORTE del archivo DIAN. '
                  'Verifique contra su declaración de IVA (Formulario 300).')
    nota.font = Font(italic=True, size=8, color='777777')
    nota.alignment = Alignment(horizontal='left', vertical='center')

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def process_liquidacion_iva(input_stream):
    """Lee el REPORTE DIAN y genera el archivo de liquidación de IVA."""
    wb_src = openpyxl.load_workbook(input_stream, data_only=True, read_only=True)

    (mi_nit, facturas_compras, facturas_gastos, notas_credito,
     _, _, base_ventas, iva_ventas, base_nc_ventas, iva_nc_ventas,
     fecha_ini, fecha_fin) = _leer_todo(wb_src)

    return generate_liquidacion_iva_file(
        facturas_compras, facturas_gastos, notas_credito,
        base_ventas, iva_ventas,
        base_nc_ventas, iva_nc_ventas,
        fecha_ini, fecha_fin,
        mi_nit=mi_nit,
    )


# ============================================================
# BALANCE DE PRUEBA POR TERCERO — GENERADOR DE REPORTE
# ============================================================

# Nombres PUC para construir el balance desde planos
PUC_NOMBRES = {
    '1': 'Activo', '2': 'Pasivo', '3': 'Patrimonio', '4': 'Ingresos',
    '5': 'Gastos', '6': 'Costos de ventas',
    '13': 'Deudores', '14': 'Inventarios', '22': 'Proveedores',
    '23': 'Cuentas por pagar', '24': 'Impuestos, gravámenes y tasas',
    '25': 'Obligaciones laborales', '41': 'Operacionales',
    '51': 'Operacionales de administración', '52': 'Operacionales de ventas',
    '1305': 'Clientes', '1435': 'Mercancías no fabricadas por la empresa',
    '2205': 'Proveedores nacionales', '2335': 'Costos y gastos por pagar',
    '2408': 'Impuesto sobre las ventas por pagar', '4100': 'Comercio al por mayor y al por menor',
    '130505': 'Clientes nacionales', '143501': 'Mercancías no fabricadas',
    '220505': 'Proveedores nacionales', '233595': 'Otros costos y gastos por pagar',
    '240805': 'IVA generado', '240810': 'IVA descontable', '240815': 'IVA descontable servicios',
    '410095': 'Comercio al por mayor y al por menor',
    '13050501': 'Clientes nacionales',
    '14350101': 'Mercancías no fabricadas por la empresa',
    '22050501': 'Proveedores nacionales',
    '23359501': 'Otros costos y gastos por pagar',
    '24080501': 'IVA generado',
    '24081001': 'IVA descontable bienes',
    '24081002': 'IVA descontable devoluciones',
    '24081501': 'IVA descontable servicios',
    '41009501': 'Comercio al por mayor y al por menor',
}


def _balance_desde_planos(wb_src):
    """
    Construye las filas del balance de prueba por tercero a partir de un
    libro de PLANOS Siigo (hojas con columnas: cuenta=6, nit=7, deb=22, cred=23).
    Retorna (rows_data, periodo).
    """
    from collections import defaultdict

    aux = defaultdict(lambda: [0.0, 0.0])   # (cuenta, nit) -> [deb, cred]
    fechas = []

    for sheet in wb_src.sheetnames:
        ws = wb_src[sheet]
        if str(ws.cell(1, 1).value or '').strip() != 'Tipo de comprobante':
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 23:
                continue
            cuenta = str(row[5] or '').strip()
            nit    = str(row[6] or '').strip()
            deb    = _n(row[21])
            cred   = _n(row[22])
            if not cuenta or not cuenta.isdigit():
                continue
            aux[(cuenta, nit)][0] += deb
            aux[(cuenta, nit)][1] += cred
            if row[2]:
                fechas.append(row[2])

    if not aux:
        raise ValueError(
            'El archivo no contiene un Balance de Siigo ni hojas de planos. '
            'Sube el Balance de prueba por tercero exportado de Siigo o un archivo de planos.')

    # Jerarquía: Clase(1) > Grupo(2) > Cuenta(4) > Subcuenta(6) > Auxiliar(8)+NIT
    niveles = {}   # prefijo -> [deb, cred]
    for (cuenta, _nit), (d, c) in aux.items():
        for ln in (1, 2, 4, 6, 8):
            pref = cuenta[:ln]
            niveles.setdefault(pref, [0.0, 0.0])
            niveles[pref][0] += d
            niveles[pref][1] += c

    rows_data = []
    def _row(nivel, codigo, nombre, ident, tercero, d, c):
        rows_data.append({
            'nivel': nivel, 'trans': 'Sí' if nivel == 'Auxiliar' else 'No',
            'codigo': codigo, 'nombre': nombre, 'ident': ident,
            'tercero': tercero, 's_ini': 0.0, 's_deb': round(d, 2),
            's_cred': round(c, 2), 's_fin': round(d - c, 2),
        })

    NIVEL_NOMBRE = {1: 'Clase', 2: 'Grupo', 4: 'Cuenta', 6: 'Subcuenta'}
    cuentas8 = sorted({cta for (cta, _n2) in aux})
    emitidos = set()
    for cta in cuentas8:
        for ln in (1, 2, 4, 6):
            pref = cta[:ln]
            if pref not in emitidos:
                emitidos.add(pref)
                d, c = niveles[pref]
                _row(NIVEL_NOMBRE[ln], pref, PUC_NOMBRES.get(pref, ''), '', '', d, c)
        # Auxiliares de esta cuenta (un renglón por tercero)
        terceros = sorted([(nit, v) for (c2, nit), v in aux.items() if c2 == cta])
        nombre8 = PUC_NOMBRES.get(cta, '')
        for nit, (d, c) in terceros:
            _row('Auxiliar', cta, nombre8, nit, '', d, c)

    fechas_dt = [f for f in fechas if hasattr(f, 'strftime')]
    if fechas_dt:
        periodo = f"De {min(fechas_dt).strftime('%d-%m-%Y')} a {max(fechas_dt).strftime('%d-%m-%Y')}"
    else:
        periodo = ''
    return rows_data, periodo


def generate_balance_prueba(input_stream):
    """
    Genera un Balance de prueba por tercero formateado.
    Acepta dos tipos de archivo:
      1. Balance de prueba por tercero exportado de Siigo (lo reformatea).
      2. Archivo de PLANOS Siigo o reporte DIAN (construye el balance
         agregando los movimientos débito/crédito por cuenta y tercero).
    """
    data = input_stream.read()
    wb_src = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws_src = wb_src.active

    encabezado = ' '.join(str(ws_src.cell(r, 1).value or '') for r in (1, 2, 3)).lower()
    es_balance_siigo = 'balance de prueba' in encabezado

    if es_balance_siigo:
        # ── Reformatear balance exportado de Siigo ───────────────────
        empresa = ws_src.cell(3, 1).value or ''
        nit_emp = ws_src.cell(4, 1).value or ''
        periodo = ws_src.cell(5, 1).value or ''

        rows_data = []
        for r in range(9, ws_src.max_row + 1):
            nivel  = ws_src.cell(r, 1).value
            if nivel is None:
                continue
            nivel = str(nivel).strip()
            if nivel.startswith('Total') or nivel.startswith('Procesado'):
                continue
            trans   = ws_src.cell(r, 2).value
            codigo  = str(ws_src.cell(r, 3).value or '').strip()
            nombre  = str(ws_src.cell(r, 4).value or '').strip()
            ident   = str(ws_src.cell(r, 5).value or '').strip()
            tercero = str(ws_src.cell(r, 7).value or '').strip()
            s_ini   = ws_src.cell(r, 8).value
            s_deb   = ws_src.cell(r, 9).value
            s_cred  = ws_src.cell(r, 10).value
            s_fin   = ws_src.cell(r, 11).value
            rows_data.append({
                'nivel': nivel, 'trans': trans, 'codigo': codigo,
                'nombre': nombre, 'ident': ident, 'tercero': tercero,
                's_ini': _n(s_ini), 's_deb': _n(s_deb),
                's_cred': _n(s_cred), 's_fin': _n(s_fin),
            })
    else:
        # ── Construir balance desde planos o reporte DIAN ────────────
        primera_celda = str(ws_src.cell(1, 1).value or '').strip()
        if primera_celda == 'Tipo de comprobante':
            wb_planos = wb_src               # ya es un archivo de planos
        else:
            # Es un reporte DIAN: generar los planos internamente
            planos_stream = process_file(io.BytesIO(data))
            wb_planos = openpyxl.load_workbook(planos_stream, data_only=True)

        rows_data, periodo = _balance_desde_planos(wb_planos)
        empresa = 'Balance generado desde movimientos'
        nit_emp = ''

    # ── Crear libro destino ───────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = 'Balance Prueba'

    # Colores por nivel
    COLOR = {
        'Clase':     ('1a3a5c', 'FFFFFF'),   # azul oscuro / blanco
        'Grupo':     ('1e6091', 'FFFFFF'),   # azul medio / blanco
        'Cuenta':    ('2980b9', 'FFFFFF'),   # azul claro / blanco
        'Subcuenta': ('aed6f1', '1a3a5c'),   # celeste / azul oscuro
        'Auxiliar':  ('FFFFFF', '1a3a5c'),   # blanco / azul oscuro
    }

    ALTO = {
        'Clase': 20, 'Grupo': 18, 'Cuenta': 16,
        'Subcuenta': 15, 'Auxiliar': 14,
    }

    def _border():
        side = Side(style='thin', color='BFCCD8')
        return Border(left=side, right=side, top=side, bottom=side)

    def _cell(row, col, value, bold=False, bg=None, fg='000000',
              align='left', fmt=None, size=10, wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, size=size, color=fg,
                      name='Calibri')
        if bg:
            c.fill = PatternFill(fill_type='solid', fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical='center',
                                wrap_text=wrap)
        c.border = _border()
        if fmt:
            c.number_format = fmt
        return c

    FMT_NUM = '#,##0.00'

    # ── Fila 1: título ────────────────────────────────────────────────
    ws.merge_cells('A1:J1')
    t = ws['A1']
    t.value = 'BALANCE DE PRUEBA POR TERCERO'
    t.font = Font(bold=True, size=14, color='1a3a5c', name='Calibri')
    t.alignment = Alignment(horizontal='center', vertical='center')
    t.fill = PatternFill(fill_type='solid', fgColor='d6eaf8')
    ws.row_dimensions[1].height = 26

    # ── Filas 2-3: empresa / periodo ─────────────────────────────────
    ws.merge_cells('A2:J2')
    e = ws['A2']
    e.value = f'{empresa}  —  NIT: {nit_emp}' if nit_emp else str(empresa)
    e.font = Font(bold=True, size=11, color='1a3a5c', name='Calibri')
    e.alignment = Alignment(horizontal='center', vertical='center')
    e.fill = PatternFill(fill_type='solid', fgColor='eaf4fb')
    ws.row_dimensions[2].height = 18

    ws.merge_cells('A3:J3')
    p = ws['A3']
    p.value = str(periodo).strip()
    p.font = Font(italic=True, size=10, color='555555', name='Calibri')
    p.alignment = Alignment(horizontal='center', vertical='center')
    p.fill = PatternFill(fill_type='solid', fgColor='f2f9fd')
    ws.row_dimensions[3].height = 16

    # ── Fila 4: encabezados de columna ───────────────────────────────
    HEADERS = [
        'Nivel', 'Código', 'Nombre cuenta',
        'NIT / Identificación', 'Nombre tercero',
        'Saldo inicial', 'Mov. débito', 'Mov. crédito', 'Saldo final', 'T'
    ]
    for col, h in enumerate(HEADERS, 1):
        _cell(4, col, h, bold=True, bg='1a3a5c', fg='FFFFFF',
              align='center', size=9)
    ws.row_dimensions[4].height = 18

    # ── Anchos de columna ────────────────────────────────────────────
    COL_W = [12, 12, 40, 16, 38, 16, 16, 16, 16, 5]
    for i, w in enumerate(COL_W, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Datos ─────────────────────────────────────────────────────────
    r = 5
    for d in rows_data:
        nivel = d['nivel']
        bg, fg = COLOR.get(nivel, ('FFFFFF', '000000'))
        is_aux = (nivel == 'Auxiliar')
        bold = not is_aux

        _cell(r, 1,  nivel,      bold=bold, bg=bg, fg=fg, align='center', size=9)
        _cell(r, 2,  d['codigo'], bold=bold, bg=bg, fg=fg, size=9)
        _cell(r, 3,  d['nombre'], bold=bold, bg=bg, fg=fg, size=9)
        _cell(r, 4,  d['ident'] if is_aux else '', bold=False, bg=bg, fg=fg, size=9)
        _cell(r, 5,  d['tercero'] if is_aux else '', bold=False, bg=bg, fg=fg, size=9)
        _cell(r, 6,  d['s_ini'],  bold=bold, bg=bg, fg=fg, align='right', fmt=FMT_NUM, size=9)
        _cell(r, 7,  d['s_deb'],  bold=bold, bg=bg, fg=fg, align='right', fmt=FMT_NUM, size=9)
        _cell(r, 8,  d['s_cred'], bold=bold, bg=bg, fg=fg, align='right', fmt=FMT_NUM, size=9)
        _cell(r, 9,  d['s_fin'],  bold=bold, bg=bg, fg=fg, align='right', fmt=FMT_NUM, size=9)
        _cell(r, 10, 'Sí' if d['trans'] == 'Sí' else 'No',
              bold=False, bg=bg, fg=fg, align='center', size=8)
        ws.row_dimensions[r].height = ALTO.get(nivel, 14)
        r += 1

    # ── Fila total general ────────────────────────────────────────────
    total_deb  = sum(d['s_deb']  for d in rows_data if d['nivel'] == 'Clase')
    total_cred = sum(d['s_cred'] for d in rows_data if d['nivel'] == 'Clase')
    total_ini  = sum(d['s_ini']  for d in rows_data if d['nivel'] == 'Clase')
    ws.merge_cells(f'A{r}:E{r}')
    t = ws[f'A{r}']
    t.value = 'TOTAL GENERAL'
    t.font = Font(bold=True, size=10, color='FFFFFF', name='Calibri')
    t.fill = PatternFill(fill_type='solid', fgColor='1a3a5c')
    t.alignment = Alignment(horizontal='right', vertical='center')
    t.border = _border()
    for c2 in ['B', 'C', 'D', 'E']:
        ws[f'{c2}{r}'].border = _border()
        ws[f'{c2}{r}'].fill = PatternFill(fill_type='solid', fgColor='1a3a5c')
    _cell(r, 6, total_ini,  bold=True, bg='1a3a5c', fg='FFFFFF', align='right', fmt=FMT_NUM)
    _cell(r, 7, total_deb,  bold=True, bg='1a3a5c', fg='FFFFFF', align='right', fmt=FMT_NUM)
    _cell(r, 8, total_cred, bold=True, bg='1a3a5c', fg='FFFFFF', align='right', fmt=FMT_NUM)
    _cell(r, 9, None,       bold=True, bg='1a3a5c', fg='FFFFFF')
    _cell(r, 10, '',        bold=True, bg='1a3a5c', fg='FFFFFF')
    ws.row_dimensions[r].height = 20

    # Congelar paneles debajo del encabezado
    ws.freeze_panes = 'A5'

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ========================================
# LIQUIDACIONES DEFINITIVAS - NORMA COLOMBIANA
# ========================================
from datetime import datetime

def calcular_dias_trabajados(fecha_ingreso, fecha_retiro):
    """Calcula días trabajados entre dos fechas."""
    try:
        if isinstance(fecha_ingreso, str):
            fecha_ingreso = datetime.strptime(fecha_ingreso, '%Y-%m-%d')
        if isinstance(fecha_retiro, str):
            fecha_retiro = datetime.strptime(fecha_retiro, '%Y-%m-%d')
        dias = (fecha_retiro - fecha_ingreso).days
        return max(0, dias)
    except:
        return 0

def calcular_prestaciones_colombianas(salario, dias_trabajados, tipo_retiro):
    """
    Calcula prestaciones sociales según normativa colombiana.

    Fórmulas:
    - Cesantías: (Salario × Días trabajados) / 360
    - Intereses cesantías: (Cesantías × 0.12 × Meses) / 12
    - Prima: (Salario × Días trabajados semestre) / 360
    - Vacaciones: (Salario × 15 días/año) / 30 (proporcional)
    - Indemnización: varía según tipo de retiro
    """
    salario = float(salario or 0)
    dias_trabajados = max(0, int(dias_trabajados or 0))

    # Cesantías (fondo de cesantía)
    cesantias = (salario * dias_trabajados) / 360

    # Intereses sobre cesantías: 12% anual
    meses_trabajados = dias_trabajados / 30
    intereses_cesantias = (cesantias * 0.12 * meses_trabajados) / 12

    # Prima de servicios (semestral: 180 días)
    # Se calcula proporcional
    prima = (salario * dias_trabajados) / 360

    # Vacaciones: 15 días al año = 0.5 días por día trabajado
    # Cálculo: (Salario / 30) × días proporcionales
    dias_vacaciones_proporcionales = (dias_trabajados * 15) / 360
    valor_vacaciones = (salario / 30) * dias_vacaciones_proporcionales

    # Indemnización según tipo de retiro
    indemnizacion = 0
    if tipo_retiro == 'despido_sin_causa':
        # Indemnización: 45 días de salario + prestaciones
        indemnizacion = salario * 1.5  # Equivalente a 45 días
    elif tipo_retiro == 'muerte' or tipo_retiro == 'jubilacion':
        indemnizacion = 0  # No aplica indemnización

    return {
        'cesantias': cesantias,
        'intereses_cesantias': intereses_cesantias,
        'prima': prima,
        'vacaciones': valor_vacaciones,
        'indemnizacion': indemnizacion,
        'dias_trabajados': dias_trabajados,
        'meses_trabajados': meses_trabajados
    }

def generate_liquidaciones(empleados_data):
    """Genera reporte de liquidaciones definitivas con norma colombiana."""
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Hoja 1: Resumen
    ws_resumen = wb.active
    ws_resumen.title = 'Resumen'

    # Estilos
    header_fill = PatternFill(fill_type='solid', fgColor='1a3a5c')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    subheader_fill = PatternFill(fill_type='solid', fgColor='2c3e50')
    subheader_font = Font(bold=True, color='FFFFFF', size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    num_fmt = '#,##0.00'
    right_align = Alignment(horizontal='right', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Encabezados resumen
    headers = ['Empleado', 'Salario', 'Días Trab.', 'Cesantías', 'Int. Cesantías',
               'Prima', 'Vacaciones', 'Indemnización', 'Total Liquidación']
    ws_resumen.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws_resumen.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Anchos
    anchos = [25, 15, 12, 15, 15, 15, 15, 15, 18]
    for idx, ancho in enumerate(anchos, 1):
        ws_resumen.column_dimensions[get_column_letter(idx)].width = ancho

    # Procesar empleados
    totales = {col: 0 for col in ['cesantias', 'intereses_cesantias', 'prima',
                                   'vacaciones', 'indemnizacion', 'total']}

    for emp_idx, emp in enumerate(empleados_data):
        nombre = emp.get('nombre', 'Sin nombre')
        salario = float(emp.get('salario', 0))
        fecha_ingreso = emp.get('fecha_ingreso', '')
        fecha_retiro = emp.get('fecha_retiro', '')
        tipo_retiro = emp.get('tipo_retiro', 'retiro_voluntario')

        # Calcular días trabajados
        dias_trabajados = calcular_dias_trabajados(fecha_ingreso, fecha_retiro)

        # Calcular prestaciones
        prestaciones = calcular_prestaciones_colombianas(salario, dias_trabajados, tipo_retiro)

        total_liquidacion = (prestaciones['cesantias'] +
                           prestaciones['intereses_cesantias'] +
                           prestaciones['prima'] +
                           prestaciones['vacaciones'] +
                           prestaciones['indemnizacion'])

        # Fila de datos
        row = ws_resumen.max_row + 1
        ws_resumen.cell(row=row, column=1, value=nombre)
        ws_resumen.cell(row=row, column=2, value=salario)
        ws_resumen.cell(row=row, column=3, value=dias_trabajados)
        ws_resumen.cell(row=row, column=4, value=prestaciones['cesantias'])
        ws_resumen.cell(row=row, column=5, value=prestaciones['intereses_cesantias'])
        ws_resumen.cell(row=row, column=6, value=prestaciones['prima'])
        ws_resumen.cell(row=row, column=7, value=prestaciones['vacaciones'])
        ws_resumen.cell(row=row, column=8, value=prestaciones['indemnizacion'])
        ws_resumen.cell(row=row, column=9, value=total_liquidacion)

        # Formato
        for col in range(1, 10):
            cell = ws_resumen.cell(row=row, column=col)
            cell.border = border
            if col in [2, 4, 5, 6, 7, 8, 9]:
                cell.number_format = num_fmt
                cell.alignment = right_align

        # Acumular totales
        totales['cesantias'] += prestaciones['cesantias']
        totales['intereses_cesantias'] += prestaciones['intereses_cesantias']
        totales['prima'] += prestaciones['prima']
        totales['vacaciones'] += prestaciones['vacaciones']
        totales['indemnizacion'] += prestaciones['indemnizacion']
        totales['total'] += total_liquidacion

        # Hoja detallada por empleado
        ws_detalle = wb.create_sheet(f'Detalles {emp_idx + 1}')
        ws_detalle.append(['LIQUIDACIÓN DEFINITIVA - NORMA COLOMBIANA', '', ''])
        ws_detalle.append(['Empleado:', nombre, ''])
        ws_detalle.append(['Salario Mensual:', f'${salario:,.2f}', ''])
        ws_detalle.append(['Fecha Ingreso:', fecha_ingreso, ''])
        ws_detalle.append(['Fecha Retiro:', fecha_retiro, ''])
        ws_detalle.append(['Días Trabajados:', dias_trabajados, ''])
        ws_detalle.append(['Meses Trabajados:', f'{prestaciones["meses_trabajados"]:.1f}', ''])
        ws_detalle.append(['Tipo de Retiro:', tipo_retiro, ''])
        ws_detalle.append(['', '', ''])

        # Detalles de conceptos
        conceptos = [
            ('Cesantías (Salario × Días / 360)', prestaciones['cesantias']),
            ('Intereses sobre Cesantías (12% anual)', prestaciones['intereses_cesantias']),
            ('Prima de Servicios (Salario × Días / 360)', prestaciones['prima']),
            ('Vacaciones (Salario × 15 días/año ÷ 30)', prestaciones['vacaciones']),
            ('Indemnización por Retiro', prestaciones['indemnizacion']),
        ]

        row = ws_detalle.max_row + 1
        for concepto, valor in conceptos:
            ws_detalle.append([concepto, f'${valor:,.2f}'])
            ws_detalle.cell(row=ws_detalle.max_row, column=1).font = Font(size=10)
            ws_detalle.cell(row=ws_detalle.max_row, column=2).number_format = num_fmt
            ws_detalle.cell(row=ws_detalle.max_row, column=2).font = Font(bold=True)

        # Total
        ws_detalle.append(['TOTAL LIQUIDACIÓN', f'${total_liquidacion:,.2f}'])
        total_row = ws_detalle.max_row
        ws_detalle.cell(row=total_row, column=1).font = Font(bold=True, size=11)
        ws_detalle.cell(row=total_row, column=2).font = Font(bold=True, size=11)
        ws_detalle.cell(row=total_row, column=2).fill = PatternFill(fill_type='solid', fgColor='d4e8f0')
        ws_detalle.cell(row=total_row, column=2).number_format = num_fmt

        # Anchos
        ws_detalle.column_dimensions['A'].width = 40
        ws_detalle.column_dimensions['B'].width = 20

    # Fila de totales en resumen
    row = ws_resumen.max_row + 1
    ws_resumen.cell(row=row, column=1, value='TOTAL').font = Font(bold=True, size=11)
    ws_resumen.cell(row=row, column=2, value='').font = Font(bold=True, size=11)
    ws_resumen.cell(row=row, column=3, value='').font = Font(bold=True, size=11)
    ws_resumen.cell(row=row, column=4, value=totales['cesantias']).number_format = num_fmt
    ws_resumen.cell(row=row, column=5, value=totales['intereses_cesantias']).number_format = num_fmt
    ws_resumen.cell(row=row, column=6, value=totales['prima']).number_format = num_fmt
    ws_resumen.cell(row=row, column=7, value=totales['vacaciones']).number_format = num_fmt
    ws_resumen.cell(row=row, column=8, value=totales['indemnizacion']).number_format = num_fmt
    ws_resumen.cell(row=row, column=9, value=totales['total']).number_format = num_fmt

    for col in range(1, 10):
        cell = ws_resumen.cell(row=row, column=col)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(fill_type='solid', fgColor='d4e8f0')
        cell.border = border
        if col in [4, 5, 6, 7, 8, 9]:
            cell.alignment = right_align

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
