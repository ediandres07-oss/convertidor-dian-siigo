import logging
from io import BytesIO
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from backend.db.session import get_db
from backend.models.user import User
from backend.models.order import Order, OrderStatus
from backend.models.inventory import Inventory
from backend.dependencies import get_current_user, get_current_admin_user

router = APIRouter()
logger = logging.getLogger(__name__)


def generate_orders_excel(orders: list[Order]):
    """Generate Excel report for orders."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel generation not available",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Headers
    headers = ["Order ID", "Order Number", "Status", "Total Amount", "Discount", "Tax", "Final Amount", "Date", "Due Date"]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Data
    for order in orders:
        ws.append([
            order.id,
            order.order_number,
            order.status.value,
            order.total_amount,
            order.discount,
            order.tax,
            order.final_amount,
            order.order_date.strftime("%Y-%m-%d"),
            order.due_date.strftime("%Y-%m-%d") if order.due_date else "",
        ])

    # Adjust column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def generate_inventory_excel(items: list[Inventory]):
    """Generate Excel report for inventory."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel generation not available",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # Headers
    headers = ["SKU", "Name", "Category", "Quantity", "Min Qty", "Max Qty", "Cost Price", "Selling Price", "Warehouse", "Low Stock"]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Data
    for item in items:
        ws.append([
            item.sku,
            item.name,
            item.category,
            item.quantity,
            item.minimum_quantity,
            item.maximum_quantity,
            item.cost_price,
            item.selling_price,
            item.warehouse,
            "Yes" if item.is_low_stock else "No",
        ])

    # Adjust column widths
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[col].width = 15

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def generate_orders_pdf(orders: list[Order]):
    """Generate PDF report for orders."""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="PDF generation not available",
        )

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title = Paragraph("Orders Report", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 0.3 * inch))

    # Date
    date_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    elements.append(Paragraph(f"Generated: {date_str}", styles["Normal"]))
    elements.append(Spacer(1, 0.2 * inch))

    # Table data
    data = [["Order ID", "Number", "Status", "Total", "Date"]]
    for order in orders:
        data.append([
            str(order.id),
            order.order_number,
            order.status.value,
            f"${order.final_amount:.2f}",
            order.order_date.strftime("%Y-%m-%d"),
        ])

    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 14),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return buffer


@router.get("/orders/excel")
async def export_orders_excel(
    current_user: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    """Export orders to Excel."""
    try:
        orders = db.query(Order).all()

        output = generate_orders_excel(orders)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=orders_report.xlsx"},
        )

    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error generating report",
        )


@router.get("/orders/pdf")
async def export_orders_pdf(
    current_user: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    """Export orders to PDF."""
    try:
        orders = db.query(Order).all()

        output = generate_orders_pdf(orders)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=orders_report.pdf"},
        )

    except Exception as e:
        logger.error(f"Error generating PDF report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error generating report",
        )


@router.get("/inventory/excel")
async def export_inventory_excel(
    current_user: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    """Export inventory to Excel."""
    try:
        items = db.query(Inventory).all()

        output = generate_inventory_excel(items)

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=inventory_report.xlsx"},
        )

    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error generating report",
        )


@router.get("/summary")
async def get_report_summary(
    current_user: User = Depends(get_current_admin_user),
    db: Session = Depends(get_db),
):
    """Get summary statistics for reports."""
    try:
        total_orders = db.query(Order).count()
        total_users = db.query(User).count()
        total_inventory = db.query(Inventory).count()

        pending_orders = db.query(Order).filter(Order.status == OrderStatus.PENDING).count()
        completed_orders = db.query(Order).filter(Order.status == OrderStatus.COMPLETED).count()

        total_revenue = sum(order.final_amount for order in db.query(Order).all())
        low_stock_items = sum(1 for item in db.query(Inventory).all() if item.is_low_stock)

        return {
            "total_orders": total_orders,
            "pending_orders": pending_orders,
            "completed_orders": completed_orders,
            "total_revenue": total_revenue,
            "total_users": total_users,
            "total_inventory_items": total_inventory,
            "low_stock_items": low_stock_items,
            "generated_at": datetime.utcnow().isoformat(),
        }

    except Exception as e:
        logger.error(f"Error generating summary: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error generating summary",
        )
