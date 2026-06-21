"""
Integración con Siigo Nube API para subir planos automáticamente.
"""
import requests
import json
from datetime import datetime
import os
from typing import Dict, List, Optional
import openpyxl

# Configuración de Siigo
SIIGO_BASE_URL = os.environ.get('SIIGO_BASE_URL', 'https://api.siigo.com/v1')
SIIGO_USERNAME = os.environ.get('SIIGO_USERNAME', '')
SIIGO_PASSWORD = os.environ.get('SIIGO_PASSWORD', '')
SIIGO_PARTNER_ID = os.environ.get('SIIGO_PARTNER_ID', 'ConvertidorDIAN')

# Tipos de documentos en Siigo (mapeo desde nuestros planos)
DOCUMENTO_TIPOS = {
    'COMPRAS': 'FC',        # Factura de compra
    'GASTOS': 'DS',         # Documento soporte
    'VENTAS': 'FV',         # Factura de venta
    'NC COMPRAS': 'NC',     # Nota crédito compra
    'NC VENTAS': 'NC',      # Nota crédito venta
    'NOMINA': 'FC',         # Factura de compra (nómina)
}


class SiigoAPIClient:
    """Cliente para interactuar con la API de Siigo Nube."""

    def __init__(self, username: str = None, password: str = None):
        self.username = username or SIIGO_USERNAME
        self.password = password or SIIGO_PASSWORD
        self.access_token = None
        self.token_expires = None
        self.session = requests.Session()

    def authenticate(self) -> bool:
        """Obtiene un JWT access_token válido por 24 horas."""
        try:
            url = f"{SIIGO_BASE_URL}/auth"
            headers = {
                'Partner-Id': SIIGO_PARTNER_ID,
                'Content-Type': 'application/json'
            }
            payload = {
                'username': self.username,
                'password': self.password
            }

            response = requests.post(url, json=payload, headers=headers, timeout=10)
            response.raise_for_status()

            data = response.json()
            self.access_token = data.get('access_token')
            self.token_expires = data.get('expires_in')  # segundos

            return bool(self.access_token)

        except Exception as e:
            print(f"❌ Error autenticando con Siigo: {e}")
            return False

    def _get_headers(self) -> Dict:
        """Retorna headers con autenticación."""
        return {
            'Authorization': self.access_token,
            'Partner-Id': SIIGO_PARTNER_ID,
            'Content-Type': 'application/json'
        }

    def crear_tercero(self, nit: str, nombre: str, es_proveedor: bool = False) -> Optional[Dict]:
        """Crea un cliente o proveedor en Siigo."""
        try:
            url = f"{SIIGO_BASE_URL}/customers"
            person_type = "company" if len(nit) > 10 else "person"
            id_type = "31" if len(nit) > 10 else "13"  # 31=NIT, 13=Cédula

            payload = {
                "type": "Supplier" if es_proveedor else "Customer",
                "person_type": person_type,
                "id_type": id_type,
                "identification": nit.replace('-', ''),
                "name": nombre,
                "active": True,
                "vat_responsible": False,
                "address": {
                    "address": "Dirección por defecto",
                    "city": {
                        "country_code": "CO",
                        "state_code": "05",
                        "city_code": "05001"
                    }
                },
                "fiscal_responsibilities": [
                    {"code": "R-99-PN"}  # No aplica
                ]
            }

            response = self.session.post(url, json=payload, headers=self._get_headers(), timeout=10)
            response.raise_for_status()
            return response.json()

        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 409:
                # Tercero ya existe
                return {"identification": nit}
            print(f"❌ Error creando tercero {nit}: {e}")
            return None
        except Exception as e:
            print(f"❌ Error: {e}")
            return None

    def crear_factura_compra(self, factura_data: Dict) -> Optional[Dict]:
        """Crea una factura de compra (documento soporte o factura)."""
        try:
            url = f"{SIIGO_BASE_URL}/invoices"

            # Asegurar que el tercero existe
            self.crear_tercero(factura_data['nit'], factura_data.get('tercero', 'Proveedor'), True)

            payload = {
                "document": {"id": 1},  # ID del tipo de factura de compra
                "date": factura_data['fecha'].strftime('%Y/%m/%d') if hasattr(factura_data['fecha'], 'strftime') else factura_data['fecha'],
                "number": factura_data.get('numero', ''),
                "customer": {
                    "identification": factura_data['nit'].replace('-', ''),
                    "branch_office": 0
                },
                "seller": 1,  # Usuario por defecto
                "items": [
                    {
                        "code": "SRV-001",  # Servicio genérico
                        "description": factura_data.get('descripcion', f"Comprobante {factura_data.get('folio')}"),
                        "quantity": 1,
                        "price": factura_data.get('total', 0)
                    }
                ]
            }

            response = self.session.post(url, json=payload, headers=self._get_headers(), timeout=10)
            response.raise_for_status()
            return response.json()

        except Exception as e:
            print(f"❌ Error creando factura: {e}")
            return None

    def crear_nota_credito(self, nc_data: Dict) -> Optional[Dict]:
        """Crea una nota crédito."""
        try:
            url = f"{SIIGO_BASE_URL}/credit-notes"

            payload = {
                "document": {"id": 2},  # ID nota crédito
                "date": nc_data['fecha'].strftime('%Y/%m/%d') if hasattr(nc_data['fecha'], 'strftime') else nc_data['fecha'],
                "number": nc_data.get('numero', ''),
                "items": [
                    {
                        "code": "SRV-001",
                        "description": nc_data.get('descripcion', f"NC {nc_data.get('folio')}"),
                        "quantity": 1,
                        "price": nc_data.get('total', 0)
                    }
                ]
            }

            response = self.session.post(url, json=payload, headers=self._get_headers(), timeout=10)
            response.raise_for_status()
            return response.json()

        except Exception as e:
            print(f"❌ Error creando nota crédito: {e}")
            return None


def extraer_plano_datos(plano_bytes: bytes, sheet_name: str) -> List[Dict]:
    """Extrae datos del plano Excel generado."""
    try:
        wb = openpyxl.load_workbook(plano_bytes)
        ws = wb[sheet_name]

        datos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Si tipo de comprobante está vacío, saltar
                continue

            datos.append({
                'tipo': row[0],
                'consecutivo': row[1],
                'fecha': row[2],
                'cuenta': row[5],
                'nit': str(row[6] or '').strip(),
                'cod_impuesto': row[7],
                'descripcion': row[8],
                'debito': row[21],
                'credito': row[22]
            })

        return datos

    except Exception as e:
        print(f"❌ Error extrayendo datos del plano: {e}")
        return []


def subir_planos_a_siigo(planos_excel: bytes, username: str, password: str) -> Dict:
    """Sube los planos generados a Siigo Nube automáticamente."""
    cliente = SiigoAPIClient(username, password)

    if not cliente.authenticate():
        return {'error': 'No se pudo autenticar con Siigo'}

    resultados = {
        'exito': 0,
        'error': 0,
        'detalles': []
    }

    try:
        wb = openpyxl.load_workbook(planos_excel)

        for sheet_name in ['COMPRAS', 'NC COMPRAS', 'GASTOS', 'VENTAS', 'NC VENTAS']:
            if sheet_name not in wb.sheetnames:
                continue

            datos = extraer_plano_datos(planos_excel, sheet_name)

            for registro in datos:
                try:
                    if 'NC' in sheet_name:
                        resultado = cliente.crear_nota_credito(registro)
                    else:
                        resultado = cliente.crear_factura_compra(registro)

                    if resultado:
                        resultados['exito'] += 1
                        resultados['detalles'].append({
                            'plano': sheet_name,
                            'status': 'OK',
                            'folio': registro.get('numero', 'N/A')
                        })
                    else:
                        resultados['error'] += 1
                        resultados['detalles'].append({
                            'plano': sheet_name,
                            'status': 'ERROR',
                            'folio': registro.get('numero', 'N/A')
                        })

                except Exception as e:
                    resultados['error'] += 1
                    resultados['detalles'].append({
                        'plano': sheet_name,
                        'status': f'ERROR: {str(e)[:50]}',
                        'folio': registro.get('numero', 'N/A')
                    })

        return resultados

    except Exception as e:
        return {
            'error': f'Error procesando planos: {str(e)}',
            'exito': 0,
            'error': resultados['error']
        }
