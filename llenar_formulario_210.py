#!/usr/bin/env python3
"""
Script para llenar el Formulario 210 con datos de Exógena
Mapea información automáticamente
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
import json

class LlenadorFormulario210:
    """Llena automáticamente el Formulario 210"""

    def __init__(self, ruta_formulario, ruta_exogena):
        self.ruta_formulario = ruta_formulario
        self.ruta_exogena = ruta_exogena
        self.wb = None
        self.ws = None
        self.datos_exogena = {}

    def extraer_exogena(self):
        """Extrae datos de la Exógena"""
        df = pd.read_excel(self.ruta_exogena, sheet_name='Reporte', header=None)

        # Extraer datos del contribuyente
        for i in range(len(df)):
            row = df.iloc[i]
            if pd.notna(row.iloc[0]):
                contenido = str(row.iloc[0]).lower()
                if 'identificación:' in contenido and pd.notna(row.iloc[2]):
                    self.datos_exogena['cedula'] = str(row.iloc[2])
                elif 'nombres' in contenido and pd.notna(row.iloc[2]):
                    self.datos_exogena['nombre'] = str(row.iloc[2])
                elif 'año' in contenido and 'refiere' in contenido and pd.notna(row.iloc[2]):
                    self.datos_exogena['ano'] = int(row.iloc[2])

        # Extraer ingresos y retenciones
        ingresos = {}
        retenciones = {}

        for i in range(14, len(df)):
            row = df.iloc[i]
            if pd.notna(row.iloc[4]) and pd.notna(row.iloc[5]):
                concepto = str(row.iloc[4]).strip()
                try:
                    valor = float(row.iloc[5])
                    if valor > 0:
                        uso = str(row.iloc[6]).lower() if pd.notna(row.iloc[6]) else ""
                        if 'retención' in uso or 'r132' in uso or 'r28' in uso:
                            if concepto not in retenciones:
                                retenciones[concepto] = 0
                            retenciones[concepto] += valor
                        else:
                            if concepto not in ingresos:
                                ingresos[concepto] = 0
                            ingresos[concepto] += valor
                except:
                    pass

        self.datos_exogena['ingresos'] = ingresos
        self.datos_exogena['retenciones'] = retenciones
        self.datos_exogena['total_ingresos'] = sum(ingresos.values())
        self.datos_exogena['total_retenciones'] = sum(retenciones.values())

        return self.datos_exogena

    def abrir_formulario(self):
        """Abre el formulario 210"""
        try:
            self.wb = load_workbook(self.ruta_formulario)
            self.ws = self.wb.active
            print("✅ Formulario 210 abierto")
            return True
        except Exception as e:
            print(f"❌ Error abriendo formulario: {e}")
            return False

    def llenar_datos_contribuyente(self):
        """Llena datos personales del contribuyente"""
        print("\n📝 Llenando datos del contribuyente...")

        # Buscar celdas para llenar
        # Año gravable (línea 1)
        for row in self.ws.iter_rows(min_row=1, max_row=30):
            for cell in row:
                if cell.value and '1. Año' in str(cell.value):
                    # Llenar año en celda cercana
                    col = cell.column
                    self.ws.cell(row=cell.row, column=col+3).value = self.datos_exogena.get('ano', 2024)
                    print(f"  ✓ Año: {self.datos_exogena.get('ano', 2024)}")

                # Cédula
                if cell.value and 'Número de Identificación' in str(cell.value):
                    cedula = self.datos_exogena.get('cedula', '')
                    # Buscar celda vacía cercana para llenar
                    for offset in range(1, 5):
                        siguiente = self.ws.cell(row=cell.row+offset, column=cell.column)
                        if siguiente.value is None:
                            siguiente.value = cedula
                            print(f"  ✓ Cédula: {cedula}")
                            break

    def llenar_ingresos(self):
        """Llena la sección de ingresos"""
        print("\n💰 Llenando ingresos...")

        total_ingresos = self.datos_exogena.get('total_ingresos', 0)

        # Buscar celda de ingresos brutos
        for row in self.ws.iter_rows(min_row=20, max_row=50):
            for cell in row:
                if cell.value and 'Ingresos brutos' in str(cell.value):
                    # Llenar en la columna correspondiente
                    col = cell.column
                    siguiente_row = cell.row + 1
                    self.ws.cell(row=siguiente_row, column=col).value = total_ingresos
                    print(f"  ✓ Ingresos Brutos: ${total_ingresos:,.0f}")
                    return True

        return False

    def llenar_retenciones(self):
        """Llena retenciones en la fuente"""
        print("\n📌 Llenando retenciones...")

        total_retenciones = self.datos_exogena.get('total_retenciones', 0)

        # Buscar sección de retenciones
        for row in self.ws.iter_rows(min_row=20, max_row=80):
            for cell in row:
                if cell.value and ('Retención' in str(cell.value) or 'R132' in str(cell.value)):
                    col = cell.column
                    siguiente_row = cell.row + 1
                    self.ws.cell(row=siguiente_row, column=col).value = total_retenciones
                    print(f"  ✓ Retenciones: ${total_retenciones:,.0f}")
                    return True

        return False

    def calcular_irpf(self):
        """Calcula y llena el IRPF"""
        print("\n🧮 Calculando IRPF...")

        base = self.datos_exogena.get('total_ingresos', 0)
        deducciones = base * 0.10
        base_liquida = base - deducciones

        # Aplicar tarifa IRPF 2024
        rangos = [
            (0, 66950000, 0),
            (66950000, 134900000, 0.05),
            (134900000, 404700000, 0.12),
            (404700000, 673500000, 0.25),
            (673500000, 1347000000, 0.32),
            (1347000000, float('inf'), 0.37),
        ]

        irpf = 0
        for inicio, fin, tasa in rangos:
            if base_liquida > inicio:
                renta_tramo = min(base_liquida, fin) - inicio
                irpf += renta_tramo * tasa

        print(f"  ✓ Base Líquida: ${base_liquida:,.0f}")
        print(f"  ✓ IRPF Calculado: ${irpf:,.0f}")

        return {
            'base_liquida': base_liquida,
            'irpf': irpf,
            'total_impuesto': irpf
        }

    def guardar(self, ruta_salida=None):
        """Guarda el formulario lleno"""
        if not ruta_salida:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ruta_salida = f"Formulario210_Llenado_{timestamp}.xlsx"

        try:
            self.wb.save(ruta_salida)
            print(f"\n✅ Formulario guardado en: {ruta_salida}")
            return ruta_salida
        except Exception as e:
            print(f"❌ Error guardando: {e}")
            return None

    def procesar(self):
        """Proceso completo"""
        print("╔════════════════════════════════════════════════════════════╗")
        print("║           LLENADOR DE FORMULARIO 210                       ║")
        print("╚════════════════════════════════════════════════════════════╝\n")

        # Extraer datos
        print("📂 Extrayendo datos de Exógena...")
        self.extraer_exogena()
        print(f"  ✓ Contribuyente: {self.datos_exogena.get('nombre')}")
        print(f"  ✓ Cédula: {self.datos_exogena.get('cedula')}")
        print(f"  ✓ Año: {self.datos_exogena.get('ano')}")
        print(f"  ✓ Ingresos: ${self.datos_exogena.get('total_ingresos', 0):,.0f}")

        # Abrir formulario
        if not self.abrir_formulario():
            return False

        # Llenar datos
        self.llenar_datos_contribuyente()
        self.llenar_ingresos()
        self.llenar_retenciones()
        calculos = self.calcular_irpf()

        # Guardar
        ruta = self.guardar()

        # Mostrar resumen
        print("\n" + "="*60)
        print("RESUMEN DE LIQUIDACIÓN")
        print("="*60)
        print(f"Total Ingresos:      ${self.datos_exogena.get('total_ingresos', 0):>15,.0f}")
        print(f"Retenciones:         ${self.datos_exogena.get('total_retenciones', 0):>15,.0f}")
        print(f"Base Líquida:        ${calculos['base_liquida']:>15,.0f}")
        print(f"IRPF:                ${calculos['irpf']:>15,.0f}")
        print(f"\nArchivo generado: {ruta}")

        return ruta

if __name__ == "__main__":
    # Procesar
    llenador = LlenadorFormulario210(
        ruta_formulario="/Users/edison/Downloads/formulario 210.xlsx",
        ruta_exogena="/Users/edison/Downloads/reporteExogena2024Elizabeth.xlsx"
    )

    ruta_salida = llenador.procesar()
