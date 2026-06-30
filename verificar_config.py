#!/usr/bin/env python3
"""
VERIFICADOR DE CONFIGURACIÓN - Exógena → Ayuda Renta
======================================================
Verifica que todos los archivos, dependencias y configuraciones
sean correctas antes de ejecutar el automatizador principal.
"""

import sys
from pathlib import Path
import subprocess

def print_header(titulo):
    """Imprime un header de sección."""
    print(f"\n{'='*70}")
    print(f"  {titulo}")
    print(f"{'='*70}\n")

def check_python_version():
    """Verifica que Python 3.7+ esté disponible."""
    print_header("1. VERIFICANDO VERSIÓN DE PYTHON")

    version = sys.version_info
    version_str = f"{version.major}.{version.minor}.{version.micro}"

    if version.major >= 3 and version.minor >= 7:
        print(f"✓ Python {version_str} (OK)")
        return True
    else:
        print(f"✗ Python {version_str} - Se requiere Python 3.7+")
        return False

def check_dependencies():
    """Verifica que pandas y openpyxl estén instalados."""
    print_header("2. VERIFICANDO DEPENDENCIAS")

    all_ok = True

    # Verificar pandas
    try:
        import pandas as pd
        print(f"✓ pandas {pd.__version__}")
    except ImportError:
        print("✗ pandas NO instalado")
        print("  Ejecuta: pip3 install pandas")
        all_ok = False

    # Verificar openpyxl
    try:
        import openpyxl
        print(f"✓ openpyxl {openpyxl.__version__}")
    except ImportError:
        print("✗ openpyxl NO instalado")
        print("  Ejecuta: pip3 install openpyxl")
        all_ok = False

    return all_ok

def check_files():
    """Verifica que los archivos requeridos existan."""
    print_header("3. VERIFICANDO ARCHIVOS REQUERIDOS")

    files_to_check = {
        "Exógena": "/Users/edison/Library/Containers/net.whatsapp.WhatsApp/Data/tmp/documents/0141329A-8688-45CB-9787-E5FD5051A852/reporteExogena2024Elizabeth.xlsx",
        "Ayuda Renta": "/Users/edison/Downloads/Programa-Ayuda-Renta-DIAN-2025/AyudaRenta 2025 V1.0 .xlsm",
        "Script Principal": "/Users/edison/Desktop/proyecto-subir info a siigo nube/exogena_to_ayuda_renta.py",
    }

    all_ok = True

    for nombre, ruta in files_to_check.items():
        path = Path(ruta)
        if path.exists():
            size_mb = path.stat().st_size / (1024 * 1024)
            print(f"✓ {nombre}: {ruta}")
            print(f"  Tamaño: {size_mb:.2f} MB")
        else:
            print(f"✗ {nombre} NO ENCONTRADO: {ruta}")
            all_ok = False

    return all_ok

def check_exogena_structure():
    """Verifica la estructura del archivo Exógena."""
    print_header("4. VERIFICANDO ESTRUCTURA DEL EXÓGENA")

    try:
        import pandas as pd

        exogena_path = "/Users/edison/Library/Containers/net.whatsapp.WhatsApp/Data/tmp/documents/0141329A-8688-45CB-9787-E5FD5051A852/reporteExogena2024Elizabeth.xlsx"

        df = pd.read_excel(exogena_path, sheet_name="Reporte", header=13)

        print(f"✓ Archivo Exógena leído correctamente")
        print(f"  Hojas disponibles: {df.shape[0]} filas, {df.shape[1]} columnas")
        print(f"  Columnas encontradas:")

        required_columns = ["Detalle", "Valor", "Uso declaración Sugerida"]
        all_ok = True

        for col in required_columns:
            if col in df.columns:
                print(f"    ✓ {col}")
            else:
                print(f"    ✗ {col} FALTA")
                all_ok = False

        # Convertir y verificar datos
        df["Valor"] = pd.to_numeric(df["Valor"], errors="coerce")
        df = df.dropna(subset=["Valor"])

        print(f"\n  Registros de datos: {len(df)}")

        # Mostrar categorías
        categorias = df["Uso declaración Sugerida"].nunique()
        print(f"  Categorías únicas: {categorias}")

        print("\n  Resumen de valores por categoría:")
        resumen = df.groupby("Uso declaración Sugerida")["Valor"].sum().sort_values(ascending=False).head(5)
        for cat, valor in resumen.items():
            if pd.notna(cat):
                print(f"    • {cat[:50]}: ${valor:,.0f}")

        return all_ok

    except Exception as e:
        print(f"✗ Error al verificar Exógena: {e}")
        return False

def check_ayuda_renta_structure():
    """Verifica la estructura del archivo Ayuda Renta."""
    print_header("5. VERIFICANDO ESTRUCTURA DEL AYUDA RENTA")

    try:
        import openpyxl

        ayuda_path = "/Users/edison/Downloads/Programa-Ayuda-Renta-DIAN-2025/AyudaRenta 2025 V1.0 .xlsm"

        wb = openpyxl.load_workbook(ayuda_path)

        print(f"✓ Archivo Ayuda Renta leído correctamente")
        print(f"  Total de hojas: {len(wb.sheetnames)}")

        # Verificar hojas clave
        hojas_esperadas = [
            "Retenciones",
            "Patrimonio",
            "Efectivo_Bancos_Cuentas",
            "Inversiones",
            "Pensiones",
            "Salarios_Demas_Pagos_Laborales",
            "Inter_Rend_Finan"
        ]

        print(f"\n  Hojas requeridas:")
        all_ok = True
        for hoja in hojas_esperadas:
            if hoja in wb.sheetnames:
                print(f"    ✓ {hoja}")
            else:
                print(f"    ✗ {hoja} FALTA")
                all_ok = False

        # Verificar si hay VBA (macros)
        if wb.vba_archive:
            print(f"\n  ✓ Macros detectadas (serán preservadas)")
        else:
            print(f"\n  ℹ️  No se detectaron macros (archivo sin VBA)")

        return all_ok

    except Exception as e:
        print(f"✗ Error al verificar Ayuda Renta: {e}")
        return False

def main():
    """Ejecuta todas las verificaciones."""
    print("\n")
    print("╔" + "="*68 + "╗")
    print("║" + " "*68 + "║")
    print("║" + "  VERIFICADOR DE CONFIGURACIÓN - EXÓGENA → AYUDA RENTA".center(68) + "║")
    print("║" + " "*68 + "║")
    print("╚" + "="*68 + "╝\n")

    results = {}

    results["Python"] = check_python_version()
    results["Dependencias"] = check_dependencies()
    results["Archivos"] = check_files()
    results["Exógena"] = check_exogena_structure()
    results["Ayuda Renta"] = check_ayuda_renta_structure()

    # Resumen final
    print_header("RESUMEN DE VERIFICACIÓN")

    all_ok = all(results.values())

    for nombre, resultado in results.items():
        estado = "✓ OK" if resultado else "✗ ERROR"
        print(f"{estado:8} {nombre}")

    print("\n" + "="*70)

    if all_ok:
        print("\n✓ ¡TODAS LAS VERIFICACIONES PASARON!")
        print("\nPuedes ejecutar el script principal con:\n")
        print("  cd /Users/edison/Desktop/proyecto-subir\\ info\\ a\\ siigo\\ nube/")
        print("  python3 exogena_to_ayuda_renta.py\n")
        return 0
    else:
        print("\n✗ ALGUNAS VERIFICACIONES FALLARON")
        print("\nPor favor, resuelve los errores anteriores antes de ejecutar")
        print("el script principal.\n")
        return 1

if __name__ == "__main__":
    exit(main())
