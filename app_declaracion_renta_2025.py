#!/usr/bin/env python3
"""
Declaración de Renta 2025 - App Integrada CORREGIDA
Carga Exógena → Calcula Cedular → Genera Formulario 210
Con mapeo correcto según guía DIAN oficial
"""

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from datetime import datetime
from mapeo_cedulas_2025 import clasificar_renta_v2

st.set_page_config(
    page_title="Declaración Renta 2025",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

VALORES_2025 = {'uvt': 47065, 'smlv': 1320000}

if 'paso_actual' not in st.session_state:
    st.session_state.paso_actual = 1
if 'datos_exogena' not in st.session_state:
    st.session_state.datos_exogena = None
if 'datos_cedular' not in st.session_state:
    st.session_state.datos_cedular = {}

def limpiar_valor(valor):
    if pd.isna(valor):
        return 0
    if isinstance(valor, str):
        valor = valor.strip().replace('$', '').replace(',', '')
    try:
        return float(valor) if valor else 0
    except:
        return 0

def extraer_datos_exogena(archivo):
    try:
        df = pd.read_excel(archivo, sheet_name=0, header=None)
        datos = {'contribuyente': {}, 'rentas': []}

        for idx, row in df.iterrows():
            row_str = str(row.iloc[0]).lower() if pd.notna(row.iloc[0]) else ""
            if 'cédula' in row_str or 'identificación' in row_str:
                if len(row) > 2 and pd.notna(row.iloc[2]):
                    datos['contribuyente']['cedula'] = str(row.iloc[2])
            elif 'nombre' in row_str:
                if len(row) > 2 and pd.notna(row.iloc[2]):
                    datos['contribuyente']['nombre'] = str(row.iloc[2])

        for idx, row in df.iterrows():
            cols = [str(c).lower() for c in row if pd.notna(c)]
            if any(x in ' '.join(cols) for x in ['concepto', 'tipo', 'valor']):
                for data_idx in range(idx + 1, len(df)):
                    if pd.notna(df.iloc[data_idx, 0]):
                        concepto = str(df.iloc[data_idx, 0])
                        valor = limpiar_valor(df.iloc[data_idx, 2] if len(df.columns) > 2 else 0)
                        if valor > 0:
                            clasificacion = clasificar_renta_v2(concepto, valor)
                            datos['rentas'].append(clasificacion)
                break
        return datos
    except Exception as e:
        st.error(f"Error: {e}")
        return None

def calcular_cedula_general(rentas):
    cedula = {
        'tipo': 'CÉDULA GENERAL',
        'ingresos_trabajo': 0,
        'ingresos_capital': 0,
        'ingresos_no_laborales': 0,
        'ingresos_totales': 0,
        'deducciones': 0,
        'renta_liquida': 0,
        'detalle_trabajo': [],
        'detalle_capital': [],
        'detalle_no_laborales': []
    }

    for renta in rentas:
        if renta['cedula'] != 'CEDULA_GENERAL':
            continue

        if renta['categoria'] == 'TRABAJO':
            cedula['ingresos_trabajo'] += renta['valor']
            cedula['detalle_trabajo'].append(renta)
        elif renta['categoria'] == 'CAPITAL':
            cedula['ingresos_capital'] += renta['valor']
            cedula['detalle_capital'].append(renta)
        elif renta['categoria'] == 'NO_LABORALES':
            cedula['ingresos_no_laborales'] += renta['valor']
            cedula['detalle_no_laborales'].append(renta)

    cedula['ingresos_totales'] = cedula['ingresos_trabajo'] + cedula['ingresos_capital'] + cedula['ingresos_no_laborales']
    cedula['deducciones'] = min(cedula['ingresos_trabajo'] * 0.25, 95000000) if cedula['ingresos_trabajo'] > 0 else 0
    cedula['renta_liquida'] = cedula['ingresos_totales'] - cedula['deducciones']

    return cedula

def calcular_cedula_pensiones(rentas):
    cedula = {
        'tipo': 'CÉDULA DE PENSIONES',
        'ingresos': 0,
        'deducciones': 0,
        'renta_liquida': 0,
        'detalle': []
    }
    for renta in rentas:
        if renta['cedula'] == 'CEDULA_PENSIONES':
            cedula['ingresos'] += renta['valor']
            cedula['detalle'].append(renta)
    cedula['renta_liquida'] = cedula['ingresos'] - cedula['deducciones']
    return cedula

def calcular_ganancias_ocasionales(rentas):
    cedula = {
        'tipo': 'GANANCIAS OCASIONALES',
        'ingresos': 0,
        'tarifa': 0.10,
        'impuesto': 0,
        'detalle': []
    }
    for renta in rentas:
        if renta['cedula'] == 'GANANCIAS_OCASIONALES':
            cedula['ingresos'] += renta['valor']
            cedula['detalle'].append(renta)
    cedula['impuesto'] = cedula['ingresos'] * cedula['tarifa']
    return cedula

def calcular_irpf_correcto(renta_neta):
    if renta_neta <= 0:
        return 0
    franjas = [
        (1650000, 0),
        (2750000, 0.05),
        (4433333, 0.08),
        (6200000, 0.11),
        (9000000, 0.13),
        (18700000, 0.17),
        (35000000, 0.26),
        (float('inf'), 0.33)
    ]
    impuesto = 0
    base_anterior = 0
    for limite, tarifa in franjas:
        if renta_neta <= base_anterior:
            break
        base_actual = min(renta_neta, limite)
        tramo = base_actual - base_anterior
        if tramo > 0:
            impuesto += tramo * tarifa
        base_anterior = limite
    return impuesto

# INTERFACE
st.sidebar.title("📋 Declaración Renta 2025")
st.sidebar.markdown("---")

col1, col2, col3 = st.sidebar.columns(3)
for i in range(1, 4):
    with [col1, col2, col3][i-1]:
        if st.button(f"Paso {i}", use_container_width=True, key=f"paso_{i}",
                    type="primary" if st.session_state.paso_actual == i else "secondary"):
            st.session_state.paso_actual = i

# PASO 1
if st.session_state.paso_actual == 1:
    st.title("📥 Cargar Archivo Exógena")
    archivo = st.file_uploader("Selecciona archivo Excel exógena", type=['xlsx', 'xls'])

    if archivo:
        with st.spinner("Procesando..."):
            datos = extraer_datos_exogena(archivo)

        if datos:
            st.session_state.datos_exogena = datos

            with st.container(border=True):
                st.subheader("✅ Datos Extraídos")
                col1, col2 = st.columns(2)
                with col1:
                    st.write(f"**Cédula:** {datos['contribuyente'].get('cedula', 'N/A')}")
                    st.write(f"**Nombre:** {datos['contribuyente'].get('nombre', 'N/A')}")
                with col2:
                    st.write(f"**Total rentas:** ${sum(r['valor'] for r in datos['rentas']):,.0f}")
                    st.write(f"**Registros:** {len(datos['rentas'])}")

            if datos['rentas']:
                st.subheader("📊 Rentas Detectadas")
                df_show = pd.DataFrame([{
                    'Concepto': r.get('descripcion', r['tipo']),
                    'Cédula': r['cedula'].replace('CEDULA_', '').replace('_', ' '),
                    'Categoría': r.get('categoria', 'N/A'),
                    'Valor': f"${r['valor']:,.0f}"
                } for r in datos['rentas']])
                st.dataframe(df_show, use_container_width=True, hide_index=True)

            if st.button("✅ Continuar →", key="paso1_continuar", type="primary"):
                st.session_state.paso_actual = 2
                st.rerun()

# PASO 2
elif st.session_state.paso_actual == 2:
    st.title("📊 Cálculo del Modo Cedular")

    if not st.session_state.datos_exogena:
        st.warning("⚠️ Carga primero el archivo")
        if st.button("← Volver"):
            st.session_state.paso_actual = 1
            st.rerun()
    else:
        datos = st.session_state.datos_exogena
        cedula_general = calcular_cedula_general(datos['rentas'])
        cedula_pensiones = calcular_cedula_pensiones(datos['rentas'])
        ganancias_ocasionales = calcular_ganancias_ocasionales(datos['rentas'])

        st.session_state.datos_cedular = [cedula_general, cedula_pensiones, ganancias_ocasionales]

        tabs = st.tabs(["Cédula General", "Cédula Pensiones", "Ganancias Ocasionales", "Resumen"])

        with tabs[0]:
            st.subheader("Cédula General")
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Ingresos Trabajo", f"${cedula_general['ingresos_trabajo']:,.0f}")
            with col2:
                st.metric("Ingresos Capital", f"${cedula_general['ingresos_capital']:,.0f}")
            with col3:
                st.metric("Ingresos No Laborales", f"${cedula_general['ingresos_no_laborales']:,.0f}")
            with col4:
                st.metric("Deducciones", f"-${cedula_general['deducciones']:,.0f}")

            st.metric("Renta Líquida General", f"${cedula_general['renta_liquida']:,.0f}")

        with tabs[1]:
            st.subheader("Cédula de Pensiones")
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Ingresos", f"${cedula_pensiones['ingresos']:,.0f}")
            with col2:
                st.metric("Renta Líquida", f"${cedula_pensiones['renta_liquida']:,.0f}")
            if not cedula_pensiones['detalle']:
                st.info("Sin ingresos de pensiones")

        with tabs[2]:
            st.subheader("Ganancias Ocasionales (Tarifa 10%)")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Ganancias", f"${ganancias_ocasionales['ingresos']:,.0f}")
            with col2:
                st.metric("Tarifa", "10%")
            with col3:
                st.metric("Impuesto", f"${ganancias_ocasionales['impuesto']:,.0f}")
            if not ganancias_ocasionales['detalle']:
                st.info("Sin ganancias ocasionales")

        with tabs[3]:
            st.subheader("Resumen Total de Impuestos")
            renta_total = cedula_general['renta_liquida'] + cedula_pensiones['renta_liquida']
            irpf = calcular_irpf_correcto(renta_total)
            aporte_solidario = max(0, (renta_total - 16000 * VALORES_2025['uvt']) * 0.01) if renta_total > 16000 * VALORES_2025['uvt'] else 0
            impuesto_ganancias = ganancias_ocasionales['impuesto']
            impuesto_total = irpf + aporte_solidario + impuesto_ganancias

            col1, col2 = st.columns(2)
            with col1:
                st.metric("Renta Total Gravable", f"${renta_total:,.0f}")
                st.metric("IRPF Cédulas", f"${irpf:,.0f}")
            with col2:
                st.metric("Aporte Solidario", f"${aporte_solidario:,.0f}")
                st.metric("Impuesto Total", f"${impuesto_total:,.0f}", delta=f"{(impuesto_total/renta_total*100):.2f}% tasa efectiva" if renta_total > 0 else "N/A")

        col1, col2 = st.columns(2)
        with col1:
            if st.button("← Volver"):
                st.session_state.paso_actual = 1
                st.rerun()
        with col2:
            if st.button("Continuar →", type="primary"):
                st.session_state.paso_actual = 3
                st.rerun()

# PASO 3
elif st.session_state.paso_actual == 3:
    st.title("📄 Generar Formulario 210")

    if not st.session_state.datos_exogena or not st.session_state.datos_cedular:
        st.warning("⚠️ Completa los pasos anteriores")
    else:
        datos_personales = st.session_state.datos_exogena['contribuyente'].copy()

        with st.expander("📝 Editar Datos"):
            col1, col2 = st.columns(2)
            with col1:
                cedula = st.text_input("Cédula", value=datos_personales.get('cedula', ''), key="cedula_form")
            with col2:
                nombre = st.text_input("Nombre", value=datos_personales.get('nombre', ''), key="nombre_form")
            datos_personales['cedula'] = cedula
            datos_personales['nombre'] = nombre

        st.subheader("✅ Resumen")
        col1, col2 = st.columns(2)
        with col1:
            st.write(f"**Cédula:** {cedula}\n**Nombre:** {nombre}")
        with col2:
            datos_cedular = st.session_state.datos_cedular
            renta_total = datos_cedular[0]['renta_liquida'] + datos_cedular[1]['renta_liquida']
            st.write(f"**Renta Total:** ${renta_total:,.0f}\n**IRPF:** ${calcular_irpf_correcto(renta_total):,.0f}")

        if st.button("✅ Generar Formulario 210", type="primary", use_container_width=True):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "210"
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            ws['A1'] = "FORMULARIO 210 - DECLARACIÓN RENTA 2025"
            ws['A1'].font = Font(bold=True, size=14)

            row = 3
            ws[f'A{row}'] = "Cédula:"
            ws[f'B{row}'] = cedula
            row += 1
            ws[f'A{row}'] = "Nombre:"
            ws[f'B{row}'] = nombre
            row += 2

            for cedula_data in st.session_state.datos_cedular:
                ws[f'A{row}'] = cedula_data['tipo']
                ws[f'A{row}'].font = header_font
                ws[f'A{row}'].fill = header_fill
                row += 1

                if 'ingresos_trabajo' in cedula_data:
                    ws[f'A{row}'] = "Ingresos Trabajo"
                    ws[f'B{row}'] = cedula_data['ingresos_trabajo']
                    ws[f'B{row}'].number_format = '$#,##0'
                    row += 1
                    ws[f'A{row}'] = "Ingresos Capital"
                    ws[f'B{row}'] = cedula_data['ingresos_capital']
                    ws[f'B{row}'].number_format = '$#,##0'
                    row += 1
                    ws[f'A{row}'] = "Deducciones"
                    ws[f'B{row}'] = cedula_data['deducciones']
                    ws[f'B{row}'].number_format = '$#,##0'
                    row += 1
                    ws[f'A{row}'] = "Renta Líquida"
                    ws[f'B{row}'] = cedula_data['renta_liquida']
                    ws[f'B{row}'].number_format = '$#,##0'
                else:
                    ws[f'A{row}'] = "Ingresos"
                    ws[f'B{row}'] = cedula_data.get('ingresos', 0)
                    ws[f'B{row}'].number_format = '$#,##0'
                    row += 1
                    ws[f'A{row}'] = "Renta Líquida"
                    ws[f'B{row}'] = cedula_data.get('renta_liquida', cedula_data.get('ingresos', 0))
                    ws[f'B{row}'].number_format = '$#,##0'
                row += 2

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            st.download_button(
                label="📥 Descargar Formulario 210",
                data=buffer,
                file_name=f"Formulario_210_{cedula}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
            st.success("✅ ¡Formulario generado!")

        if st.button("← Volver"):
            st.session_state.paso_actual = 2
            st.rerun()
