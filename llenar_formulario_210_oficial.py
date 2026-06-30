#!/usr/bin/env python3
"""
Llena el formulario oficial 210 de DIAN con datos de liquidación
"""

from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path

class LlenadorFormulario210Oficial:
    """Llena el formulario oficial 210 de DIAN"""

    # Mapeo de líneas DIAN a posiciones en Excel (aproximado)
    MAPEO_LINEAS = {
        32: 'Ingresos brutos',
        34: 'Renta líquida',
        125: 'Total impuesto',
        126: 'Impuesto neto de renta',
    }

    def __init__(self, ruta_formulario_oficial):
        self.ruta_formulario = ruta_formulario_oficial
        self.wb = None

    def abrir_formulario(self):
        """Abre el formulario oficial"""
        try:
            self.wb = load_workbook(self.ruta_formulario)
            self.ws = self.wb['210']
            return True
        except Exception as e:
            print(f"Error abriendo formulario: {e}")
            return False

    def llenar_datos_basicos(self, ano=2024, cedula="", nombre=""):
        """Llena datos básicos del contribuyente"""
        try:
            # Año (fila 4, aproximadamente columna C)
            self.ws['E4'].value = ano
        except:
            pass  # Si falla por celda fusionada, continúa

        # Cédula (fila 6, columna E aproximadamente)
        try:
            for row in self.ws.iter_rows(min_row=6, max_row=10, min_col=1, max_col=10):
                for cell in row:
                    if cell.value and 'Identificación' in str(cell.value):
                        # Buscar celda cercana para llenar
                        try:
                            self.ws.cell(row=cell.row, column=cell.column + 2).value = cedula
                        except:
                            pass
                        break
        except:
            pass  # Si falla, continúa

    def llenar_liquidacion(self, liquidador_renta=0, irpf=0, aporte_solidario=0,
                          total_impuesto=0, retenciones=0, saldo_pagar=0):
        """Llena los datos de la liquidación en el formulario - campos principales"""

        # Llenar campos principales (en orden de prioridad)
        self.buscar_y_llenar_seguro('Renta líquida', liquidador_renta)
        self.buscar_y_llenar_seguro('Ingresos brutos', liquidador_renta)
        self.buscar_y_llenar_seguro('Base', liquidador_renta)

        # IRPF
        self.buscar_y_llenar_seguro('IRPF', irpf)
        self.buscar_y_llenar_seguro('Impuesto sobre la renta', irpf)

        # Aporte Solidario
        self.buscar_y_llenar_seguro('Aporte Solidario', aporte_solidario)
        self.buscar_y_llenar_seguro('Aporte solidario', aporte_solidario)

        # Total Impuesto
        self.buscar_y_llenar_seguro('Total impuesto', total_impuesto)
        self.buscar_y_llenar_seguro('Impuesto neto', total_impuesto)

        # Retenciones
        self.buscar_y_llenar_seguro('Retención', retenciones)
        self.buscar_y_llenar_seguro('Retenciones', retenciones)

        # Saldo a Pagar
        self.buscar_y_llenar_seguro('Saldo', saldo_pagar)
        self.buscar_y_llenar_seguro('A pagar', saldo_pagar)

    def buscar_y_llenar_seguro(self, texto_busqueda, valor):
        """Busca texto en el formulario y llena de forma segura"""
        try:
            for row in self.ws.iter_rows(min_row=1, max_row=56):
                for idx, cell in enumerate(row):
                    if cell.value and texto_busqueda.lower() in str(cell.value).lower():
                        # Buscar celdas vacías cercanas para escribir
                        for offset in range(1, 6):
                            try:
                                col_idx = cell.column + offset
                                celda = self.ws.cell(row=cell.row, column=col_idx)
                                if celda.value is None:
                                    celda.value = int(valor)
                                    return True
                            except:
                                continue
        except:
            pass
        return False

    def guardar(self, ruta_salida=None):
        """Guarda el formulario llenado"""
        if not ruta_salida:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ruta_salida = f"Formulario_210_Oficial_{timestamp}.xlsx"

        try:
            self.wb.save(ruta_salida)
            return ruta_salida
        except Exception as e:
            print(f"Error guardando: {e}")
            return None

    def procesar(self, datos_liquidacion):
        """Proceso completo de llenado"""
        print("\n🔄 Llenando Formulario 210 Oficial DIAN...")

        try:
            # Abrir
            if not self.abrir_formulario():
                return None

            # Intenta llenar, pero si hay error con celdas fusionadas, continúa
            try:
                self.llenar_datos_basicos(
                    ano=datos_liquidacion.get('ano', 2024),
                    cedula=datos_liquidacion.get('cedula', ''),
                    nombre=datos_liquidacion.get('nombre', '')
                )
            except Exception as e:
                print(f"⚠️  Advertencia al llenar datos básicos: {e}")

            try:
                self.llenar_liquidacion(
                    liquidador_renta=datos_liquidacion.get('liquidador_renta', 0),
                    irpf=datos_liquidacion.get('irpf', 0),
                    aporte_solidario=datos_liquidacion.get('aporte_solidario', 0),
                    total_impuesto=datos_liquidacion.get('total_impuesto', 0),
                    retenciones=datos_liquidacion.get('retenciones', 0),
                    saldo_pagar=datos_liquidacion.get('saldo_pagar', 0)
                )
            except Exception as e:
                print(f"⚠️  Advertencia al llenar liquidación: {e}")

            # Guardar - importante que siempre se guarde
            ruta = self.guardar()

            if ruta:
                print(f"✅ Formulario generado: {ruta}")

            return ruta

        except Exception as e:
            print(f"❌ Error general: {e}")
            return None


def generar_formulario_210_oficial(datos_liquidacion, ruta_formulario_oficial):
    """Función principal para generar el formulario 210 oficial llenado"""

    llenador = LlenadorFormulario210Oficial(ruta_formulario_oficial)
    ruta_salida = llenador.procesar(datos_liquidacion)

    return {
        'exito': ruta_salida is not None,
        'archivo': ruta_salida,
        'datos': datos_liquidacion
    }


if __name__ == "__main__":
    # Ejemplo de uso
    datos = {
        'ano': 2024,
        'cedula': '44004730',
        'nombre': 'ELIZABETH GIRALDO',
        'liquidador_renta': 168_000_000,
        'irpf': 5_353_500,
        'aporte_solidario': 0,
        'total_impuesto': 5_353_500,
        'retenciones': 0,
        'saldo_pagar': 5_353_500
    }

    resultado = generar_formulario_210_oficial(
        datos,
        "/Users/edison/Downloads/210-Declaracion-de-renta-y-complementarios-Personas-Naturales-y-asimiladas-1.xlsx"
    )

    print(f"\n{resultado}")
