# REBUILD FORCE v2
from flask import Flask, request, send_file, render_template, jsonify, redirect, session
import io
import json
import traceback
from datetime import datetime
from functools import wraps
from converter import process_file, process_liquidacion_iva, _cargar_nomina_balance, convert_nomina, generate_balance_prueba, generate_liquidaciones
from siigo_integration import subir_planos_a_siigo
from dian_integration import descargar_reporte_dian
from liquidacion_pdf_premium import generar_liquidacion_pdf_premium
from retencion import calcular_retencion, generar_reporte_retenciones_excel
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024  # 20 MB máximo
app.secret_key = 'dev-key-change-in-production'

# Usuarios válidos (cambiar en producción)
USUARIOS_VALIDOS = {
    'ediandres@hotmail.com': '12345678'
}


def require_login(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('autenticado'):
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
@require_login
def index():
    return render_template('app.html')



@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de login"""
    if request.method == 'POST':
        usuario = request.form.get('usuario', '').strip()
        password = request.form.get('password', '').strip()
        
        if usuario in USUARIOS_VALIDOS and USUARIOS_VALIDOS[usuario] == password:
            from flask import session
            session['usuario'] = usuario
            session['autenticado'] = True
            return jsonify({'success': True, 'mensaje': 'Login exitoso'}), 200
        else:
            return jsonify({'success': False, 'error': 'Usuario o contraseña incorrectos'}), 401
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Cerrar sesión"""
    from flask import session
    session.clear()
    return redirect('/login')



@app.route('/api/test-validar', methods=['POST'])
def test_validar():
    """Test simple de validación"""
    return jsonify({
        'status': 'ok',
        'mensaje': 'Test funcionando',
        'timestamp': str(datetime.now())
    }), 200

@app.route('/legacy')
def legacy_index():
    return render_template('index.html')

@app.route('/api/convert', methods=['POST'])
def convert():
    if 'file' not in request.files:
        return {'error': 'No se proporcionó ningún archivo'}, 400
        
    file = request.files['file']
    if file.filename == '':
        return {'error': 'Archivo no seleccionado'}, 400
        
    try:
        consec_compras = int(request.form.get('consec_compras', 371))
        consec_nc = int(request.form.get('consec_nc', 271))
        consec_gastos = int(request.form.get('consec_gastos', 798))
        consec_ventas    = int(request.form.get('consec_ventas', 1))
        consec_nc_ventas = int(request.form.get('consec_nc_ventas', 1))
    except ValueError:
        return {'error': 'Los consecutivos deben ser números válidos'}, 400

    try:
        # Leer el archivo en memoria
        input_stream = io.BytesIO(file.read())

        # Procesar el archivo
        output_stream = process_file(
            input_stream,
            consec_compras=consec_compras,
            consec_nc=consec_nc,
            consec_gastos=consec_gastos,
            consec_ventas=consec_ventas,
            consec_nc_ventas=consec_nc_ventas
        )

        original_name = file.filename.rsplit('.', 1)[0]
        download_name = f"PLANOS_{original_name}.xlsx"

        return send_file(
            output_stream,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        traceback.print_exc()
        return {'error': f'Error procesando el archivo: {str(e)}'}, 500

@app.route('/api/liquidar-iva', methods=['POST'])
def liquidar_iva():
    if 'file' not in request.files:
        return {'error': 'No se proporcionó ningún archivo'}, 400

    file = request.files['file']
    if file.filename == '':
        return {'error': 'Archivo no seleccionado'}, 400

    try:
        input_stream = io.BytesIO(file.read())
        output_stream = process_liquidacion_iva(input_stream)

        original_name = file.filename.rsplit('.', 1)[0]
        download_name = f"LIQUIDACION_IVA_{original_name}.xlsx"

        return send_file(
            output_stream,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        traceback.print_exc()
        return {'error': f'Error generando liquidación: {str(e)}'}, 500

@app.route('/api/liquidar-retenciones', methods=['POST'])
def liquidar_retenciones():
    if 'file' not in request.files:
        return {'error': 'No se proporcionó ningún archivo'}, 400

    file = request.files['file']
    if file.filename == '':
        return {'error': 'Archivo no seleccionado'}, 400

    try:
        import openpyxl

        # Leer el archivo DIAN
        file_data = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)

        # Extraer documentos del DIAN
        planos_data = {'compras': [], 'ventas': [], 'gastos': []}

        # Buscar la hoja de reportes
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Detectar si es un reporte DIAN o planos
            primera_celda = str(ws.cell(1, 1).value or '').strip()

            if 'REPORTE' in sheet_name.upper() or 'Rp_Doc' in sheet_name:
                # Es un reporte DIAN - extraer documentos
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) < 13:
                        continue

                    nit_emisor = str(row[9] or '').strip() if len(row) > 9 else ''
                    nombre_emisor = str(row[10] or '').strip() if len(row) > 10 else ''
                    nit_receptor = str(row[11] or '').strip() if len(row) > 11 else ''
                    nombre_receptor = str(row[12] or '').strip() if len(row) > 12 else ''
                    folio = str(row[2] or '').strip() if len(row) > 2 else ''
                    fecha = row[7] if len(row) > 7 else ''
                    total = float(row[29]) if len(row) > 29 and row[29] else 0
                    iva = float(row[13]) if len(row) > 13 and row[13] else 0
                    grupo = str(row[31] or '').strip() if len(row) > 31 else ''

                    if not total or not folio:
                        continue

                    base = round(total - iva, 2)
                    doc = {
                        'folio': folio,
                        'fecha': fecha,
                        'nit': nit_emisor if grupo == 'Recibido' else nit_receptor,
                        'nombre': nombre_emisor if grupo == 'Recibido' else nombre_receptor,
                        'total': total,
                        'iva': iva
                    }

                    # Clasificar
                    if grupo == 'Recibido':
                        planos_data['compras'].append(doc)
                    elif grupo == 'Emitido':
                        planos_data['ventas'].append(doc)

        # Generar reporte de retenciones
        wb = generar_reporte_retenciones_excel(planos_data)
        output_stream = io.BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)

        original_name = file.filename.rsplit('.', 1)[0]
        download_name = f"RETENCIONES_{original_name}.xlsx"

        return send_file(
            output_stream,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        traceback.print_exc()
        return {'error': f'Error generando retenciones: {str(e)}'}, 500

@app.route('/api/nomina', methods=['POST'])
def nomina():
    if 'balance' not in request.files:
        return {'error': 'Se requiere el archivo de Balance de prueba por tercero'}, 400

    balance_file = request.files['balance']
    if balance_file.filename == '':
        return {'error': 'Archivo de balance no seleccionado'}, 400

    try:
        consec_nomina = int(request.form.get('consec_nomina', 1))
    except ValueError:
        return {'error': 'El consecutivo debe ser un número válido'}, 400

    try:
        import tempfile, os
        # Guardar balance en archivo temporal (openpyxl necesita path real para read_only)
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            balance_file.save(tmp)
            tmp_path = tmp.name

        from openpyxl import Workbook
        empleados = _cargar_nomina_balance(tmp_path)
        os.unlink(tmp_path)

        wb_dst = Workbook()
        if wb_dst.active:
            wb_dst.remove(wb_dst.active)

        convert_nomina(empleados, wb_dst, consec_nomina)

        output_stream = io.BytesIO()
        wb_dst.save(output_stream)
        output_stream.seek(0)

        return send_file(
            output_stream,
            as_attachment=True,
            download_name='PLANO_NOMINA.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        traceback.print_exc()
        return {'error': f'Error generando nómina: {str(e)}'}, 500

@app.route('/api/balance', methods=['POST'])
def balance():
    if 'file' not in request.files:
        return {'error': 'No se proporcionó ningún archivo'}, 400
    file = request.files['file']
    if file.filename == '':
        return {'error': 'Archivo no seleccionado'}, 400
    try:
        input_stream = io.BytesIO(file.read())
        output_stream = generate_balance_prueba(input_stream)
        original_name = file.filename.rsplit('.', 1)[0]
        download_name = f"BALANCE_TERCERO_{original_name}.xlsx"
        return send_file(
            output_stream,
            as_attachment=True,
            download_name=download_name,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        traceback.print_exc()
        return {'error': f'Error generando balance: {str(e)}'}, 500

@app.route('/api/download-dian', methods=['POST'])
def download_dian():
    """Descarga reporte DIAN automáticamente."""
    try:
        usuario = request.json.get('usuario', '')
        password = request.json.get('password', '')

        if not usuario or not password:
            return {'error': 'Credenciales DIAN requeridas'}, 400

        # Intentar descargar
        archivo = descargar_reporte_dian(usuario, password)

        if not archivo:
            return {
                'error': 'No se pudo descargar automáticamente.',
                'instrucciones': [
                    '1. Ve a https://catalogo-vpfe.dian.gov.co/User/PersonLogin',
                    '2. Ingresa tus credenciales',
                    '3. Menú: "Consulta de Documentos" → "Buscar Documento"',
                    '4. Filtra por período (ej: 2025)',
                    '5. Clickea "Descargar Reporte" (Excel)',
                    '6. Sube el archivo aquí en "Selecciona archivo"'
                ]
            }, 400

        return send_file(
            archivo,
            as_attachment=True,
            download_name=f'Reporte_DIAN_{datetime.now().strftime("%Y%m%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        traceback.print_exc()
        return {'error': f'Error descargando: {str(e)}'}, 500

@app.route('/api/upload-siigo', methods=['POST'])
def upload_siigo():
    """Genera planos desde DIAN y los sube a Siigo Nube automáticamente."""
    if 'file' not in request.files:
        return {'error': 'No se proporcionó archivo DIAN'}, 400

    file = request.files['file']
    if file.filename == '':
        return {'error': 'Archivo no seleccionado'}, 400

    try:
        username = request.form.get('siigo_user', '')
        password = request.form.get('siigo_pass', '')

        if not username or not password:
            return {'error': 'Credenciales de Siigo requeridas'}, 400

        # Parámetros de consecutivos
        consec_compras = int(request.form.get('consec_compras', 371))
        consec_nc = int(request.form.get('consec_nc', 271))
        consec_gastos = int(request.form.get('consec_gastos', 798))
        consec_ventas = int(request.form.get('consec_ventas', 1))
        consec_nc_ventas = int(request.form.get('consec_nc_ventas', 1))

        # Paso 1: Procesar archivo DIAN y generar planos
        file_data = io.BytesIO(file.read())
        planos_stream = process_file(
            file_data,
            consec_compras=consec_compras,
            consec_nc=consec_nc,
            consec_gastos=consec_gastos,
            consec_ventas=consec_ventas,
            consec_nc_ventas=consec_nc_ventas
        )
        planos_stream.seek(0)

        # Paso 2: Subir planos a Siigo
        resultado = subir_planos_a_siigo(planos_stream, username, password)

        if 'error' in resultado:
            return resultado, 400

        return {
            'success': True,
            'mensaje': f'✅ Planos generados y {resultado["exito"]} registros subidos a Siigo',
            'detalles': resultado
        }, 200

    except Exception as e:
        traceback.print_exc()
        return {'error': f'Error: {str(e)}'}, 500

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 8080))
    app.run(debug=True, port=port, host='127.0.0.1')

@app.route('/api/retenciones', methods=['POST'])
def calcular_retenciones():
    """Calcula retenciones de un monto"""
    try:
        data = request.json
        base = float(data.get('base', 0))
        tipo = data.get('tipo', 'compras')  # compras, ventas, gastos
        nit = data.get('nit', '')
        
        resultado = calcular_retencion(base, tipo, nit)
        return resultado
    except Exception as e:
        return {'error': str(e)}, 400

# ===== LIQUIDACIONES MEJORADAS =====

@app.route('/api/validar-liquidacion', methods=['POST'])
def validar_liquidacion():
    """Valida el Excel de liquidación sin procesar"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        if not file or file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        try:
            input_stream = io.BytesIO(file.read())
            wb = openpyxl.load_workbook(input_stream, data_only=False)
        except Exception as e:
            return jsonify({'error': f'Archivo Excel inválido: {str(e)}'}), 400

        # Validación flexible
        if not wb.sheetnames:
            return jsonify({'error': 'El archivo no tiene hojas'}), 400

        ws = wb.active
        if ws.max_row < 2:
            return jsonify({'error': 'El archivo está vacío'}), 400

        empleados_count = ws.max_row - 1

        return jsonify({
            'valido': True,
            'empleados_count': empleados_count,
            'mensaje': f'Excel válido con {empleados_count} registros'
        }), 200

    except Exception as e:
        import traceback
        return jsonify({'error': f'Error: {str(e)}', 'details': traceback.format_exc()}), 500

@app.route('/api/previsualizar-liquidacion/<int>/<nombre>', methods=['POST'])
def previsualizar_liquidacion(index, nombre):
    """Previsualiza PDF de un empleado específico"""
    try:
        if 'file' not in request.files:
            return {'error': 'No file uploaded'}, 400
        
        file = request.files['file']
        input_stream = io.BytesIO(file.read())
        wb = openpyxl.load_workbook(input_stream)
        
        ws_emp = wb['Empleados']
        
        # Obtener datos del empleado en la fila especificada
        row = index + 2  # +2 porque row 1 es header, índice es 0-based
        if row > ws_emp.max_row:
            return {'error': f'Empleado {index} no existe'}, 400
        
        nombre_emp = ws_emp.cell(row, 1).value
        documento = ws_emp.cell(row, 2).value
        salario = ws_emp.cell(row, 3).value
        dias = ws_emp.cell(row, 4).value
        cesantias = ws_emp.cell(row, 5).value or 0
        vacaciones = ws_emp.cell(row, 6).value or 0
        
        # Obtener parámetros
        ws_param = wb['Parametros']
        params = {}
        for i in range(2, ws_param.max_row + 1):
            param_name = ws_param.cell(i, 1).value
            param_value = ws_param.cell(i, 2).value
            if param_name:
                params[param_name] = param_value
        
        # Generar datos de liquidación
        datos_emp = {
            'nombre': nombre_emp,
            'documento': documento,
            'salario_mensual': salario,
            'dias_laborados': dias
        }
        
        datos_liq = {
            'cesantias': cesantias,
            'vacaciones': vacaciones,
            'salario_basico': salario * dias / 30
        }
        
        # Generar PDF
        empresa_data = {
            'nombre': request.json.get('empresa_nombre', 'Mi Empresa') if request.is_json else 'Mi Empresa',
            'nit': request.json.get('empresa_nit', '') if request.is_json else ''
        }
        
        pdf_stream = generar_liquidacion_pdf_premium(datos_emp, datos_liq, None, empresa_data)
        
        return send_file(
            pdf_stream,
            as_attachment=True,
            download_name=f'Preview_Liquidacion_{nombre}.pdf',
            mimetype='application/pdf'
        )
    
    except Exception as e:
        return {'error': f'Error en previsualización: {str(e)}'}, 500

def listar_empleados_liquidar():
    """Lista empleados para descargar PDFs individuales"""
    try:
        if 'file' not in request.files:
            return {'error': 'No file uploaded'}, 400
        
        file = request.files['file']
        input_stream = io.BytesIO(file.read())
        wb = openpyxl.load_workbook(input_stream)
        
        ws_emp = wb['Empleados']
        empleados = []
        
        for row in range(2, ws_emp.max_row + 1):
            nombre = ws_emp.cell(row, 1).value
            documento = ws_emp.cell(row, 2).value
            
            if nombre:
                empleados.append({
                    'id': row - 2,
                    'nombre': nombre,
                    'documento': documento
                })
        
        return {'empleados': empleados}, 200
    
    except Exception as e:
        return {'error': f'Error: {str(e)}'}, 400

@app.route('/api/descargar-liquidacion/<int:emp_id>', methods=['POST'])
def descargar_liquidacion(emp_id):
    """Descarga PDF de un empleado específico"""
    try:
        if 'file' not in request.files:
            return {'error': 'No file uploaded'}, 400
        
        file = request.files['file']
        input_stream = io.BytesIO(file.read())
        wb = openpyxl.load_workbook(input_stream)
        
        ws_emp = wb['Empleados']
        ws_param = wb['Parametros']
        
        row = emp_id + 2
        if row > ws_emp.max_row:
            return {'error': 'Empleado no existe'}, 400
        
        nombre = ws_emp.cell(row, 1).value
        documento = ws_emp.cell(row, 2).value
        salario = ws_emp.cell(row, 3).value
        dias = ws_emp.cell(row, 4).value
        cesantias = ws_emp.cell(row, 5).value or 0
        vacaciones = ws_emp.cell(row, 6).value or 0
        
        # Obtener parámetros
        params = {}
        for i in range(2, ws_param.max_row + 1):
            param_name = ws_param.cell(i, 1).value
            param_value = ws_param.cell(i, 2).value
            if param_name:
                params[param_name] = param_value
        
        datos_emp = {
            'nombre': nombre,
            'documento': documento,
            'salario_mensual': salario,
            'dias_laborados': dias
        }
        
        datos_liq = {
            'cesantias': cesantias,
            'vacaciones': vacaciones,
            'salario_basico': salario * dias / 30 if salario and dias else 0
        }
        
        empresa_data = {
            'nombre': request.form.get('empresa_nombre', 'Mi Empresa'),
            'nit': request.form.get('empresa_nit', '')
        }
        
        pdf_stream = generar_liquidacion_pdf_premium(
            datos_emp, datos_liq, None, empresa_data
        )
        
        return send_file(
            pdf_stream,
            as_attachment=True,
            download_name=f'Liquidacion_{nombre.replace(" ", "_")}.pdf',
            mimetype='application/pdf'
        )
    
    except Exception as e:
        return {'error': f'Error: {str(e)}'}, 500

