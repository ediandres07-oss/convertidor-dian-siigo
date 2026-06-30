#!/usr/bin/env python3
"""
Llenador Avanzado de Formulario 210 - DIAN
Mapeo completo y automático de campos
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import json

class LlenadorFormulario210Completo:
    """Mapea y llena automáticamente el Formulario 210 DIAN"""

    # Mapeo de líneas DIAN a posiciones Excel (aproximado, depende del template)
    MAPEO_LINEAS_DIAN = {
        '1': 'Año',
        '5': 'NIT/Cédula',
        '10': 'Patrimonio Bruto',
        '110': 'Retención en la Fuente',
        '200': 'Ingresos por Servicios/Honorarios',
        '300': 'Rendimientos Financieros',
        '400': 'Ingresos por Arrendamiento',
        '500': 'Ganancia en Venta de Bienes',
        '600': 'Otros Ingresos',
    }

    def __init__(self, ruta_formulario, ruta_exogena):
        self.ruta_formulario = ruta_formulario
        self.ruta_exogena = ruta_exogena
        self.wb = None
        self.ws = None
        self.datos = {}

    def extraer_exogena(self):
        """Extrae todos los datos de la Exógena"""
        df = pd.read_excel(self.ruta_exogena, sheet_name='Reporte', header=None)

        # Extraer contribuyente
        for i in range(len(df)):
            row = df.iloc[i]
            if pd.notna(row.iloc[0]):
                txt = str(row.iloc[0]).lower()
                if 'identificación:' in txt and pd.notna(row.iloc[2]):
                    self.datos['cedula'] = str(row.iloc[2])
                elif 'nombres' in txt and pd.notna(row.iloc[2]):
                    self.datos['nombre'] = str(row.iloc[2])
                elif 'año' in txt and 'refiere' in txt and pd.notna(row.iloc[2]):
                    self.datos['ano'] = int(row.iloc[2])

        # Extraer ingresos y retenciones por categoría
        self.datos['ingresos'] = {}
        self.datos['retenciones'] = {}

        for i in range(14, len(df)):
            row = df.iloc[i]
            if pd.notna(row.iloc[4]) and pd.notna(row.iloc[5]):
                concepto = str(row.iloc[4]).strip()
                try:
                    valor = float(row.iloc[5])
                    if valor > 0:
                        uso = str(row.iloc[6]).lower() if pd.notna(row.iloc[6]) else ""

                        # Clasificar automáticamente
                        if 'retención' in uso or 'r132' in uso or 'r28' in uso or 'cdt retención' in concepto.lower():
                            if concepto not in self.datos['retenciones']:
                                self.datos['retenciones'][concepto] = 0
                            self.datos['retenciones'][concepto] += valor
                        elif 'patrimonio' in concepto.lower():
                            cat = '10'  # Patrimonio
                        elif 'cdt' in concepto.lower() or 'rendimiento' in concepto.lower():
                            cat = '300'  # Rendimientos
                        elif 'aporte' in concepto.lower() or 'pensión' in concepto.lower():
                            cat = '35'   # Rentas exentas
                        elif 'salario' in concepto.lower():
                            cat = '200'  # Ingresos laborales
                        elif 'factura' in concepto.lower():
                            cat = '200'  # Ingresos servicios
                        else:
                            cat = '600'  # Otros

                        if cat not in self.datos['ingresos']:
                            self.datos['ingresos'][cat] = {'valor': 0, 'conceptos': []}
                        self.datos['ingresos'][cat]['valor'] += valor
                        self.datos['ingresos'][cat]['conceptos'].append(concepto)
                except:
                    pass

        # Calcular totales
        self.datos['total_ingresos'] = sum(item['valor'] for item in self.datos['ingresos'].values())
        self.datos['total_retenciones'] = sum(self.datos['retenciones'].values())

        return self.datos

    def abrir_formulario(self):
        """Abre el archivo del formulario 210"""
        try:
            self.wb = load_workbook(self.ruta_formulario)
            self.ws = self.wb.active
            return True
        except Exception as e:
            print(f"❌ Error: {e}")
            return False

    def buscar_y_llenar(self, texto_busqueda, valor, offset_row=1, offset_col=2):
        """Busca texto en el formulario y llena celda cercana"""
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and texto_busqueda in str(cell.value):
                    # Buscar celda válida (no fusionada)
                    for attempt_col in range(cell.column, cell.column + 5):
                        for attempt_row in range(cell.row + offset_row, cell.row + offset_row + 3):
                            try:
                                destino = self.ws.cell(row=attempt_row, column=attempt_col)
                                # Verificar que no es celda fusionada
                                if destino.value is not None:
                                    continue
                                destino.value = valor
                                # Formato
                                destino.alignment = Alignment(horizontal='right')
                                if isinstance(valor, (int, float)):
                                    destino.number_format = '#,##0'
                                return True
                            except AttributeError:
                                # Celda fusionada, pasar a la siguiente
                                continue
        return False

    def llenar_estructura_completa(self):
        """Llena el formulario con estructura completa"""
        print("\n📋 LLENANDO FORMULARIO 210")
        print("="*70)

        # 1. Datos del declarante
        print("\n1️⃣  DATOS DEL DECLARANTE")
        self.buscar_y_llenar("Año", self.datos.get('ano', 2024))
        print(f"   ✓ Año: {self.datos.get('ano')}")

        self.buscar_y_llenar("Número de Identificación", self.datos.get('cedula', ''))
        print(f"   ✓ Cédula: {self.datos.get('cedula')}")

        # 2. Sección de Ingresos
        print("\n2️⃣  INGRESOS BRUTOS")
        total = self.datos.get('total_ingresos', 0)
        self.buscar_y_llenar("Ingresos brutos", total)
        print(f"   ✓ Total Ingresos: ${total:,.0f}")

        # Por categoría
        for linea, info in sorted(self.datos['ingresos'].items()):
            linea_nombre = self.MAPEO_LINEAS_DIAN.get(str(linea), f'Línea {linea}')
            print(f"   ✓ {linea_nombre}: ${info['valor']:,.0f}")

        # 3. Retenciones
        print("\n3️⃣  RETENCIONES EN LA FUENTE")
        total_ret = self.datos.get('total_retenciones', 0)
        self.buscar_y_llenar("Retención", total_ret)
        if total_ret > 0:
            print(f"   ✓ Total Retenciones: ${total_ret:,.0f}")
        else:
            print(f"   ℹ️  No hay retenciones reportadas")

        # 4. Cálculos
        print("\n4️⃣  CÁLCULOS TRIBUTARIOS")
        deduccion = total * 0.10
        base_liquida = total - deduccion

        print(f"   ✓ Deducción (10%): ${deduccion:,.0f}")
        print(f"   ✓ Base Líquida: ${base_liquida:,.0f}")

        # IRPF progresivo
        irpf = self._calcular_irpf(base_liquida)
        print(f"   ✓ IRPF: ${irpf:,.0f}")

        # Aporte solidario
        aporte_sol = self._calcular_aporte_solidario(base_liquida)
        if aporte_sol > 0:
            print(f"   ✓ Aporte Solidario: ${aporte_sol:,.0f}")

        total_impuesto = irpf + aporte_sol
        saldo = max(0, total_impuesto - total_ret)

        print(f"   ✓ Total Impuesto: ${total_impuesto:,.0f}")
        print(f"   ✓ Saldo a Pagar: ${saldo:,.0f}")

        self.datos['calculos'] = {
            'deduccion': deduccion,
            'base_liquida': base_liquida,
            'irpf': irpf,
            'aporte_solidario': aporte_sol,
            'total_impuesto': total_impuesto,
            'saldo': saldo
        }

    def _calcular_irpf(self, base):
        """Calcula IRPF progresivo 2024"""
        rangos = [
            (0, 66950000, 0),
            (66950000, 134900000, 0.05),
            (134900000, 404700000, 0.12),
            (404700000, 673500000, 0.25),
            (673500000, 1347000000, 0.32),
            (1347000000, float('inf'), 0.37),
        ]
        irpf = 0
        for inicio, fin, tasa in rangos:
            if base > inicio:
                tramo = min(base, fin) - inicio
                irpf += tramo * tasa
        return irpf

    def _calcular_aporte_solidario(self, base):
        """Calcula aporte solidario (1% > 16.000 UVT)"""
        UVT = 44654
        limite = 16000 * UVT
        if base > limite:
            return (base - limite) * 0.01
        return 0

    def guardar(self, ruta_salida=None):
        """Guarda el formulario completo"""
        if not ruta_salida:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            cedula = self.datos.get('cedula', 'XXX')
            ruta_salida = f"Formulario210_Elizabeth_{cedula}_{timestamp}.xlsx"

        self.wb.save(ruta_salida)
        return ruta_salida

    def generar_resumen_json(self):
        """Genera JSON con los resultados"""
        return {
            'contribuyente': {
                'cedula': self.datos.get('cedula'),
                'nombre': self.datos.get('nombre'),
                'ano_gravable': self.datos.get('ano'),
            },
            'ingresos_por_categoria': {
                'linea_' + k: v['valor'] for k, v in self.datos.get('ingresos', {}).items()
            },
            'total_ingresos': self.datos.get('total_ingresos'),
            'retenciones': self.datos.get('total_retenciones'),
            'calculos': self.datos.get('calculos', {}),
        }

    def procesar_completo(self):
        """Proceso completo: extraer, mapear, llenar y guardar"""
        print("\n╔════════════════════════════════════════════════════════════╗")
        print("║  LLENADOR AVANZADO - FORMULARIO 210 COMPLETO              ║")
        print("║  Colombia DIAN - 2024-2025                               ║")
        print("╚════════════════════════════════════════════════════════════╝")

        # Paso 1: Extraer exógena
        print("\n📂 Paso 1: Extrayendo Exógena...")
        self.extraer_exogena()
        print(f"   ✓ {self.datos.get('nombre')} (CC {self.datos.get('cedula')})")
        print(f"   ✓ {len(self.datos['ingresos'])} categorías de ingresos")
        print(f"   ✓ Total: ${self.datos.get('total_ingresos', 0):,.0f}")

        # Paso 2: Abrir formulario
        print("\n📋 Paso 2: Abriendo Formulario 210...")
        if not self.abrir_formulario():
            return None

        # Paso 3: Llenar
        print("\n✏️  Paso 3: Rellenando formulario...")
        self.llenar_estructura_completa()

        # Paso 4: Guardar
        print("\n💾 Paso 4: Guardando...")
        ruta = self.guardar()
        print(f"   ✓ Archivo: {ruta}")

        # Paso 5: Resumen
        print("\n" + "="*70)
        print("📊 RESUMEN FINAL")
        print("="*70)
        calc = self.datos.get('calculos', {})
        print(f"Ingresos Brutos:      ${self.datos.get('total_ingresos', 0):>17,.0f}")
        print(f"Retenciones:          ${self.datos.get('total_retenciones', 0):>17,.0f}")
        print(f"Deducción (10%):      ${calc.get('deduccion', 0):>17,.0f}")
        print(f"─" * 50)
        print(f"Base Líquida:         ${calc.get('base_liquida', 0):>17,.0f}")
        print(f"IRPF:                 ${calc.get('irpf', 0):>17,.0f}")
        print(f"Aporte Solidario:     ${calc.get('aporte_solidario', 0):>17,.0f}")
        print(f"─" * 50)
        print(f"Total Impuesto:       ${calc.get('total_impuesto', 0):>17,.0f}")
        print(f"Menos Retenciones:    ${self.datos.get('total_retenciones', 0):>17,.0f}")
        print(f"═" * 50)
        print(f"SALDO A PAGAR:        ${calc.get('saldo', 0):>17,.0f}")
        print("="*70)

        print(f"\n✅ Proceso completado exitosamente!")
        print(f"📄 Archivo: {ruta}")

        return {
            'archivo': ruta,
            'datos': self.generar_resumen_json(),
            'calculos': calc
        }

if __name__ == "__main__":
    # Ejecutar
    llenador = LlenadorFormulario210Completo(
        ruta_formulario="/Users/edison/Downloads/formulario 210.xlsx",
        ruta_exogena="/Users/edison/Downloads/reporteExogena2024Elizabeth.xlsx"
    )

    resultado = llenador.procesar_completo()

    # Guardar JSON
    if resultado:
        with open('resultado_formulario_210.json', 'w', encoding='utf-8') as f:
            import json
            json.dump(resultado['datos'], f, indent=2, ensure_ascii=False, default=str)
        print(f"\n📋 JSON guardado en: resultado_formulario_210.json")
