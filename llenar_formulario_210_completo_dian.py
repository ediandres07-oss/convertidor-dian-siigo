#!/usr/bin/env python3
"""
Llenador Formulario 210 - ULTRA SIMPLE
Escribe directamente sin buscar
"""

from openpyxl import load_workbook
from datetime import datetime

class LlenadorFormulario210CompletoDIAN:
    """Llena formulario 210 de forma directa"""

    def __init__(self, ruta_formulario_oficial):
        self.ruta_formulario = ruta_formulario_oficial
        self.wb = None
        self.ws = None

    def abrir_formulario(self):
        """Abre el formulario oficial"""
        try:
            self.wb = load_workbook(self.ruta_formulario)
            self.ws = self.wb['210']
            print("✓ Formulario abierto")
            return True
        except Exception as e:
            print(f"✗ Error: {e}")
            return False

    def llenar_datos_completos(self, datos):
        """Escribe valores directamente en el formulario"""

        print("\n📝 Escribiendo valores en formulario...")

        # Escribir en las primeras filas (área de datos del contribuyente y ingresos)
        valores = [
            (5, 2, str(datos.get('cedula', ''))),
            (6, 2, str(datos.get('nombre', ''))),
            (10, 2, datos.get('rentas_trabajo', 0)),
            (15, 2, datos.get('rentas_capital', 0)),
            (25, 2, datos.get('rentas_no_laborales', 0)),
            (34, 2, datos.get('liquidador_renta', 0)),
            (40, 2, datos.get('irpf', 0)),
            (45, 2, datos.get('aporte_solidario', 0)),
            (50, 2, datos.get('retenciones', 0)),
            (55, 2, datos.get('total_impuesto', 0)),
            (60, 2, datos.get('saldo_pagar', 0)),
        ]

        for fila, col, valor in valores:
            try:
                celda = self.ws.cell(row=fila, column=col)
                celda.value = valor
                print(f"  ✓ Celda B{fila}: {valor}")
            except Exception as e:
                print(f"  ✗ Error B{fila}: {e}")

        # También escribir en columnas alternas para asegurar
        valores_alt = [
            (10, 4, datos.get('rentas_trabajo', 0)),
            (15, 4, datos.get('rentas_capital', 0)),
            (25, 4, datos.get('rentas_no_laborales', 0)),
            (34, 4, datos.get('liquidador_renta', 0)),
            (40, 4, datos.get('irpf', 0)),
            (45, 4, datos.get('aporte_solidario', 0)),
            (50, 4, datos.get('retenciones', 0)),
            (55, 4, datos.get('total_impuesto', 0)),
            (60, 4, datos.get('saldo_pagar', 0)),
        ]

        for fila, col, valor in valores_alt:
            try:
                celda = self.ws.cell(row=fila, column=col)
                celda.value = valor
            except:
                pass

        print("✓ Valores escritos")

    def guardar(self, ruta_salida=None):
        """Guarda el formulario"""
        if not ruta_salida:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ruta_salida = f"Formulario_210_{timestamp}.xlsx"

        try:
            self.wb.save(ruta_salida)
            print(f"✓ Guardado: {ruta_salida}")
            return ruta_salida
        except Exception as e:
            print(f"✗ Error al guardar: {e}")
            return None

    def procesar_completo(self, datos_liquidacion):
        """Proceso completo"""
        print("\n" + "=" * 60)
        print("FORMULARIO 210 - LLENADO AUTOMÁTICO")
        print("=" * 60)

        try:
            if not self.abrir_formulario():
                return None

            self.llenar_datos_completos(datos_liquidacion)
            ruta = self.guardar()

            if ruta:
                print("\n✅ COMPLETADO")
                return ruta
            else:
                return None

        except Exception as e:
            print(f"❌ Error: {e}")
            return None
