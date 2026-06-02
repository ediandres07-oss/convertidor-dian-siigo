#!/usr/bin/env python3
"""
Script para convertir datos de la DIAN (Bimestre) al formato Siigo Nube (Planos).
Genera 3 hojas: COMPRAS, NC COMPRAS, GASTOS
"""

import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os

# ========================================
# CONFIGURACIÓN
# ========================================
INPUT_FILE = "BIMESTRE 1 YIMMY-1.xlsx"
OUTPUT_FILE = "PLANOS BIMESTRE 1 YIMMY.xlsx"

# Consecutivos iniciales (se asignan secuencialmente)
COMPRAS_CONSEC_INICIO = 371  # Siguiente al último del plano anterior (370)
NC_COMPRAS_CONSEC_INICIO = 271  # Siguiente al último del plano anterior (270)
GASTOS_CONSEC_INICIO = 798  # Siguiente al último del plano anterior (797 aprox)

# ========================================
# MAPEO DE CONCEPTO -> CUENTA CONTABLE DE GASTO
# ========================================
# Basado en análisis del plano de referencia
CONCEPTO_CUENTA_MAP = {
    "PIEZAS VEHICULOS": "52454001",
    "COMBUSTIBLE": "52953501",
    "PROCESAMIENTO DE DATOS": "52352001",
    "SEGUROS": "52304001",
    "FLETES": "52953001",
    "GASTOS LEGALES": "52509501",
    "ADECUACIONES": "52055101",
    "UTILES": "52355001",
}

# Excepciones por NIT (algunos NITs tienen cuenta diferente a la del concepto)
NIT_CUENTA_OVERRIDE = {
    # PIEZAS VEHICULOS - mecánicos/servicios van a 52509501
    "70051912": "52509501",   # MIRANDA JARAMILLO ALBERTO - Mecánico
    # COMBUSTIBLE - conductores específicos
    "70162940": "52956501",   # OSCAR ALONSO CANO MARIN
    # PROCESAMIENTO DE DATOS
    "800153993": "52353501",  # COMCEL - Telecomunicaciones
    "805018674": "52352001",  # ECOM S A S
    "830059699": "52352001",  # Satrack
    "890903407": "52304001",  # SEGUROS GENERALES SURAMERICANA (es seguro aunque dice proc datos)
    # Servicios generales (concepto None con NIT 811036875)
    "811036875": "53052001",  # SERVICIOS GENERALES SURAMERICA
}

# ========================================
# ENCABEZADOS DEL PLANO SIIGO
# ========================================
HEADERS = [
    'Tipo de comprobante', 'Consecutivo comprobante', 'Fecha de elaboración ',
    'Sigla moneda', 'Tasa de cambio', 'Código cuenta contable',
    'Identificación tercero', 'Sucursal', 'Código producto',
    'Código de bodega', 'Acción', 'Cantidad producto',
    'Prefijo', 'Consecutivo', 'No. cuota',
    'Fecha vencimiento', 'Código impuesto', 'Código grupo activo fijo',
    'Código activo fijo', 'Descripción', 'Código centro/subcentro de costos',
    'Débito', 'Crédito', 'Observaciones',
    'Base gravable libro compras/ventas  ', 'Base exenta libro compras/ventas',
    'Mes de cierre'
]


def make_row(tipo_comp, consec, fecha, cuenta, nit, cod_imp, descripcion, debito, credito):
    """Crea una fila en formato Siigo"""
    return [
        tipo_comp,  # Tipo de comprobante
        consec,     # Consecutivo
        fecha,      # Fecha
        None, None, # Moneda, Tasa
        cuenta,     # Cuenta contable
        str(nit) if nit else None,   # NIT tercero
        None, None, None, None, None,  # Sucursal, CodProd, Bodega, Acción, Cantidad
        None, None, None, None,  # Prefijo, Consecutivo, NoCuota, FechaVenc
        cod_imp,    # Código impuesto
        None, None, # CodGrupoAF, CodAF
        descripcion,# Descripción
        None,       # Centro de costos
        debito,     # Débito
        credito,    # Crédito
        None, None, None, None  # Observaciones, BaseGravable, BaseExenta, MesCierre
    ]


def get_cuenta_gasto(concepto, nit):
    """Determina la cuenta contable de gasto basándose en el concepto y NIT"""
    nit_str = str(nit) if nit else ""
    
    # Primero verificar si hay override por NIT
    if nit_str in NIT_CUENTA_OVERRIDE:
        return NIT_CUENTA_OVERRIDE[nit_str]
    
    # Luego por concepto
    if concepto and concepto in CONCEPTO_CUENTA_MAP:
        return CONCEPTO_CUENTA_MAP[concepto]
    
    # Si no hay concepto (None), asignar cuenta genérica
    if concepto is None:
        return "52454001"  # Por defecto para documentos sin concepto
    
    return "52454001"  # Fallback


def convert_compras(wb_src, wb_dst, consec_inicio):
    """Convierte la hoja COMPRAS del bimestre al formato Siigo"""
    ws_src = wb_src['COMPRAS']
    ws_dst = wb_dst.create_sheet('COMPRAS')
    
    # Escribir encabezados
    for col, header in enumerate(HEADERS, 1):
        ws_dst.cell(row=1, column=col, value=header)
    
    # Leer datos del bimestre (empezando en fila 3, fila 2 es encabezado)
    facturas = []
    for row_idx in range(3, ws_src.max_row + 1):
        folio = ws_src.cell(row=row_idx, column=3).value
        if folio is None:
            continue  # Saltar filas vacías o de totales
        
        fecha = ws_src.cell(row=row_idx, column=5).value
        nit = ws_src.cell(row=row_idx, column=6).value
        gravado = ws_src.cell(row=row_idx, column=10).value or 0
        iva = ws_src.cell(row=row_idx, column=11).value or 0
        total = ws_src.cell(row=row_idx, column=12).value or 0
        
        facturas.append({
            'folio': folio,
            'fecha': fecha,
            'nit': nit,
            'gravado': gravado,
            'iva': iva,
            'total': total,
        })
    
    # Generar filas del plano: 4 bloques (14350103, 14350101, 24081001, 22050501)
    row_out = 2
    consec = consec_inicio
    
    # Bloque 1: cuenta 14350103 (inventario - siempre 0,0)
    for f in facturas:
        row_data = make_row(9, consec, f['fecha'], '14350103', f['nit'], None,
                           f"FC: {f['folio']}", 0, 0)
        for col, val in enumerate(row_data, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1
    
    # Bloque 2: cuenta 14350101 (inventario - débito = gravado)
    consec = consec_inicio
    for f in facturas:
        row_data = make_row(9, consec, f['fecha'], '14350101', f['nit'], None,
                           f"FC: {f['folio']}", f['gravado'], 0)
        for col, val in enumerate(row_data, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1
    
    # Bloque 3: cuenta 24081001 (IVA descontable - débito = iva, codImpuesto=1)
    consec = consec_inicio
    for f in facturas:
        row_data = make_row(9, consec, f['fecha'], '24081001', f['nit'], 1,
                           f"FC: {f['folio']}", f['iva'], 0)
        for col, val in enumerate(row_data, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1
    
    # Bloque 4: cuenta 22050501 (proveedor - crédito = total)
    consec = consec_inicio
    for f in facturas:
        row_data = make_row(9, consec, f['fecha'], '22050501', f['nit'], None,
                           f"FC: {f['folio']}", 0, f['total'])
        for col, val in enumerate(row_data, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1
    
    print(f"  COMPRAS: {len(facturas)} facturas -> {(row_out - 2)} filas generadas")
    return consec_inicio + len(facturas)


def convert_nc_compras(wb_src, wb_dst, consec_inicio):
    """Convierte la hoja NC COMPRAS del bimestre al formato Siigo"""
    ws_src = wb_src['NC COMPRAS']
    ws_dst = wb_dst.create_sheet('NC COMPRAS')
    
    # Escribir encabezados
    for col, header in enumerate(HEADERS, 1):
        ws_dst.cell(row=1, column=col, value=header)
    
    # Leer datos
    notas = []
    for row_idx in range(3, ws_src.max_row + 1):
        folio = ws_src.cell(row=row_idx, column=3).value
        if folio is None:
            continue
        
        fecha = ws_src.cell(row=row_idx, column=5).value
        nit = ws_src.cell(row=row_idx, column=6).value
        no_gravado = ws_src.cell(row=row_idx, column=10).value or 0
        gravado = ws_src.cell(row=row_idx, column=11).value or 0
        iva = ws_src.cell(row=row_idx, column=12).value or 0
        total = ws_src.cell(row=row_idx, column=13).value or 0
        
        notas.append({
            'folio': folio,
            'fecha': fecha,
            'nit': nit,
            'no_gravado': no_gravado,
            'gravado': gravado,
            'iva': iva,
            'total': total,
        })
    
    # Generar 4 bloques: 14350103, 14350101, 24081002, 22050501
    row_out = 2
    consec = consec_inicio
    
    # Bloque 1: cuenta 14350103 - Crédito = NO GRAVADO (si hay NO GRAVADO sin GRAVADO)
    # En NC, cuando hay NO GRAVADO, va en 14350103 como crédito
    for n in notas:
        credito_14350103 = n['no_gravado'] if n['no_gravado'] > 0 and n['gravado'] == 0 else 0
        row_data = make_row(10, consec, n['fecha'], '14350103', n['nit'], None,
                           f"NC: {n['folio']}", 0, credito_14350103)
        for col, val in enumerate(row_data, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1
    
    # Bloque 2: cuenta 14350101 - Crédito = GRAVADO
    consec = consec_inicio
    for n in notas:
        row_data = make_row(10, consec, n['fecha'], '14350101', n['nit'], None,
                           f"NC: {n['folio']}", 0, n['gravado'])
        for col, val in enumerate(row_data, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1
    
    # Bloque 3: cuenta 24081002 (IVA NC - crédito = IVA, codImpuesto=1)
    consec = consec_inicio
    for n in notas:
        row_data = make_row(10, consec, n['fecha'], '24081002', n['nit'], 1,
                           f"NC: {n['folio']}", 0, n['iva'])
        for col, val in enumerate(row_data, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1
    
    # Bloque 4: cuenta 22050501 (proveedor - débito = total)
    consec = consec_inicio
    for n in notas:
        row_data = make_row(10, consec, n['fecha'], '22050501', n['nit'], None,
                           f"NC: {n['folio']}", n['total'], 0)
        for col, val in enumerate(row_data, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1
    
    print(f"  NC COMPRAS: {len(notas)} notas crédito -> {(row_out - 2)} filas generadas")
    return consec_inicio + len(notas)


def convert_gastos(wb_src, wb_dst, consec_inicio):
    """Convierte la hoja GASTOS del bimestre al formato Siigo"""
    ws_src = wb_src['GASTOS']
    ws_dst = wb_dst.create_sheet('GASTOS')
    
    # Escribir encabezados
    for col, header in enumerate(HEADERS, 1):
        ws_dst.cell(row=1, column=col, value=header)
    
    # Leer datos
    gastos = []
    for row_idx in range(3, ws_src.max_row + 1):
        folio = ws_src.cell(row=row_idx, column=3).value
        if folio is None:
            continue
        
        tipo_doc = ws_src.cell(row=row_idx, column=1).value
        fecha = ws_src.cell(row=row_idx, column=5).value
        nit = ws_src.cell(row=row_idx, column=6).value
        concepto = ws_src.cell(row=row_idx, column=8).value
        cod_iva_tipo = ws_src.cell(row=row_idx, column=9).value  # 24081001-1 o 24081501
        no_gravado = ws_src.cell(row=row_idx, column=12).value or 0
        base_gastos = ws_src.cell(row=row_idx, column=13).value or 0
        base_servicios = ws_src.cell(row=row_idx, column=14).value or 0
        iva_gastos = ws_src.cell(row=row_idx, column=15).value or 0
        iva_servicios = ws_src.cell(row=row_idx, column=16).value or 0
        total = ws_src.cell(row=row_idx, column=17).value or 0
        
        # Asegurarse de que los valores sean numéricos
        try:
            no_gravado = float(no_gravado) if no_gravado else 0
            base_gastos = float(base_gastos) if base_gastos else 0
            base_servicios = float(base_servicios) if base_servicios else 0
            iva_gastos = float(iva_gastos) if iva_gastos else 0
            iva_servicios = float(iva_servicios) if iva_servicios else 0
            total = float(total) if total else 0
        except (ValueError, TypeError):
            continue
        
        gastos.append({
            'folio': folio,
            'tipo_doc': tipo_doc,
            'fecha': fecha,
            'nit': nit,
            'concepto': concepto,
            'cod_iva_tipo': str(cod_iva_tipo) if cod_iva_tipo else None,
            'no_gravado': no_gravado,
            'base_gastos': base_gastos,
            'base_servicios': base_servicios,
            'iva_gastos': iva_gastos,
            'iva_servicios': iva_servicios,
            'total': total,
        })
    
    row_out = 2
    consec = consec_inicio
    
    # Bloque 1: Cuentas de gasto (5xxxxx) - Débito = base_gastos + base_servicios + no_gravado
    for g in gastos:
        cuenta = get_cuenta_gasto(g['concepto'], g['nit'])
        # El débito del gasto es la base gravable (gastos o servicios) + no gravado
        debito_gasto = int(g['base_gastos'] + g['base_servicios'] + g['no_gravado'])
        
        row_data = make_row(12, consec, g['fecha'], cuenta, g['nit'], None,
                           f"G: {g['folio']}", debito_gasto, 0)
        for col, val in enumerate(row_data, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1
    
    # Bloque 2: IVA (24081001 o 24081501) - Débito = IVA
    consec = consec_inicio
    for g in gastos:
        iva_total = int(g['iva_gastos'] + g['iva_servicios'])
        
        # Determinar cuenta IVA basándose en el tipo
        if g['cod_iva_tipo'] == '24081001-1':
            cuenta_iva = '24081001'
            cod_imp = 1
        elif g['cod_iva_tipo'] == '24081501':
            cuenta_iva = '24081501'
            cod_imp = None  # En el plano los 24081501 no tienen codImpuesto
        else:
            # Si tiene IVA gastos, usar 24081001
            if g['iva_gastos'] > 0:
                cuenta_iva = '24081001'
                cod_imp = 1
            elif g['iva_servicios'] > 0:
                cuenta_iva = '24081501'
                cod_imp = None
            else:
                cuenta_iva = '24081001'
                cod_imp = 1
        
        row_data = make_row(12, consec, g['fecha'], cuenta_iva, g['nit'], cod_imp,
                           f"G: {g['folio']}", iva_total, 0)
        for col, val in enumerate(row_data, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1
    
    # Bloque 3: Proveedor (23359501) - Crédito = Total
    consec = consec_inicio
    for g in gastos:
        row_data = make_row(12, consec, g['fecha'], '23359501', g['nit'], None,
                           f"G: {g['folio']}", 0, int(g['total']))
        for col, val in enumerate(row_data, 1):
            ws_dst.cell(row=row_out, column=col, value=val)
        row_out += 1
        consec += 1
    
    print(f"  GASTOS: {len(gastos)} gastos -> {(row_out - 2)} filas generadas")
    return consec_inicio + len(gastos)


def main():
    print(f"Leyendo archivo: {INPUT_FILE}")
    wb_src = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    
    print(f"Hojas encontradas: {wb_src.sheetnames}")
    
    wb_dst = Workbook()
    # Eliminar la hoja por defecto
    wb_dst.remove(wb_dst.active)
    
    print("\nConvirtiendo...")
    
    # Convertir COMPRAS
    last_consec = convert_compras(wb_src, wb_dst, COMPRAS_CONSEC_INICIO)
    print(f"    Último consecutivo COMPRAS: {last_consec - 1}")
    
    # Convertir NC COMPRAS
    last_consec = convert_nc_compras(wb_src, wb_dst, NC_COMPRAS_CONSEC_INICIO)
    print(f"    Último consecutivo NC COMPRAS: {last_consec - 1}")
    
    # Convertir GASTOS
    last_consec = convert_gastos(wb_src, wb_dst, GASTOS_CONSEC_INICIO)
    print(f"    Último consecutivo GASTOS: {last_consec - 1}")
    
    # Guardar
    wb_dst.save(OUTPUT_FILE)
    print(f"\n✅ Archivo generado: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
